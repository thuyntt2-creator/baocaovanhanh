import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v18.xlsx"

if not os.path.exists(excel_path):
    print("File không tồn tại")
    sys.exit(1)

# Đọc file không dùng data_only để xem công thức
wb = openpyxl.load_workbook(excel_path, data_only=False)

if '0.3 Bưu cục Detail' in wb.sheetnames:
    sheet = wb['0.3 Bưu cục Detail']
    print(f"=== ĐỌC CÔNG THỨC SHEET 0.3 Bưu cục Detail trong {os.path.basename(excel_path)} ===")
    # Đọc thử một số ô xung quanh hàng 71 hoặc tìm các ô chứa Nha Trang
    for r in range(1, 100):
        row_vals = [sheet.cell(r, c).value for c in range(1, 15)]
        # Kiểm tra xem dòng có chứa Nha Trang hoặc tổng
        row_str = str(row_vals).lower()
        if 'nha trang' in row_str or 'tổng' in row_str or 'tong' in row_str or 'chi phí' in row_str:
            print(f"Row {r:02d}: {[str(v) if v is not None else '' for v in row_vals]}")
else:
    print("Không tìm thấy sheet 0.3 Bưu cục Detail")
