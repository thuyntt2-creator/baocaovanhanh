import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"C:\Users\lap4all\Downloads\NTB_ Kế hoạch tách bưu cục Cam Linh và Bắc Cam Ranh-2026.xlsx"

if not os.path.exists(excel_path):
    print("File not found at exact path. Searching Downloads folder...")
    dl = r"C:\Users\lap4all\Downloads"
    for f in os.listdir(dl):
        if "cam linh" in f.lower() or "bắc cam ranh" in f.lower() or "tách" in f.lower():
            print("Candidate:", f)
            excel_path = os.path.join(dl, f)

print(f"\nReading Excel file: {excel_path}")
wb = openpyxl.load_workbook(excel_path, data_only=True)
print("Sheet names:", wb.sheetnames)

for sname in wb.sheetnames:
    print(f"\n--- SHEET: {sname} ---")
    sheet = wb[sname]
    for r in range(1, min(sheet.max_row + 1, 50)):
        row_vals = [sheet.cell(r, c).value for c in range(1, min(sheet.max_column + 1, 20))]
        if any(v is not None for v in row_vals):
            print(f"Row {r:2d}:", [str(v).strip() if v is not None else "" for v in row_vals])
