import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\V3 AOP_NTB_T70-T12_2026.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

print("=== Formulas in Mặt bằng sheet ===")
sheet_mb = wb['Mặt bằng']
for r in range(1, 10):
    for c in range(1, 10):
        val = sheet_mb.cell(r, c).value
        if isinstance(val, str) and val.startswith('='):
            col = openpyxl.utils.get_column_letter(c)
            print(f"  Cell {col}{r}: {val}")

print("\n=== Formulas in Nguồn lực & chi phí row 13 ===")
sheet_nlcp = wb['Nguồn lực & chi phí']
for c in range(1, 9):
    val = sheet_nlcp.cell(13, c).value
    col = openpyxl.utils.get_column_letter(c)
    print(f"  Cell {col}13: {val}")

wb.close()
