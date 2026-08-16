import docx, openpyxl, sys
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from docx.shared import Inches, Pt, RGBColor
from docx.enum.text import WD_ALIGN_PARAGRAPH
from docx.enum.table import WD_TABLE_ALIGNMENT
from docx.oxml import OxmlElement, parse_xml
from docx.oxml.ns import nsdecls, qn

sys.stdout.reconfigure(encoding='utf-8')

# Read original docx
doc_src = docx.Document(r'C:\Users\lap4all\Downloads\Quy_Hoach MANG LUOI NTB.docx')

parsed_36 = []
current = None

for p in doc_src.paragraphs:
    txt = p.text.strip()
    if not txt: continue
    
    is_heading = False
    for n in range(1, 37):
        if txt.startswith(f'{n}. '):
            is_heading = True
            break
            
    if is_heading:
        if current: parsed_36.append(current)
        current = {'title': txt, 'communes': '', 'proposal': '', 'reason': ''}
    elif current:
        if 'Các xã cũ sáp nhập' in txt:
            current['communes'] = txt.replace('❖ Các xã cũ sáp nhập & Sản lượng Giao/Lấy từng xã:', '').strip()
        elif 'ĐỀ XUẤT PHƯƠNG ÁN' in txt:
            current['proposal'] = txt.replace('❖ ĐỀ XUẤT PHƯƠNG ÁN CỦA AM:', '').replace('❖ ĐỀ XUẤT PHƯƠNG ÁN:', '').strip()
        elif 'LÝ DO VÀ GIẢI THÍCH' in txt:
            current['reason'] = txt.replace('❖ LÝ DO VÀ GIẢI THÍCH NGUYÊN NHÂN CHI TIẾT TỪ AM:', '').replace('❖ LÝ DO VÀ GIẢI THÍCH NGUYÊN NHÂN CHI TIẾT:', '').strip()
        elif current['communes'] and not current['proposal'] and not txt.startswith('❖'):
            current['communes'] += ' ' + txt
        elif current['proposal'] and not current['reason'] and not txt.startswith('❖'):
            current['proposal'] += ' ' + txt
        elif current['reason'] and not txt.startswith('❖'):
            current['reason'] += ' ' + txt

if current: parsed_36.append(current)

print(f"Total parsed 36 items: {len(parsed_36)}")

# Create Excel File with exact 36 items un-truncated
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "36_DVHC_Tu_Docx_Goc"

font_header = Font(name="Calibri", size=11, bold=True, color="000000")
fill_header = PatternFill(start_color="F7A059", end_color="F7A059", fill_type="solid")
font_dvhc = Font(name="Calibri", size=10, bold=True, color="000000")
font_normal = Font(name="Calibri", size=9.5, color="000000")
font_buucuc = Font(name="Calibri", size=10, bold=True, color="003399")

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
)

headers = ["STT & Tên ĐVHC Mới", "Các Xã / Phường Cũ sáp nhập & Sản lượng Giao/Lấy", "Bưu cục Cover (Đề xuất AM)", "Phương án Quy hoạch & Lý do chi tiết từ AM"]

for col_idx, h_text in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=h_text)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

ws.row_dimensions[1].height = 28

for r_idx, item in enumerate(parsed_36, start=2):
    c1 = ws.cell(row=r_idx, column=1, value=item['title'])
    c2 = ws.cell(row=r_idx, column=2, value=item['communes'])
    c3 = ws.cell(row=r_idx, column=3, value=item['proposal'])
    c4 = ws.cell(row=r_idx, column=4, value=item['reason'] if item['reason'] else item['proposal'])
    
    c1.font = font_dvhc
    c2.font = font_normal
    c3.font = font_buucuc
    c4.font = font_normal
    
    for col_idx in range(1, 5):
        ws.cell(row=r_idx, column=col_idx).border = thin_border
        ws.cell(row=r_idx, column=col_idx).alignment = align_left

ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 38
ws.column_dimensions['D'].width = 55

excel_out = r'C:\Users\lap4all\Downloads\Quy_Hoach_NTB_2026_Full_36_Khoan_Chinh_Xac_Docx.xlsx'
wb.save(excel_out)
print(f"Exported Excel: {excel_out}")
