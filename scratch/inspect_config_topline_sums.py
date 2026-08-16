import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

config_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB.xlsx"
wb = openpyxl.load_workbook(config_path, data_only=True)

for sheet_name in ['Topline_Giao_NTB', 'Topline_Lay_NTB']:
    sheet = wb[sheet_name]
    print(f"\n=== Summing {sheet_name} by Province, Band, Month ===")
    
    sums = {}
    for r in range(2, sheet.max_row + 1):
        month_val = sheet.cell(r, 1).value # e.g. '2026-07-01' or a date object
        prov = sheet.cell(r, 5).value
        band = sheet.cell(r, 4).value
        vol = sheet.cell(r, 6).value
        
        if month_val is None or prov is None or band is None:
            continue
            
        # extract month string like 'T7'
        if isinstance(month_val, str):
            month_num = int(month_val.split('-')[1])
        else:
            month_num = month_val.month
        month_str = f"T{month_num}"
        
        key = (prov.strip(), band.strip())
        sums.setdefault(key, {})
        sums[key].setdefault(month_str, 0.0)
        if vol is not None:
            sums[key][month_str] += float(vol)
            
    # Print sorted
    for key in sorted(sums.keys()):
        m_vals = [round(sums[key].get(m, 0.0), 1) for m in ['T7', 'T8', 'T9', 'T10', 'T11', 'T12']]
        print(f"  {key}: {m_vals}")

print("\n" + "="*50)
v2_path = r"C:\Users\lap4all\Downloads\[V2] AOP_Hang_NTB_T7-T12_2026.xlsx"
wb_v2 = openpyxl.load_workbook(v2_path, data_only=True)
print("Sheets in [V2]:", wb_v2.sheetnames)

