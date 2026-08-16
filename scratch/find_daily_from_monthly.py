import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

config_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB (1).xlsx"
if not os.path.exists(config_path):
    config_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB.xlsx"

wb_cfg = openpyxl.load_workbook(config_path, data_only=True)

# 1. Sum up monthly volumes for target groups (BCCKs) in 3_KQ_BC_Detail
# Target groups:
# Nha Trang: any BC in Nha Trang (we can find from 1_Config_Chuan where Quận_Tên == 'Thành phố Nha Trang')
# Đơn Dương, Di Linh, Đức Linh: from their names

# Let's get BC IDs for Nha Trang from 1_Config_Chuan
sheet_chuan = wb_cfg['1_Config_Chuan']
chuan_rows = list(sheet_chuan.iter_rows(values_only=True))
chuan_headers = chuan_rows[0]
h_idx = {h: i for i, h in enumerate(chuan_headers)}

nt_giao_bc_ids = set()
for row in chuan_rows[1:]:
    if row[h_idx['Quận_Tên']] == 'Thành phố Nha Trang' and row[h_idx['BC_Giao_ID']]:
        nt_giao_bc_ids.add(row[h_idx['BC_Giao_ID']])

print("Nha Trang Giao BC IDs in Config:", nt_giao_bc_ids)

# Let's sum monthly volumes for Nha Trang, Đơn Dương, Di Linh, Đức Linh
sheet_3 = wb_cfg['3_KQ_BC_Detail']
totals = {
    'Nha Trang': {f'T{m}': 0.0 for m in range(7, 13)},
    'Đơn Dương': {f'T{m}': 0.0 for m in range(7, 13)},
    'Di Linh': {f'T{m}': 0.0 for m in range(7, 13)},
    'Đức Linh': {f'T{m}': 0.0 for m in range(7, 13)}
}

bulky_groups = {'10-15kg', '15-30kg', '≥30kg'}

for r in range(2, sheet_3.max_row + 1):
    bc_id = sheet_3.cell(r, 2).value
    bc_name = str(sheet_3.cell(r, 3).value or '')
    weight_group = sheet_3.cell(r, 5).value
    
    if weight_group in bulky_groups:
        # Determine which group
        group = None
        if bc_id in nt_giao_bc_ids:
            group = 'Nha Trang'
        elif 'Đơn Dương' in bc_name:
            group = 'Đơn Dương'
        elif 'Di Linh' in bc_name:
            group = 'Di Linh'
        elif 'Đức Linh' in bc_name:
            group = 'Đức Linh'
            
        if group:
            for m in range(7, 13):
                col_idx = 5 + (m - 6) # T7 is col 6, T8 is col 7, etc.
                val = sheet_3.cell(r, col_idx).value
                totals[group][f'T{m}'] += float(val) if val is not None else 0.0

print("\n=== Monthly Sums of ≥10kg from 3_KQ_BC_Detail ===")
for g, m_vals in totals.items():
    print(f"{g}:")
    for m in range(7, 13):
        print(f"  T{m}: {m_vals[f'T{m}']:.2f}")

# Now let's compare these with the sums of daily Forecast in v18
v18_path = r"C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v18.xlsx"
wb_v18 = openpyxl.load_workbook(v18_path, data_only=True)

print("\n=== Sums of daily Forecast in v18 ===")
for m in range(7, 13):
    sheet_name = f"Forecast T{m}"
    sheet = wb_v18[sheet_name]
    print(f"Month T{m}:")
    # Rows 4 to 7 are Nha Trang, Đơn Dương, Di Linh, Đức Linh
    for r in range(4, 8):
        name = sheet.cell(r, 1).value
        # sum of row values starting from col B to last non-empty col
        vols = []
        for c in range(2, sheet.max_column + 1):
            val = sheet.cell(r, c).value
            if val is not None:
                vols.append(val)
        print(f"  {name}: Daily Sum = {sum(vols)}")
