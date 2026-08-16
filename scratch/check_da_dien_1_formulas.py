import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

filepath = r"C:\Users\lap4all\Downloads\NTB_Input_Da_Dien_FLM_CRC_1.xlsx"
if os.path.exists(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=False)
    sheet = wb['NTB – Input']
    print("=== NTB_Input_Da_Dien_FLM_CRC_1.xlsx - NTB – Input rows 41-45 formulas ===")
    cols = ['D', 'E', 'F', 'G', 'H', 'I']
    for r in range(41, 46):
        label = sheet.cell(r, 2).value or sheet.cell(r, 1).value
        print(f"Row {r:2d} | {str(label)[:35]:<35}")
        for c in cols:
            print(f"  {c}: {sheet[f'{c}{r}'].value}")
    wb.close()
else:
    print("NTB_Input_Da_Dien_FLM_CRC_1.xlsx not found")
