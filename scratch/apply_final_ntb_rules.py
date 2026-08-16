import openpyxl
import win32com.client
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

downloads_dir = r"C:\Users\lap4all\Downloads"
files_to_fix = [
    os.path.join(downloads_dir, "NTB_Input_Da_Dien_FLM_CRC_2.xlsx"),
    os.path.join(downloads_dir, "NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx")
]

cols = ['D', 'E', 'F', 'G', 'H', 'I']

# Source of truth for new volumes
# config_psbba_NTB new.xlsx volumes:
# Giao Band 3: 77765, 93821, 100952, 115910, 123254, 132076
# Giao Band 4: 39757, 48820, 53078, 61451, 65146, 69431
# Giao Band 5: 71073, 87538, 95594, 111084, 117527, 124892
# Lay Band 3: 12394, 13953, 14591, 16405, 17767, 19463
# Lay Band 4: 5672, 6477, 6743, 7536, 8110, 8857
# Lay Band 5: 9870, 11351, 11893, 13363, 14336, 15592

new_volumes = {
    8: [77765, 93821, 100952, 115910, 123254, 132076],
    9: [39757, 48820, 53078, 61451, 65146, 69431],
    10: [71073, 87538, 95594, 111084, 117527, 124892],
    11: [12394, 13953, 14591, 16405, 17767, 19463],
    12: [5672, 6477, 6743, 7536, 8110, 8857],
    13: [9870, 11351, 11893, 13363, 14336, 15592]
}

# Number of vehicles based on recalculation with new volumes:
# T7: 39, T8: 43, T9: 45, T10: 50, T11: 51, T12: 57
vehicles = [39, 43, 45, 50, 51, 57]

for filepath in files_to_fix:
    if not os.path.exists(filepath):
        continue
        
    print(f"\nApplying crew salary and new volumes to: {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath, data_only=False)
    sheet = wb['NTB – Input']
    
    # 1. Update volumes (Row 8-13)
    for row_idx, vals in new_volumes.items():
        for col_idx, col in enumerate(cols):
            sheet[f"{col}{row_idx}"] = vals[col_idx]
            
    # 2. Update Row 7 (TỔNG đơn)
    for col in cols:
        sheet[f"{col}7"] = f"=SUM({col}8:{col}10)"
        
    # 3. Update Row 25 (Số xe cần có)
    for col_idx, col in enumerate(cols):
        sheet[f"{col}25"] = vehicles[col_idx]
        
    # 4. Update Row 19 (Số NV theo xe = Row 25 * 2)
    for col in cols:
        sheet[f"{col}19"] = f"={col}25*2"
        
    # 5. Update Row 20 (Lương NV theo xe = 15 million)
    for col in cols:
        sheet[f"{col}20"] = 15
        
    # 6. Update Row 30 (A. Thuê xe) = D25 * D28 * D6 / 1000000
    for col in cols:
        sheet[f"{col}30"] = f"={col}25*{col}28*{col}6/1000000"
        
    # 7. Update Row 31 (B1. Chi phí NV kênh nhẹ) = D15 * D18
    for col in cols:
        sheet[f"{col}31"] = f"={col}15*{col}18"
        
    # 8. Update Row 32 (B2. Chi phí NV theo xe) = D19 * D20
    for col in cols:
        sheet[f"{col}32"] = f"={col}19*{col}20"
        
    # 9. Update Row 33 (B3. Chi phí NV xử lý/kho&QL) = D21 * D23
    for col in cols:
        sheet[f"{col}33"] = f"={col}21*{col}23"
        
    wb.save(filepath)
    wb.close()
    
    # Recalculate via Excel COM
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        workbook = excel.Workbooks.Open(filepath)
        workbook.Save()
        workbook.Close()
        print("Success!")
    except Exception as e:
        print(f"Error during recalculation: {e}")
    finally:
        excel.Quit()

# Verify values in both files
for filepath in files_to_fix:
    if not os.path.exists(filepath):
        continue
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb['NTB – Input']
    print(f"\n=== VERIFIED VALUES IN: {os.path.basename(filepath)} ===")
    for r in [7, 8, 9, 19, 20, 25, 30, 31, 32, 33, 38, 39, 41, 42, 43, 44, 45]:
        label = sheet.cell(r, 2).value or sheet.cell(r, 1).value
        vals = [sheet.cell(r, openpyxl.utils.column_index_from_string(c)).value for c in cols]
        print(f"Row {r:2d} | {str(label)[:35]:<35} | {vals}")
    wb.close()
