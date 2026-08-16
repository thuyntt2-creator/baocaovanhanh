import openpyxl
import sys
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB new.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

# Check Topline_Giao_NTB for weight bands >= 10kg
print("=== Topline_Giao_NTB - Unique rank_weight values ===")
sheet = wb['Topline_Giao_NTB']
rank_weights = set()
for r in range(2, sheet.max_row + 1):
    rw = sheet.cell(r, 4).value
    if rw:
        rank_weights.add(rw)
print(sorted(rank_weights))

# Sum GIAO volumes by rank_weight and month for relevant bands (>=10kg)
print("\n=== GIAO Volume by rank_weight >= 10kg ===")
months_map = {'T07': 7.0, 'T08': 8.0, 'T09': 9.0, 'T10': 10.0, 'T11': 11.0, 'T12': 12.0}
giao_by_band_month = defaultdict(lambda: defaultdict(float))

for r in range(2, sheet.max_row + 1):
    rw = sheet.cell(r, 4).value  # rank_weight
    vol = sheet.cell(r, 6).value or 0  # volume
    month = sheet.cell(r, 9).value   # THÁNG
    if rw and month and any(s in str(rw) for s in ['10', '15', '20', '30', 'gte', 'ge']):
        giao_by_band_month[rw][int(month)] += vol

for band in sorted(giao_by_band_month.keys()):
    months_vals = [f"T{m}:{giao_by_band_month[band][m]:,.0f}" for m in [7,8,9,10,11,12]]
    print(f"  {band}: {' | '.join(months_vals)}")

# Same for Topline_Lay_NTB
print("\n\n=== Topline_Lay_NTB - Unique rank_weight values ===")
sheet_lay = wb['Topline_Lay_NTB']
rank_weights_lay = set()
for r in range(2, sheet_lay.max_row + 1):
    rw = sheet_lay.cell(r, 4).value
    if rw:
        rank_weights_lay.add(rw)
print(sorted(rank_weights_lay))

print("\n=== LAY Volume by rank_weight >= 10kg ===")
lay_by_band_month = defaultdict(lambda: defaultdict(float))
for r in range(2, sheet_lay.max_row + 1):
    rw = sheet_lay.cell(r, 4).value
    vol = sheet_lay.cell(r, 6).value or 0
    month = sheet_lay.cell(r, 9).value
    if rw and month and any(s in str(rw) for s in ['10', '15', '20', '30']):
        lay_by_band_month[rw][int(month)] += vol

for band in sorted(lay_by_band_month.keys()):
    months_vals = [f"T{m}:{lay_by_band_month[band][m]:,.0f}" for m in [7,8,9,10,11,12]]
    print(f"  {band}: {' | '.join(months_vals)}")

wb.close()
