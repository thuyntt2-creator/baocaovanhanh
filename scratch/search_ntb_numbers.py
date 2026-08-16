import openpyxl, os, sys
sys.stdout.reconfigure(encoding='utf-8')

target_files = [
    r'C:\Users\lap4all\Downloads\AOP_Hang_NTB_T7-T12_2026.xlsx',
    r'C:\Users\lap4all\Downloads\AOP_Hang_NTB_T7-T12_2026_fix.xlsx',
    r'C:\Users\lap4all\Downloads\AOP_Hang_NTB_T7-T12_2026_calculated.xlsx',
    r'C:\Users\lap4all\Downloads\AOP_NTB_calculated_v2.xlsx',
    r'C:\Users\lap4all\Downloads\AOP_NTB_v3.xlsx'
]

for f in target_files:
    if not os.path.exists(f):
        print(f'{f} does not exist.')
        continue
    print(f'Searching {os.path.basename(f)}...')
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        for s in wb.sheetnames:
            sheet = wb[s]
            for r in range(1, 100):
                for c in range(1, 50):
                    val = sheet.cell(r, c).value
                    if isinstance(val, (int, float)) and round(val) == 624:
                        print(f'  Found 624 at {s}!{sheet.cell(r, c).coordinate} -> {val}')
                    if isinstance(val, (int, float)) and round(val) == 211:
                        print(f'  Found 211 at {s}!{sheet.cell(r, c).coordinate} -> {val}')
    except Exception as e:
        print(f'  Error: {e}')
print('Search finished.')
