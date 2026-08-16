import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\V3_AOP_Hang_NTB_T7-T12_2026 mới.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

print("=== Forecast sums by BCCK and Month ===")
for m in range(7, 13):
    sheet_name = f"Forecast T{m}"
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\nMonth T{m}:")
        for r in range(4, 9):
            bc_name = sheet.cell(r, 1).value
            if bc_name is None:
                continue
            row_sum = sum(sheet.cell(r, c).value or 0.0 for c in range(2, sheet.max_column + 1))
            print(f"  {bc_name}: {row_sum:.1f}")

wb.close()
