import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

xlsx_aop_path = r"C:\Users\lap4all\Downloads\[V2] AOP_Hang_NTB_T7-T12_2026.xlsx"
wb = openpyxl.load_workbook(xlsx_aop_path, data_only=True)
sheet = wb['0.3 Bưu cục Detail']

print("=== Template 0.3 Bưu cục Detail rows ===")
for r in range(1, sheet.max_row + 1):
    bc_name = sheet.cell(r, 2).value
    stt = sheet.cell(r, 1).value
    if bc_name is not None or stt is not None:
        print(f"Row {r:2d}: STT={stt} | Name={bc_name}")

