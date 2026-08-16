import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\V3_AOP_Hang_NTB_T7-T12_2026 mới.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

if 'Bưu cục detai;' in wb.sheetnames:
    sheet = wb['Bưu cục detai;']
    print(f"=== Sheet: Bưu cục detai; (Rows 1 to 40) ===")
    for r in range(1, 41):
        row_vals = [sheet.cell(r, c).value for c in range(1, 14)]
        if any(v is not None for v in row_vals):
            print(f"Row {r:2d}: {row_vals}")
else:
    print("Sheet 'Bưu cục detai;' not found!")

wb.close()
