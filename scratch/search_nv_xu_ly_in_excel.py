import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v18.xlsx"

if not os.path.exists(excel_path):
    print("File không tồn tại")
    sys.exit(1)

wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)

print("=== TÌM KIẾM NHÂN SỰ KHO TRONG EXCEL V18 ===")
for sname in wb.sheetnames:
    if "nguồn lực" in sname.lower() or "chi phí" in sname.lower() or "kế hoạch" in sname.lower():
        sheet = wb[sname]
        print(f"\nSheet: {sname}")
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            row_str = " | ".join([str(cell) for cell in row if cell is not None])
            if "kho" in row_str.lower() or "xử lý" in row_str.lower() or "quản lý" in row_str.lower() or "ql" in row_str.lower():
                print(f"Row {r_idx+1}: {row_str[:200]}")
