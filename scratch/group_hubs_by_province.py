import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\lap4all\Documents\Auto report\NTB_Phan_Tuyen_Hanh_Chinh_Quy_Hoach_Chi_Tiet.xlsx"
if not os.path.exists(excel_path):
    excel_path = r"c:\Users\lap4all\Documents\Auto report\NTB_Phan_Tuyen_Hanh_Chinh_Quy_Hoach_Moi.xlsx"

wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb.active

print("Sheet title:", sheet.title)
headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column+1)]
print("Headers:", headers)

data_by_prov = {}
for r in range(2, sheet.max_row+1):
    row_vals = {headers[i]: sheet.cell(r, i+1).value for i in range(len(headers)) if headers[i]}
    prov = str(row_vals.get('Tỉnh', '') or row_vals.get('Tỉnh/TP', '')).strip()
    if prov:
        if prov not in data_by_prov:
            data_by_prov[prov] = []
        data_by_prov[prov].append(row_vals)

for prov, rows in data_by_prov.items():
    print(f"\n================ PROVINCE: {prov} ({len(rows)} rows) ================")
    for r in rows[:5]:
        print(r)
