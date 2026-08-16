import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\lap4all\Documents\Auto report\NTB_Phan_Tuyen_Hanh_Chinh_Quy_Hoach_Chi_Tiet.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Sheet1']

headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column+1)]

rows_by_xa_moi = {}

for r in range(2, sheet.max_row+1):
    row = {headers[i]: sheet.cell(r, i+1).value for i in range(len(headers)) if headers[i]}
    prov = str(row.get('Tỉnh, thành phố mới') or '').strip()
    xa_moi = str(row.get('Tên Xã mới') or '').strip()
    xa_cu = str(row.get('Tên Xã cũ') or '').strip()
    bc_moi = str(row.get('Tên Bưu cục giao mới đề xuất') or '').strip()
    dexuat = str(row.get('Đánh giá & Phương án đề xuất') or '').strip()
    lydo = str(row.get('Lý do & Bố trí nhân sự') or '').strip()
    am = str(row.get('Quản lý khu vực (AM)') or '').strip()
    
    if not prov or not xa_moi:
        continue
        
    if xa_moi not in rows_by_xa_moi:
        rows_by_xa_moi[xa_moi] = {
            'prov': prov,
            'xa_moi': xa_moi,
            'bc_moi': bc_moi,
            'dexuat': dexuat,
            'lydo': lydo,
            'am': am,
            'xa_cu': []
        }
    rows_by_xa_moi[xa_moi]['xa_cu'].append(xa_cu)

print(f"Total Xã mới analyzed: {len(rows_by_xa_moi)}")
for idx, (xa_moi, info) in enumerate(rows_by_xa_moi.items(), 1):
    print(f"{idx:2d}. [{info['prov']}] {xa_moi}")
    print(f"    -> BC đề xuất: {info['bc_moi']}")
    print(f"    -> Đề xuất: {info['dexuat']}")
    print(f"    -> Lý do: {info['lydo']}")
    print(f"    -> Cover xã cũ: {', '.join(info['xa_cu'])}")
    print()
