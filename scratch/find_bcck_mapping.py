import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

config_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB.xlsx"
wb = openpyxl.load_workbook(config_path, data_only=True)
sheet = wb['1_Config_Chuan']

districts = {}
for r in range(2, sheet.max_row + 1):
    dist_name = sheet.cell(r, 3).value
    bc_giao = sheet.cell(r, 9).value
    if dist_name and bc_giao:
        districts.setdefault(dist_name.strip(), set())
        districts[dist_name.strip()].add(bc_giao.strip())

print("=== District to BC Giao Mapping ===")
for dist, bcs in sorted(districts.items()):
    print(f"District: {dist}")
    for bc in sorted(bcs):
        print(f"  - {bc}")

