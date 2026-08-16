import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')
wb_cfg = openpyxl.load_workbook(r'C:\Users\lap4all\Downloads\config_psbba_NTB.xlsx', data_only=True)

sheet_cfg = wb_cfg['1_Config_Chuan']
ward_data = {}
bc_ward_count = {}
for r in range(2, sheet_cfg.max_row+1):
    ward_id = sheet_cfg.cell(r, 4).value # Phường_ID
    ward_name = sheet_cfg.cell(r, 5).value # Phường_Tên
    bc = sheet_cfg.cell(r, 9).value # BC_Giao_Tên
    grp = sheet_cfg.cell(r, 11).value # B2B_Nhóm
    if ward_id and bc and ward_name:
        ward_id = str(ward_id).strip()
        bc = str(bc).strip()
        ward_data[ward_id] = {'bc': bc, 'group': grp, 'name': str(ward_name).strip()}
        bc_ward_count[bc] = bc_ward_count.get(bc, 0) + 1

sheet_3 = wb_cfg['3_KQ_BC_Detail']
t7_col_3 = 6
bc_reg_vol = {}
for r in range(2, sheet_3.max_row+1):
    bc = sheet_3.cell(r, 3).value # routed_bc_name
    weight = str(sheet_3.cell(r, 5).value or '')
    if bc and weight in ['10-15kg', '15-30kg', '≥30kg']:
        vol = sheet_3.cell(r, t7_col_3).value or 0
        bc = str(bc).strip()
        bc_reg_vol[bc] = bc_reg_vol.get(bc, 0) + vol

sheet_4 = wb_cfg['4_KQ_B2B']
t7_col_4 = 12
ward_b2b_vol = {}
for r in range(2, sheet_4.max_row+1):
    ward_id = sheet_4.cell(r, 4).value # Phường_ID
    weight = str(sheet_4.cell(r, 11).value or '') # Weight_Band
    if ward_id and weight in ['15to30', 'gte30']:
        vol = sheet_4.cell(r, t7_col_4).value or 0
        ward_id = str(ward_id).strip()
        ward_b2b_vol[ward_id] = ward_b2b_vol.get(ward_id, 0) + vol

print("\n--- KẾT QUẢ TÍNH NHÁP (MỚI) ---")
for w_id, w_info in ward_data.items():
    if 'Di Linh' in w_info['name'] or 'Đinh Văn' in w_info['name'] or 'Hòa Ninh' in w_info['name'] or 'Bình Thạnh' in w_info['name']:
        bc = w_info['bc']
        grp = w_info['group']
        w_name = w_info['name']
        w_count = bc_ward_count.get(bc, 1)
        tot_bc_reg = bc_reg_vol.get(bc, 0)
        avg_reg = tot_bc_reg / w_count
        b2b_vol = ward_b2b_vol.get(w_id, 0)
        
        if grp in ['A', 'B']:
            final_vol = b2b_vol + avg_reg
        else:
            final_vol = avg_reg
            
        print(f"{w_name} (Nhóm {grp}): BC={bc} | T.Số Xã={w_count} | Tổng BC GTC={tot_bc_reg:.1f} => GTC TB={avg_reg:.1f} | B2B={b2b_vol:.1f} || Tổng = {final_vol:.1f} (Làm tròn: {round(final_vol)})")
