import openpyxl
import os
import glob
import sys

sys.stdout.reconfigure(encoding='utf-8')

downloads_dir = r"C:\Users\lap4all\Downloads"
files = glob.glob(os.path.join(downloads_dir, "AOP_Hang_Nang_*.xlsx"))

for fpath in files:
    if os.path.basename(fpath).startswith("~$"):
        continue
    print(f"\n--- Checking File: {os.path.basename(fpath)} ---")
    try:
        wb = openpyxl.load_workbook(fpath, data_only=False)
        for sname in wb.sheetnames:
            if "Forecast T" in sname:
                sheet = wb[sname]
                # Look at first few cells in row 4
                cells = [sheet.cell(4, c).value for c in range(1, 10)]
                has_formula = any(isinstance(c, str) and c.startswith("=") for c in cells)
                print(f"  Sheet {sname}: row 4 = {cells[:6]} (has formulas: {has_formula})")
    except Exception as e:
        print(f"  Error reading {os.path.basename(fpath)}: {e}")
