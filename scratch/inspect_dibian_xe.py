import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['Định biên & Sản lượng']

print("=== Định biên & Sản lượng Row 17 (Số xe cần/đơn vị) ===")
row_vals = [sheet.cell(17, c).value for c in range(1, 16)]
print(row_vals)

print("\n=== Định biên & Sản lượng Row 31 (Số NV kênh nhẹ) ===")
row_vals_31 = [sheet.cell(31, c).value for c in range(1, 16)]
print(row_vals_31)

print("\n=== Định biên & Sản lượng Row 33 (Số NV theo xe) ===")
row_vals_33 = [sheet.cell(33, c).value for c in range(1, 16)]
print(row_vals_33)

wb.close()
