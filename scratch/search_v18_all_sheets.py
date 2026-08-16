import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v18.xlsx"

if not os.path.exists(excel_path):
    print("File không tồn tại")
    sys.exit(1)

wb = openpyxl.load_workbook(excel_path, data_only=True)

print(f"=== Tìm kiếm trong file: {os.path.basename(excel_path)} ===")
targets = [715, 690, 1746, 126, 76]

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    found = False
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(r, c).value
            if val is not None:
                # Kiểm tra xem giá trị có khớp hoặc chứa các số mục tiêu không
                val_str = str(val).lower()
                for target in targets:
                    # Kiểm tra số chính xác hoặc dạng float
                    if (isinstance(val, (int, float)) and abs(val - target) < 1) or \
                       (isinstance(val, (int, float)) and abs(val/1e6 - target) < 1) or \
                       (isinstance(val, (int, float)) and abs(val/1e3 - target) < 1):
                        print(f"[{sheet_name}] Ô {openpyxl.utils.get_column_letter(c)}{r}: Giá trị = {val} (Khớp với target {target})")
                        found = True
                    elif str(target) in val_str:
                        print(f"[{sheet_name}] Ô {openpyxl.utils.get_column_letter(c)}{r}: Giá trị = {val} (Khớp chuỗi với target {target})")
                        found = True
