import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v18.xlsx"

if not os.path.exists(excel_path):
    print("File không tồn tại")
    sys.exit(1)

wb = openpyxl.load_workbook(excel_path, data_only=True)

# Các bưu cục cần trích xuất
bccs = ['BCCK Nha Trang', 'BCCK Di Linh', 'BCCK Đơn Dương', 'BCCK Đức Linh']
months = ['T7', 'T10', 'T12']

print("=== TRÍCH XUẤT SỐ XE VÀ CHUYẾN XE TỪ PLAN EXCEL V18 STANDARD ===")

for m in months:
    sheet_name = f"Kế hoạch {m}"
    if sheet_name not in wb.sheetnames:
        print(f"Không tìm thấy sheet: {sheet_name}")
        continue
    sheet = wb[sheet_name]
    print(f"\n--- {sheet_name} ---")
    
    # Quét toàn bộ sheet để tìm bưu cục và số xe
    # Thông thường cấu trúc mỗi bưu cục là một block
    # Chúng ta sẽ in ra các dòng chứa bưu cục và thông tin xe
    for r_idx in range(1, 100):
        row_vals = [sheet.cell(row=r_idx, column=c).value for c in range(1, 15)]
        row_str = " | ".join([str(v) for v in row_vals if v is not None])
        # Kiểm tra xem dòng này có chứa tên bưu cục nào không
        for bc in bccs:
            if bc.lower() in row_str.lower() or (bc.replace('BCCK ', '').lower() in row_str.lower()):
                # In ra block 15 dòng kể từ dòng này để xem chi tiết
                print(f"\n[KHỚP] Bưu cục {bc} tìm thấy tại dòng {r_idx}:")
                for offset in range(12):
                    line_vals = [sheet.cell(row=r_idx+offset, column=c).value for c in range(1, 15)]
                    line_str = " | ".join([str(v) for v in line_vals if v is not None])
                    print(f"  Row {r_idx+offset}: {line_str}")
                break
wb.close()
