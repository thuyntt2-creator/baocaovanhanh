import win32com.client
import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

draft_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC_updated_draft.xlsx"

# Force recalculation via Excel COM interface
print("Opening Excel to force recalculation...")
excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = False
excel.DisplayAlerts = False

try:
    wb = excel.Workbooks.Open(draft_path)
    wb.Save()
    wb.Close()
    print("Recalculation and save complete.")
except Exception as e:
    print(f"Error during Excel COM call: {e}")
finally:
    excel.Quit()

# Read the recalculated draft with openpyxl data_only=True
wb_op = openpyxl.load_workbook(draft_path, data_only=True)
sheet = wb_op['NTB – Input']

print("\n=== Evaluated values in NTB - Input ===")
cols = ['D', 'E', 'F', 'G', 'H', 'I']
months = ['T7', 'T8', 'T9', 'T10', 'T11', 'T12']

rows_to_check = [7, 8, 9, 10, 11, 12, 13, 15, 17, 20, 22, 25, 27, 30, 31, 32, 33, 34, 36, 37, 38, 39, 41, 42, 43, 44, 45]

for r in rows_to_check:
    row_label = sheet.cell(r, 2).value or sheet.cell(r, 1).value or f"Row {r}"
    vals = [sheet[f"{c}{r}"].value for c in cols]
    print(f"Row {r:2d} ({row_label}): {vals}")

wb_op.close()
