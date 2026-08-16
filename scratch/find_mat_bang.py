import openpyxl
import os
import glob
import sys

sys.stdout.reconfigure(encoding='utf-8')

downloads_dir = r"C:\Users\lap4all\Downloads"
search_pattern = os.path.join(downloads_dir, "*.xlsx")

print("Searching for 'Đức Linh' and 'Diện tích' in all xlsx files...")

for fpath in glob.glob(search_pattern):
    if os.path.basename(fpath).startswith("~$"):
        continue
    try:
        wb = openpyxl.load_workbook(fpath, read_only=True)
        for sname in wb.sheetnames:
            # Check if sheet contains some indicator of mat bang
            if any(k in sname.lower() for k in ['mặt bằng', 'phương án', 'mật độ', 'tiếp nhận']):
                print(f"\nFound sheet '{sname}' in file '{os.path.basename(fpath)}'")
                # Let's read first 40 rows and search for 'Đức Linh'
                # Open with data_only=True to search values
                wb_val = openpyxl.load_workbook(fpath, data_only=True)
                sheet = wb_val[sname]
                for r in range(1, sheet.max_row + 1):
                    row_vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
                    if any(isinstance(v, str) and 'Đức Linh' in v for v in row_vals if v is not None):
                        print(f"  Row {r}: {row_vals[:15]}")
    except Exception as e:
        pass

print("\nDone searching.")
