import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\V3_AOP_Hang_NTB_T7-T12_2026 mới.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

targets = [1303, 1303.0, 905, 905.0, 83, 83.0, 120, 120.0]
for sname in wb.sheetnames:
    sheet = wb[sname]
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(r, c).value
            if val in targets:
                print(f"Found {val} in Sheet: '{sname}', Cell: {openpyxl.utils.get_column_letter(c)}{r}")
wb.close()
