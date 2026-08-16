import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

# Path to the monthly file
monthly_file = r"C:\Users\lap4all\Downloads\AOP_Hang_NTB_T7-T12_2026.xlsx"
if not os.path.exists(monthly_file):
    print("Monthly file not found!")
    sys.exit(1)

wb = openpyxl.load_workbook(monthly_file, data_only=True)
sheet = wb['0.3 Bưu cục Detail']

print("=== 0.3 Bưu cục Detail (Monthly) ===")
# Print headers
headers = [cell.value for cell in sheet[2]]
print("Headers:", headers)

for r in range(3, sheet.max_row + 1):
    vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
    if any(v is not None for v in vals):
        bc_name = sheet.cell(r, 2).value
        if bc_name and any(name in bc_name for name in ['Nha Trang', 'Đơn Dương', 'Di Linh', 'Đức Linh']):
            # Print row info
            print(f"Row {r}: BC = {bc_name}")
            for c_idx, val in enumerate(vals):
                if val is not None:
                    print(f"  {headers[c_idx]}: {val}")
