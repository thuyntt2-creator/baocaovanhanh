import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

fpath = r"C:\Users\lap4all\Downloads\AOP_Hang_NTB_T7-T12_2026.xlsx"
wb = openpyxl.load_workbook(fpath, data_only=True)

# Find sheets containing 'mặt bằng' or 'mật độ' or 'phương án'
for sname in wb.sheetnames:
    if 'mặt bằng' in sname.lower() or 'mật độ' in sname.lower() or 'phương án' in sname.lower():
        sheet = wb[sname]
        print(f"\n==========================================")
        print(f"Sheet: {sname} (Rows: {sheet.max_row}, Cols: {sheet.max_column})")
        print(f"==========================================")
        
        # Scan all cells for 'Đức Linh' or 'Diện tích hiện tại'
        for r in range(1, sheet.max_row + 1):
            row_vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
            row_str = " | ".join([f"{openpyxl.utils.get_column_letter(c+1)}: {repr(val)}" for c, val in enumerate(row_vals) if val is not None])
            
            # Check if Đức Linh or Diện tích or trống is in the row
            has_duc_linh = any('Đức Linh' in str(val) for val in row_vals if val is not None)
            has_dt = any('Diện tích' in str(val) for val in row_vals if val is not None)
            
            if has_duc_linh or has_dt or r <= 10:
                print(f"Row {r:2d}: {row_str}")
