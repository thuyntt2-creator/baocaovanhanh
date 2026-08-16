import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\V3_AOP_Hang_NTB_T7-T12_2026 mới.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['Kế hoạch T7']

print("=== Rows containing 'BCCK' or 'Bưu cục' in Kế hoạch T7 ===")
for r in range(1, sheet.max_row + 1):
    val = sheet.cell(r, 1).value or sheet.cell(r, 2).value
    if val and any(x in str(val).lower() for x in ['bcck', 'bưu cục', 'mặt bằng', 'thuê', 'nha trang', 'di linh', 'đơn dương', 'đức linh']):
        row_vals = [sheet.cell(r, c).value for c in range(1, 10)]
        print(f"Row {r:3d}: {row_vals}")

wb.close()
