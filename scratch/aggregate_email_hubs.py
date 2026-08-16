import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\lap4all\Documents\Auto report\NTB_Phan_Tuyen_Hanh_Chinh_Quy_Hoach_Chi_Tiet.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Sheet1']

headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column+1)]

# Group by Province -> Hub
hubs = {}

for r in range(2, sheet.max_row+1):
    row = {headers[i]: sheet.cell(r, i+1).value for i in range(len(headers)) if headers[i]}
    prov = str(row.get('Tỉnh, thành phố mới') or '').strip()
    bc_new = str(row.get('Tên Bưu cục giao mới đề xuất') or '').strip()
    xa_moi = str(row.get('Tên Xã mới') or '').strip()
    xa_cu = str(row.get('Tên Xã cũ') or '').strip()
    phuong_an = str(row.get('Đánh giá & Phương án đề xuất') or '').strip()
    ly_do = str(row.get('Lý do & Bố trí nhân sự') or '').strip()
    am = str(row.get('Quản lý khu vực (AM)') or '').strip()
    
    if not prov or not bc_new:
        continue
        
    key = (prov, bc_new)
    if key not in hubs:
        hubs[key] = {
            'prov': prov,
            'bc': bc_new,
            'xa_moi': set(),
            'xa_cu': set(),
            'phuong_an': phuong_an,
            'ly_do': ly_do,
            'am': am
        }
    if xa_moi:
        hubs[key]['xa_moi'].add(xa_moi)
    if xa_cu:
        hubs[key]['xa_cu'].add(xa_cu)

print(f"Total unique (Province, BC) pairs: {len(hubs)}")

provinces = {}
for (prov, bc), info in hubs.items():
    if prov not in provinces:
        provinces[prov] = []
    provinces[prov].append(info)

for prov, items in provinces.items():
    print(f"\n=================== TỈNH: {prov.upper()} ({len(items)} BƯU CỤC) ===================")
    for item in items:
        print(f"❖ BC: {item['bc']}")
        print(f"   Cover Xã mới: {', '.join(sorted(item['xa_moi']))}")
        print(f"   Cover Xã cũ: {', '.join(sorted(item['xa_cu']))}")
        print(f"   Phương án: {item['phuong_an']}")
        print(f"   AM: {item['am']}")
