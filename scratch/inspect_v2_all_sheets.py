import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

v2_path = r"C:\Users\lap4all\Downloads\[V2] AOP_Hang_NTB_T7-T12_2026.xlsx"
wb = openpyxl.load_workbook(v2_path, data_only=False)

for sname in ['1. Thông số', 'Volume Giao', 'Volume Lấy', '0.3 Bưu cục Detail', 'Kênh & nhu cầu', 'Nguồn lực & chi phí', 'Mặt bằng']:
    sheet = wb[sname]
    print(f"\n==================== Sheet: {sname} (Rows: {sheet.max_row}, Cols: {sheet.max_column}) ====================")
    for r_idx in range(1, min(sheet.max_row + 1, 15)):
        row_vals = []
        for c_idx in range(1, min(sheet.max_column + 1, 15)):
            cell = sheet.cell(r_idx, c_idx)
            val = cell.value
            if isinstance(val, str) and val.startswith('='):
                row_vals.append(f"F:{val}")
            else:
                row_vals.append(val)
        if any(v is not None for v in row_vals):
            print(f"Row {r_idx:2d}: {row_vals}")

print("\n" + "="*50)
# Check sheet Mặt bằng row 3-10 specifically
mb_sheet = wb['Mặt bằng']
print("Mặt bằng sheet row 2-8:")
for r in range(2, 9):
    row_vals = [mb_sheet.cell(r, c).value for c in range(1, 15)]
    print(f"  Row {r}: {row_vals}")

