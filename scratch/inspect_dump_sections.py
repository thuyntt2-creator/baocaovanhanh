import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r'C:\Users\lap4all\Downloads\BaoCao_Tuan_NTB_W32_2026.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    print(f"\n==========================================")
    print(f"=== SHEET: {sheetname} ({ws.max_row}x{ws.max_column}) ===")
    print(f"==========================================")
    for r in range(1, ws.max_row+1):
        vals = [ws.cell(r, c).value for c in range(1, min(15, ws.max_column+1))]
        if any(v is not None for v in vals):
            row_str = []
            for v in vals:
                if v is None:
                    row_str.append("")
                elif isinstance(v, float):
                    if -1.0 <= v <= 1.0 and v != 0:
                        row_str.append(f"{v*100:.1f}%")
                    else:
                        row_str.append(f"{v:,.1f}".rstrip('0').rstrip('.'))
                else:
                    row_str.append(str(v))
            while row_str and row_str[-1] == "":
                row_str.pop()
            print(f"R{r:2d}:", row_str)
