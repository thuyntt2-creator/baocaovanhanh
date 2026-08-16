import openpyxl
import os
import glob
import sys

sys.stdout.reconfigure(encoding='utf-8')

downloads_dir = r"C:\Users\lap4all\Downloads"

# Find files
v3_files = glob.glob(os.path.join(downloads_dir, "*V3_AOP_Hang_NTB*"))
input_files = glob.glob(os.path.join(downloads_dir, "*NTB_Input*"))

print("V3 files found:", v3_files)
print("Input files found:", input_files)

# Check V3 sheets and sample contents
if v3_files:
    v3_path = v3_files[0]
    wb = openpyxl.load_workbook(v3_path, read_only=True)
    print("\nSheets in V3:", wb.sheetnames)
    wb.close()

# Check Input sheets and sample contents
if input_files:
    input_path = input_files[0]
    wb = openpyxl.load_workbook(input_path, read_only=True)
    print("\nSheets in Input file:", wb.sheetnames)
    wb.close()
