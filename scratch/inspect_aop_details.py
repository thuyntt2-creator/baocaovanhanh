import zipfile
import xml.etree.ElementTree as ET
import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

docx_path = r"C:\Users\lap4all\Downloads\AOP_BCCK_Plan_new.docx"
xlsx_aop_path = r"C:\Users\lap4all\Downloads\AOP_Hang_NTB_T7-T12_2026.xlsx"
xlsx_config_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB.xlsx"
xlsx_compare_path = r"C:\Users\lap4all\Downloads\[NTB] So sánh topline H2 Mới - Cũ.xlsx"

def get_docx_all_elements(path):
    if not os.path.exists(path):
        return f"File not found: {path}"
    try:
        with zipfile.ZipFile(path) as z:
            xml_content = z.read('word/document.xml')
            root = ET.fromstring(xml_content)
            namespaces = {
                'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'
            }
            
            # We want to iterate through body elements in order
            body = root.find('w:body', namespaces)
            if body is None:
                return "No body found"
                
            out = []
            for child in body:
                tag = child.tag.split('}')[-1]
                if tag == 'p':
                    # Paragraph
                    p_text = "".join(t.text for t in child.findall('.//w:t', namespaces) if t.text)
                    if p_text.strip():
                        out.append(f"[P] {p_text}")
                elif tag == 'tbl':
                    # Table
                    out.append("[Table]")
                    rows = child.findall('w:tr', namespaces)
                    for r_idx, r in enumerate(rows):
                        cells = r.findall('w:tc', namespaces)
                        cell_texts = []
                        for cell in cells:
                            cell_text = "".join(t.text for t in cell.findall('.//w:t', namespaces) if t.text)
                            cell_texts.append(cell_text.strip())
                        out.append(f"  Row {r_idx}: " + " | ".join(cell_texts))
            return "\n".join(out)
    except Exception as e:
        return f"Error: {e}"

def inspect_sheet_info(path, sheet_name):
    if not os.path.exists(path):
        return f"File not found: {path}"
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        if sheet_name not in wb.sheetnames:
            return f"Sheet {sheet_name} not in {os.path.basename(path)}"
        sheet = wb[sheet_name]
        out = [f"--- Sheet {sheet_name} in {os.path.basename(path)} (Rows: {sheet.max_row}, Cols: {sheet.max_column}) ---"]
        # Print first 20 rows
        for r_idx in range(1, min(sheet.max_row + 1, 40)):
            row_vals = [sheet.cell(r_idx, c_idx).value for c_idx in range(1, min(sheet.max_column + 1, 15))]
            if any(v is not None for v in row_vals):
                out.append(f"Row {r_idx}: {row_vals}")
        return "\n".join(out)
    except Exception as e:
        return f"Error: {e}"

print("=== ALL DOCX CONTENT ===")
print(get_docx_all_elements(docx_path))

print("\n" + "="*40 + "\n")
print(inspect_sheet_info(xlsx_aop_path, '1. Thông số'))
print("\n" + "="*40 + "\n")
print(inspect_sheet_info(xlsx_aop_path, '0. Thông tin vùng'))
print("\n" + "="*40 + "\n")
print(inspect_sheet_info(xlsx_aop_path, 'Mặt bằng'))

