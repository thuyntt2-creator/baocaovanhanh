import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\V3_AOP_Hang_NTB_T7-T12_2026 mới.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

for name in ['0. Thông tin vùng', '1. Thông số']:
    sheet = wb[name]
    print(f"\n=== Sheet: {name} (max_row: {sheet.max_row}, max_col: {sheet.max_column}) ===")
    for r in range(1, 40):
        row_vals = [sheet.cell(r, c).value for c in range(1, 15)]
        if any(v is not None for v in row_vals):
            print(f"Row {r:2d}: {row_vals}")

wb.close()
