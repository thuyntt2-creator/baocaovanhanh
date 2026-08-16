import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\V3_AOP_Hang_NTB_T7-T12_2026 mới.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['Tổng hợp xe']

print("=== Sheet: Tổng hợp xe ===")
for r in range(1, sheet.max_row + 1):
    row_vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
    if any(v is not None for v in row_vals):
        print(f"Row {r:2d}: {row_vals}")

wb.close()
