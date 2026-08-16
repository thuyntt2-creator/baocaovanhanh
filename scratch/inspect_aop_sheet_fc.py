import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v18.xlsx"

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    sys.exit(1)

wb_formula = openpyxl.load_workbook(file_path, data_only=False)
wb_value = openpyxl.load_workbook(file_path, data_only=True)

print("Sheet names in the file:")
for idx, name in enumerate(wb_formula.sheetnames):
    print(f"{idx}: {name}")

# Let's inspect the formulas in all sheets and find where monthly FC or forecast is calculated.
for sheet_name in wb_formula.sheetnames:
    sheet_f = wb_formula[sheet_name]
    sheet_v = wb_value[sheet_name]
    
    print(f"\n==================================================")
    print(f"SHEET: {sheet_name}")
    print(f"==================================================")
    
    # We will search for any cells containing 'FC' or 'Forecast' (case-insensitive) in formulas or values
    matches = []
    for r in range(1, sheet_f.max_row + 1):
        for c in range(1, sheet_f.max_column + 1):
            val_f = sheet_f.cell(r, c).value
            val_v = sheet_v.cell(r, c).value
            
            if val_f is not None:
                str_f = str(val_f).upper()
                str_v = str(val_v).upper() if val_v is not None else ""
                
                # Check for "FC" or "FORECAST" or "ROUNDUP" or "AVERAGE"
                if "FC" in str_f or "FORECAST" in str_f or "FC" in str_v or "FORECAST" in str_v:
                    matches.append((r, c, val_f, val_v))
                    
    print(f"Found {len(matches)} cells matching 'FC' or 'Forecast'.")
    if len(matches) > 0:
        print("First 30 matches:")
        for r, c, val_f, val_v in matches[:30]:
            col_letter = openpyxl.utils.get_column_letter(c)
            print(f"  {col_letter}{r}: Formula/Text = {repr(val_f)} | Value = {repr(val_v)}")

    # Let's also look for column or row headers containing "FC" or "forecast"
    # and print some rows around them to see the context.
