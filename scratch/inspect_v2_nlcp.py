import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

v2_path = r"C:\Users\lap4all\Downloads\[V2] AOP_Hang_NTB_T7-T12_2026.xlsx"
wb = openpyxl.load_workbook(v2_path, data_only=False)
sheet = wb['Nguồn lực & chi phí']

print("=== [V2] Nguồn lực & chi phí ===")
for r in range(1, sheet.max_row + 1):
    vals = []
    for c in range(1, 10):
        cell = sheet.cell(r, c)
        val = cell.value
        if isinstance(val, str) and val.startswith('='):
            vals.append(f"F:{val}")
        else:
            vals.append(val)
    if any(v is not None for v in vals):
        print(f"Row {r:2d}: {vals}")

