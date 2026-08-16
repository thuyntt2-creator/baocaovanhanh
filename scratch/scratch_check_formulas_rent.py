import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['NTB – Input']

cols = ['D', 'E', 'F', 'G', 'H', 'I']
print("=== Formulas in NTB_Input_FLM_CRC.xlsx (Rows 34, 35, 36) ===")
for r in [34, 35, 36]:
    row_label = sheet.cell(r, 2).value or sheet.cell(r, 1).value or f"Row {r}"
    row_vals = [sheet.cell(r, openpyxl.utils.column_index_from_string(c)).value for c in cols]
    print(f"Row {r:2d} | {str(row_label)[:35]:<35} | {row_vals}")

wb.close()
