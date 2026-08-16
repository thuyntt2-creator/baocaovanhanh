import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['NTB – Input']

print("=== NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx summary rows formulas ===")
cols = ['D', 'E', 'F', 'G', 'H', 'I']
for r in range(41, 46):
    label = sheet.cell(r, 2).value or sheet.cell(r, 1).value
    print(f"Row {r:2d} | {str(label)[:35]:<35}")
    for c in cols:
        print(f"  {c}: {sheet[f'{c}{r}'].value}")
wb.close()
