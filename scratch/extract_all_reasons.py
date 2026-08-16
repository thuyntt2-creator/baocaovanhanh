import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\lap4all\Documents\Auto report\NTB_Phan_Tuyen_Hanh_Chinh_Quy_Hoach_Chi_Tiet.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Sheet1']

headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column+1)]

by_prov = {}

for r in range(2, sheet.max_row+1):
    row = {headers[i]: sheet.cell(r, i+1).value for i in range(len(headers)) if headers[i]}
    prov = str(row.get('Tỉnh, thành phố mới') or '').strip()
    xa_moi = str(row.get('Tên Xã mới') or '').strip()
    bc_moi = str(row.get('Tên Bưu cục giao mới đề xuất') or '').strip()
    dexuat = str(row.get('Đánh giá & Phương án đề xuất') or '').strip()
    lydo = str(row.get('Lý do & Bố trí nhân sự') or '').strip()
    vol_giao = row.get('Sản lượng giao/ngày (xã mới)')
    vol_lay = row.get('Sản lượng lấy/ngày (xã mới)')
    
    if not prov or not xa_moi:
        continue
        
    if prov not in by_prov:
        by_prov[prov] = {}
        
    if bc_moi not in by_prov[prov]:
        by_prov[prov][bc_moi] = {
            'xa_moi': [],
            'dexuat': dexuat,
            'lydo': lydo,
            'vol_giao': 0,
            'vol_lay': 0
        }
    by_prov[prov][bc_moi]['xa_moi'].append(xa_moi)
    if vol_giao:
        by_prov[prov][bc_moi]['vol_giao'] += float(vol_giao)
    if vol_lay:
        by_prov[prov][bc_moi]['vol_lay'] += float(vol_lay)

for prov, hubs in by_prov.items():
    print(f"================ {prov} ================")
    for bc, info in hubs.items():
        print(f"BC: {bc}")
        print(f"  Action: {info['dexuat']}")
        print(f"  Cover: {', '.join(set(info['xa_moi']))}")
        print(f"  Lý do: {info['lydo']}")
        print(f"  Vol: Giao {int(info['vol_giao'])}, Lấy {int(info['vol_lay'])}")
        print()
