import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

print("=== Checking references to row 36 of 'NTB – Input' ===")
referenced = False
for sname in wb.sheetnames:
    sheet = wb[sname]
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(r, c).value
            if isinstance(val, str) and "Input'!" in val:
                # check if it references row 36
                if "36" in val:
                    print(f"Sheet: {sname}, Cell: {openpyxl.utils.get_column_letter(c)}{r} contains formula: {val}")
                    referenced = True
if not referenced:
    print("No formulas reference Row 36 of 'NTB – Input'. It is just for display/reporting.")
wb.close()
