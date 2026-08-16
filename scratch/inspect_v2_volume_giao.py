import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

v2_path = r"C:\Users\lap4all\Downloads\[V2] AOP_Hang_NTB_T7-T12_2026.xlsx"
wb = openpyxl.load_workbook(v2_path, data_only=True)

for sheet_name in ['Volume Giao', 'Volume Lấy']:
    sheet = wb[sheet_name]
    print(f"\n--- [V2] Sheet: {sheet_name} (max_row: {sheet.max_row}, max_column: {sheet.max_column}) ---")
    non_empty = 0
    sample_rows = []
    for r in range(1, sheet.max_row + 1):
        vals = [sheet.cell(r, c).value for c in range(1, min(sheet.max_column + 1, 15))]
        if any(v is not None for v in vals):
            non_empty += 1
            if len(sample_rows) < 15:
                sample_rows.append((r, vals))
    print(f"Non-empty rows count: {non_empty}")
    print("First few rows:")
    for r, vals in sample_rows:
        print(f"  Row {r:2d}: {vals[:10]}")

