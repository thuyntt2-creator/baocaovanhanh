import openpyxl
import os
import glob
import sys

sys.stdout.reconfigure(encoding='utf-8')

downloads_dir = r"C:\Users\lap4all\Downloads"
xlsx_files = [
    os.path.join(downloads_dir, "AOP_MAU_NTB_T7-T12_2026_v4.xlsx"),
    os.path.join(downloads_dir, "AOP_MAU_NTB_T7-T12_2026_v3.xlsx"),
    os.path.join(downloads_dir, "V2 AOP_Hang_NTB_T7-T12_2026.xlsx"),
    os.path.join(downloads_dir, "[V2] AOP_Hang_NTB_T7-T12_2026.xlsx"),
]
xlsx_files = [f for f in xlsx_files if os.path.exists(f)]

print("=== Checking Row 38 & Row 39 in all Excel files ===")
for f in xlsx_files:
    if "~$" in f:
        continue
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        for sname in wb.sheetnames:
            if "input" in sname.lower() or sname == "NTB":
                sheet = wb[sname]
                # Try to locate the label "Setup mở mới" or row 38/39
                for r in range(1, min(60, sheet.max_row + 1)):
                    row_lbl = sheet.cell(r, 2).value or sheet.cell(r, 1).value
                    if row_lbl and any(term in str(row_lbl).lower() for term in ["setup", "di dời"]):
                        vals = [sheet.cell(r, c).value for c in range(4, 10)]
                        print(f"File: {os.path.basename(f)} | Sheet: {sname} | Row {r} ({row_lbl}): {vals}")
        wb.close()
    except Exception as e:
        pass
