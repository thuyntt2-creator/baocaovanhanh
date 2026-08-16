import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\V3_AOP_Hang_NTB_T7-T12_2026 mới.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} (max_row: {sheet.max_row}, max_col: {sheet.max_column}) ===")
    
    # print first 5 rows and 8 columns
    for r in range(1, min(6, sheet.max_row + 1)):
        row_vals = [sheet.cell(r, c).value for c in range(1, min(9, sheet.max_column + 1))]
        if any(v is not None for v in row_vals):
            print(f"  Row {r:2d}: {row_vals}")

wb.close()
