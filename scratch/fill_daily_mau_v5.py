import openpyxl, sys, shutil
sys.stdout.reconfigure(encoding='utf-8')

# Paths
config_path = r'C:\Users\lap4all\Downloads\config_psbba_NTB.xlsx'
monthly_path = r'C:\Users\lap4all\Downloads\V2 AOP_Hang_NTB_T7-T12_2026.xlsx'
mau_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026.xlsx'
out_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v5.xlsx'

groups = {
    'BCCK Nha Trang': ['Nha Trang'],
    'BCCK Đơn Dương': ['Đơn Dương', 'Hiệp Thạnh'],
    'BCCK Di Linh': ['Di Linh', 'Hòa Ninh', 'Lâm Hà'],
    'BCCK Đức Linh': ['Đức Linh']
}

print("1. Đọc data từ 7_FC_Giao_Daily để lấy số liệu T7, T8 và tính Tỷ trọng...")
wb_cfg = openpyxl.load_workbook(config_path, data_only=True)
sheet_fc = wb_cfg['7_FC_Giao_Daily']

col_map = {}
for c in range(6, sheet_fc.max_column):
    header = sheet_fc.cell(1, c).value
    if header:
        parts = str(header).split()
        if len(parts) >= 2:
            col_map[parts[1]] = c # e.g. '01/08': col_idx

daily_vol = {g: {} for g in groups.keys()}

for r in range(2, sheet_fc.max_row+1):
    bc = sheet_fc.cell(r, 4).value
    band = sheet_fc.cell(r, 5).value
    if bc and band and 'Bulky' in str(band).strip():
        bc_str = str(bc).strip()
        if bc_str == 'Kho Giao Hàng Nặng - Nha Trang - Khánh Hoà':
            continue
        
        matched_group = None
        for g_name, keywords in groups.items():
            for kw in keywords:
                if kw in bc_str:
                    matched_group = g_name
                    break
            if matched_group: break
        
        if matched_group:
            for date_str, c in col_map.items():
                vol = sheet_fc.cell(r, c).value or 0
                daily_vol[matched_group][date_str] = daily_vol[matched_group].get(date_str, 0) + vol


print("2. Đọc Tổng Sản Lượng T9-T12 từ file Monthly AOP...")
wb_mon = openpyxl.load_workbook(monthly_path, data_only=True)
sheet_bc = wb_mon['0.3 Bưu cục Detail']

monthly_totals = {g: {'T9': 0, 'T10': 0, 'T11': 0, 'T12': 0} for g in groups.keys()}

for r in range(3, 50):
    bc_name = sheet_bc.cell(r, 2).value
    if bc_name:
        bc_str = str(bc_name).strip()
        matched_group = None
        for g_name, keywords in groups.items():
            for kw in keywords:
                if kw in bc_str:
                    matched_group = g_name
                    break
            if matched_group: break
        
        if matched_group:
            # Columns in 0.3 Bưu cục Detail for T9-T12
            # Assuming T7=C, T8=D, T9=E(5), T10=F(6), T11=G(7), T12=H(8)
            monthly_totals[matched_group]['T9'] += sheet_bc.cell(r, 5).value or 0
            monthly_totals[matched_group]['T10'] += sheet_bc.cell(r, 6).value or 0
            monthly_totals[matched_group]['T11'] += sheet_bc.cell(r, 7).value or 0
            monthly_totals[matched_group]['T12'] += sheet_bc.cell(r, 8).value or 0

print("3. Đắp data vào file MAU V5...")
shutil.copy(mau_path, out_path)
wb_mau = openpyxl.load_workbook(out_path)

# Cập nhật thông số
sheet_param = wb_mau['1. Thông số']
sheet_param['B2'] = 1.4
sheet_param['B3'] = 20
sheet_param['B4'] = 30
sheet_param['B5'] = 47
sheet_param['B6'] = 64
sheet_param['B10'] = 0.855
sheet_param['B11'] = 0.858
sheet_param['B12'] = 0.859
sheet_param['B13'] = 0.858
sheet_param['B14'] = 0.858
sheet_param['B15'] = 0.846
sheet_param['B16'] = 0.825
sheet_param['B19'] = 55
sheet_param['B20'] = 2.0
sheet_param['B21'] = 3.3
sheet_param['B22'] = 150000
sheet_param['B23'] = 1200000
sheet_param['B24'] = 15000000

months_info = {
    'T7': 31, 'T8': 31, 'T9': 30, 'T10': 31, 'T11': 30, 'T12': 31
}

group_names = list(groups.keys())

for m_idx, (m_name, num_days) in enumerate(months_info.items()):
    sheet_f = wb_mau[f'Forecast {m_name}']
    
    for i in range(8):
        r_bc = 4 + i
        if i < len(group_names):
            g_name = group_names[i]
            sheet_f.cell(r_bc, 1).value = g_name
            
            if m_name in ['T7', 'T8']:
                # Dùng số thực tế
                for day in range(1, num_days + 1):
                    date_col = day + 1
                    month_num = 7 + m_idx
                    date_str = f"{day:02d}/{month_num:02d}"
                    val = daily_vol[g_name].get(date_str, 0)
                    sheet_f.cell(r_bc, date_col).value = round(val) if val else None
            else:
                # Dùng trọng số T8
                # 1. Tính mảng số T8
                t8_vals = []
                for day in range(1, num_days + 1):
                    date_str = f"{day:02d}/08"
                    t8_vals.append(daily_vol[g_name].get(date_str, 0))
                
                sum_t8 = sum(t8_vals)
                target_total = monthly_totals[g_name][m_name]
                
                for day in range(1, num_days + 1):
                    date_col = day + 1
                    if sum_t8 > 0:
                        weight = t8_vals[day-1] / sum_t8
                        gen_val = target_total * weight
                        sheet_f.cell(r_bc, date_col).value = round(gen_val)
                    else:
                        # Fallback nếu T8 không có số (tránh lỗi div/0)
                        sheet_f.cell(r_bc, date_col).value = round(target_total / num_days)
        else:
            sheet_f.cell(r_bc, 1).value = None
            for day in range(1, 32):
                sheet_f.cell(r_bc, day + 1).value = None

wb_mau.save(out_path)
print(f"Hoàn thành! Đã lưu: {out_path}")
