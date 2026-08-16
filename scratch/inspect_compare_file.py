import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

xlsx_compare_path = r"C:\Users\lap4all\Downloads\[NTB] So sánh topline H2 Mới - Cũ.xlsx"

wb = openpyxl.load_workbook(xlsx_compare_path, data_only=True)
sheet = wb['NTB']
print(f"Rows: {sheet.max_row}, Cols: {sheet.max_column}")
for r_idx in range(1, min(sheet.max_row + 1, 30)):
    row_vals = [sheet.cell(r_idx, c_idx).value for c_idx in range(1, min(sheet.max_column + 1, 15))]
    if any(v is not None for v in row_vals):
        print(f"Row {r_idx:2d}: {row_vals}")

