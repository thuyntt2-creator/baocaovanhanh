import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

tg_dir = r"C:\Users\lap4all\Downloads\Telegram Desktop"
filepath = os.path.join(tg_dir, "CRC_DCL_Input_FLM_2026.xlsx")

wb = openpyxl.load_workbook(filepath, data_only=True)
sheet = wb['ĐCL – Input']

print("=== ĐCL – Input rows 1-60 values ===")
for r in range(1, 61):
    c1 = sheet.cell(r, 1).value
    c2 = sheet.cell(r, 2).value
    c3 = sheet.cell(r, 3).value
    d = sheet.cell(r, 4).value
    if c1 or c2 or d:
        print(f"Row {r:2d} | {str(c1):<4} | {str(c2)[:35]:<35} | {str(c3):<10} | T7={d}")
wb.close()
