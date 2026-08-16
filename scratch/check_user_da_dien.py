import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

downloads_dir = r"C:\Users\lap4all\Downloads"

def check_file(filename):
    path = os.path.join(downloads_dir, filename)
    if os.path.exists(path):
        print(f"\n=== File: {filename} ===")
        wb = openpyxl.load_workbook(path, data_only=True)
        sheet = wb['NTB – Input']
        cols = ['D', 'E', 'F', 'G', 'H', 'I']
        for r in range(41, 46):
            label = sheet.cell(r, 2).value or sheet.cell(r, 1).value
            vals = [sheet.cell(r, openpyxl.utils.column_index_from_string(c)).value for c in cols]
            print(f"Row {r:2d} | {str(label)[:35]:<35} | {vals}")
        wb.close()

check_file("NTB_Input_Da_Dien_FLM_CRC_1.xlsx")
check_file("NTB_Input_Da_Dien_FLM_CRC_2.xlsx")
check_file("NTB_Input_Da_Dien_FLM_CRC.xlsx")
