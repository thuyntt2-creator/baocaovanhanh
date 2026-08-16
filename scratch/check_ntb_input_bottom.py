import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['NTB – Input']

print("=== NTB – Input rows 50-70 ===")
for r in range(50, 71):
    c1 = sheet.cell(r, 1).value
    c2 = sheet.cell(r, 2).value
    c3 = sheet.cell(r, 3).value
    d = sheet.cell(r, 4).value
    if c1 or c2 or d:
        print(f"Row {r:2d} | {str(c1):<10} | {str(c2)[:35]:<35} | T7={d}")

wb.close()
