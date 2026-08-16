import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\AOP_Hang_NTB_T7-T12_2026_calculated.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
print("Sheet names:", wb.sheetnames)
# Try to find a sheet containing 'Chi phí FLM' or similar
sheet_name = [s for s in wb.sheetnames if 'Chi' in s and 'FLM' in s][0]
sheet = wb[sheet_name]

print(f"=== Formulas for Row 13 in Calculated '{sheet_name}' ===")
cols = ['D', 'E', 'F', 'G', 'H', 'I']
for col in cols:
    cell = sheet[f"{col}13"]
    print(f"Cell {col}13: {cell.value}")

wb.close()
