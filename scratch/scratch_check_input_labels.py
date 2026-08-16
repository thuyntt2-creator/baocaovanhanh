import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['NTB – Input']

print("=== Labels and formulas for Rows 40-45 in 'NTB – Input' ===")
for r in range(40, 46):
    lbl = sheet.cell(r, 2).value or sheet.cell(r, 1).value
    val = sheet.cell(r, 4).value # Column D
    print(f"Row {r:2d} | Label: '{lbl}' | Formula: {val}")

wb.close()
