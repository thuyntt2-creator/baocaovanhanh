import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['NTB – Input']

cols = ['D', 'E', 'F', 'G', 'H', 'I']
months = ['T7', 'T8', 'T9', 'T10', 'T11', 'T12']

print("=== Kết quả cuối cùng - NTB Input ===\n")
rows_to_check = [8, 9, 11, 12, 13, 17, 20, 22, 27, 32, 37, 38, 39, 43, 44, 45]

print(f"{'Dòng':<4} | {'Tên chỉ số':<35} | " + " | ".join(f"{m:>12}" for m in months))
print("-" * 130)
for r in rows_to_check:
    row_label = sheet.cell(r, 2).value or sheet.cell(r, 1).value or f"Row {r}"
    row_label = str(row_label)[:34]
    vals = []
    for c in cols:
        val = sheet[f"{c}{r}"].value
        if isinstance(val, float):
            if val == int(val):
                vals.append(f"{int(val):,}")
            else:
                vals.append(f"{val:,.2f}")
        elif isinstance(val, int):
            vals.append(f"{val:,}")
        else:
            vals.append(str(val) if val is not None else "–")
    print(f"{r:<4} | {row_label:<35} | " + " | ".join(f"{v:>12}" for v in vals))

wb.close()
