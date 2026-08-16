import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

config_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB.xlsx"
wb = openpyxl.load_workbook(config_path, data_only=True)

for sname in ['0_README', 'GHI_CHU_Topline']:
    if sname in wb.sheetnames:
        sheet = wb[sname]
        print(f"\n=== Sheet: {sname} ===")
        for r in range(1, min(sheet.max_row + 1, 30)):
            row_vals = [sheet.cell(r, c).value for c in range(1, min(sheet.max_column + 1, 15))]
            if any(v is not None for v in row_vals):
                print(f"Row {r:2d}: {row_vals}")

