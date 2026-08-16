import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

monthly_file = r"C:\Users\lap4all\Downloads\AOP_Hang_NTB_T7-T12_2026.xlsx"
wb = openpyxl.load_workbook(monthly_file, data_only=True)
sheet = wb['0.3 Bưu cục Detail']

print("=== 0.3 Bưu cục Detail Sums ===")
total_t7 = 0.0
total_t12 = 0.0
for r in range(3, sheet.max_row + 1):
    t7_val = sheet.cell(r, 4).value # col D is T7
    t12_val = sheet.cell(r, 9).value # col I is T12
    if t7_val is not None:
        total_t7 += float(t7_val)
    if t12_val is not None:
        total_t12 += float(t12_val)

print(f"Total T7 across all BCs: {total_t7:.2f}")
print(f"Total T12 across all BCs: {total_t12:.2f}")
