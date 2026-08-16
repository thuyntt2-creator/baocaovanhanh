import openpyxl
import os
import glob
import sys

sys.stdout.reconfigure(encoding='utf-8')

downloads_dir = r"C:\Users\lap4all\Downloads"
files = [
    os.path.join(downloads_dir, "[V2] AOP_Hang_NTB_T7-T12_2026.xlsx"),
    os.path.join(downloads_dir, "V2 AOP_Hang_NTB_T7-T12_2026.xlsx"),
    os.path.join(downloads_dir, "AOP_NTB_v3.xlsx"),
    os.path.join(downloads_dir, "AOP_NTB_calculated_v2.xlsx"),
]
# Filter out non-existent files
files = [f for f in files if os.path.exists(f)]

targets = [198219, 68578, 129641]

print("=== Searching in files ===")
for f in files:
    if "~$" in f:
        continue
    try:
        wb = openpyxl.load_workbook(f, read_only=True)
        # Check sheet names
        for sname in wb.sheetnames:
            if "giao" in sname.lower() or "lấy" in sname.lower() or "input" in sname.lower() or "detail" in sname.lower():
                print(f"File: {os.path.basename(f)} | Sheet: {sname}")
        wb.close()
    except Exception as e:
        pass

wb_in = openpyxl.load_workbook(r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx", data_only=True)
print("\n=== Sheets in Input File: ===")
for s in wb_in.sheetnames:
    print(f"  {s}")
wb_in.close()
