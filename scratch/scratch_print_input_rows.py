import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['NTB – Input']

cols = ['C', 'D', 'E', 'F', 'G', 'H', 'I']
print("=== VERIFIED VALUES IN NTB_Input_FLM_CRC.xlsx ===")
for r in range(7, 46):
    row_label = sheet.cell(r, 2).value or sheet.cell(r, 1).value or f"Row {r}"
    row_vals = [sheet.cell(r, openpyxl.utils.column_index_from_string(c)).value for c in cols]
    if any(v is not None for v in row_vals):
        print(f"Row {r:2d} | {str(row_label)[:35]:<35} | {row_vals}")

wb.close()
