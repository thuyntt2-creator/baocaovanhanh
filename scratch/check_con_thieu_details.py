import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

filepath = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(filepath, data_only=True)
sheet = wb['NTB – Input']

cols = ['D', 'E', 'F', 'G', 'H', 'I']
print("=== NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx values (Rows 15-39) ===")
for r in range(15, 40):
    label = sheet.cell(r, 2).value or sheet.cell(r, 1).value
    vals = [sheet.cell(r, openpyxl.utils.column_index_from_string(c)).value for c in cols]
    print(f"Row {r:2d} | {str(label)[:35]:<35} | {vals}")
wb.close()
