import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

# Paths
template_path = r"C:\Users\lap4all\Downloads\AOP_V2_updated_2.xlsx"
config_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB.xlsx"

# Load workbooks
wb_temp = openpyxl.load_workbook(template_path, data_only=True)
wb_conf = openpyxl.load_workbook(config_path, data_only=True)

sheet_tl = wb_temp['Timeline tiếp nhận']
sheet_cfg = wb_conf['1_Config_Chuan']

# 1. Read config mappings: (District, Ward) -> BC_Giao_Tên
cfg_map = {}
for r in range(2, sheet_cfg.max_row + 1):
    dist = sheet_cfg.cell(r, 3).value
    ward = sheet_cfg.cell(r, 5).value
    bc_giao = sheet_cfg.cell(r, 9).value
    if dist and bc_giao:
        key = (dist.strip().lower(), ward.strip().lower() if ward else "")
        cfg_map[key] = bc_giao.strip()

# 2. Read timeline mappings: (District, Ward) -> BCCK_Dự_Kiến
bcck_bcs = {}
missing = []
for r in range(3, sheet_tl.max_row + 1):
    ward = sheet_tl.cell(r, 1).value
    dist = sheet_tl.cell(r, 2).value
    bcck = sheet_tl.cell(r, 4).value
    
    if not ward and not dist and not bcck:
        # Check if empty row
        row_vals = [sheet_tl.cell(r, c).value for c in range(1, 10)]
        if not any(row_vals):
            continue
            
    if bcck:
        bcck = bcck.strip()
        ward_clean = ward.strip().lower() if ward else ""
        dist_clean = dist.strip().lower() if dist else ""
        
        # Try matching by (dist, ward)
        key = (dist_clean, ward_clean)
        bc_name = cfg_map.get(key)
        
        # If not found, try matching by ward alone
        if not bc_name:
            # search config map for ward
            for (d_cfg, w_cfg), bc in cfg_map.items():
                if w_cfg == ward_clean:
                    bc_name = bc
                    break
                    
        if bc_name:
            bcck_bcs.setdefault(bcck, set())
            bcck_bcs[bcck].add(bc_name)
        else:
            missing.append((dist, ward, bcck))

print("=== Joined BCCK to BC Giao Mapping ===")
for bcck, bcs in sorted(bcck_bcs.items()):
    print(f"\nBCCK: {bcck}")
    for bc in sorted(bcs):
        print(f"  - {bc}")

if missing:
    print(f"\n⚠️ Missing mappings for {len(missing)} rows:")
    for dist, ward, bcck in missing[:10]:
        print(f"  District: {dist} | Ward: {ward} | BCCK: {bcck}")
