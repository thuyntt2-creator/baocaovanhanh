import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\AOP_Hang_NTB_T7-T12_2026_calculated.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['Kênh & nhu cầu']

wb_eval = openpyxl.load_workbook(file_path, data_only=True)
sheet_eval = wb_eval['Kênh & nhu cầu']

print("=== Formulas and values in 'Kênh & nhu cầu' ===")
for r in [8, 9]:
    label = sheet.cell(r, 1).value or sheet.cell(r, 2).value
    row_formulas = [sheet.cell(r, c).value for c in range(2, 9)]
    row_vals = [sheet_eval.cell(r, c).value for c in range(2, 9)]
    print(f"Row {r:2d} | Label: '{label}'")
    print(f"  Formulas: {row_formulas}")
    print(f"  Values  : {row_vals}")

wb_eval.close()
wb.close()
