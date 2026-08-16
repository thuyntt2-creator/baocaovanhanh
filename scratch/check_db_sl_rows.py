import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['Định biên & Sản lượng']

print("=== Sheet 'Định biên & Sản lượng' - rows 70-85 ===")
for r in range(70, 86):
    row_label = sheet.cell(r, 2).value or sheet.cell(r, 1).value or f"Row {r}"
    c_val = sheet[f"C{r}"].value  # column C (usually T7 or formula/base)
    d_val = sheet[f"D{r}"].value  # column D
    print(f"Row {r:2d} | {str(row_label)[:35]:<35} | C: {c_val} | D: {d_val}")
wb.close()
