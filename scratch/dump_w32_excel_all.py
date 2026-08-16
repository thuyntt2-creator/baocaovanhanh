import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r'C:\Users\lap4all\Downloads\BaoCao_Tuan_NTB_W32_2026.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

def dump_sheet(sheetname):
    ws = wb[sheetname]
    print(f"\n==========================================")
    print(f"=== FULL DUMP SHEET: {sheetname} ({ws.max_row}x{ws.max_column}) ===")
    print(f"==========================================")
    for r in range(1, ws.max_row+1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
        if any(v is not None for v in vals):
            row_str = [str(v) if v is not None else "" for v in vals]
            # print non-empty trailing
            while row_str and row_str[-1] == "":
                row_str.pop()
            print(f"R{r:2d}:", row_str[:12])

for name in wb.sheetnames:
    dump_sheet(name)
