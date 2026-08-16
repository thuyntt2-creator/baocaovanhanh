import openpyxl, sys
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Quy_Hoach_NTB_2026"

# Styles
font_header = Font(name="Calibri", size=11, bold=True, color="000000")
fill_header = PatternFill(start_color="F7A059", end_color="F7A059", fill_type="solid") # Orange header matching screenshot

font_dvhc = Font(name="Calibri", size=10.5, bold=True, color="000000")
font_normal = Font(name="Calibri", size=10, color="000000")
font_buucuc = Font(name="Calibri", size=10, bold=True, color="003399") # Dark Blue font

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

headers = ["ĐVHC Mới", "Các Xã / Phường Cũ sáp nhập", "Bưu cục Cover", "Phương án & Ghi chú"]

# Data per province matching user screenshot perfectly
province_tables = [
    {
        "province": "TỈNH LÂM ĐỒNG",
        "rows": [
            ["Phường 1 Bảo Lộc", "P.1 + P. Lộc Phát + X. Lộc Thanh", "BC (LDO) B'Lao Mới", "Gộp 100% ➔ Đóng BC 1 Bảo Lộc cũ"],
            ["Phường B'Lao", "P. Lộc Sơn + P. B'Lao + X. Lộc Nga", "BC (LDO) B'Lao Mới", "Gộp sản lượng về BC mới"],
            ["Phường 2 Bảo Lộc", "P.2 + X. Đạm Bri + X. Lộc Tân", "BC (LDO) 3 Bảo Lộc", "Gộp về BC 3 Bảo Lộc"],
            ["Cụm Đơn Dương (Bắc)", "Lạc Lâm + Lạc Xuân + D'Ran + Ka Đô", "BC (LDO) Lạc Xuân (MỚI)", "Mở mới BC (vùng màu vàng)"],
            ["Cụm Đơn Dương (Nam)", "Đạ Ròn + TT. Thạnh Mỹ + Tu Tra + Ka Đơn...", "BC Nghĩa Đức (GỐC)", "Giữ BC gốc Nghĩa Đức cover 6 xã"],
            ["Xã Di Linh", "TT. Di Linh + X. Liên Đầm + Tân Châu...", "BC Hàng Vừa Di Linh (MỚI) & BC Di Linh", "Tách BC tại Đinh Trang Thượng"],
            ["Xã Đam Rông 4", "X. Đạ Tông + X. Đạ Long + X. Đưng KNớ", "BC Đam Rông 3 & BC Lang Biang 1", "Giữ 2 BC (địa hình xa >50km)"],
            ["Xã Đức Trọng", "TT. Liên Nghĩa + X. Phú Hội", "BC Đức Trọng 2 & BC Đức Trọng 1", "Giữ 2 BC (địa bàn xa >15km)"]
        ]
    },
    {
        "province": "TỈNH KHÁNH HÒA",
        "rows": [
            ["Phường Nam Nha Trang", "P. Vĩnh Trường + P. Phước Long + X. Vĩnh Hiệp", "BC Nam Nha Trang 1 Mới & BC Nam Nha Trang 5", "Gộp về BC mới (Đóng BC 3)"],
            ["Phường Nha Trang", "P. Vĩnh Nguyên + P. Phước Tiến + P. Phước Tân", "BC (KHO) Nha Trang", "Gộp phân vùng về BC Nha Trang"],
            ["Phường Tây Nha Trang", "P. Vĩnh Hải + X. Vĩnh Ngọc + X. Vĩnh Thạnh", "BC (KHO) Tây Nha Trang", "Gộp về BC Tây Nha Trang"],
            ["Xã Diên Khánh", "TT. Diên Khánh + X. Diên An + X. Diên Toàn", "BC (KHO) Diên Khánh 2", "Gộp về DK2 (Đóng BC DK1 <40m²)"],
            ["Xã Vạn Thắng", "X. Vạn Thắng + X. Vạn Bình", "BC (KHO) Tu Bông", "Gộp dồn sản lượng về BC Tu Bông"],
            ["Cụm TP. Cam Ranh", "Ba Ngòi + Cam Bình + Cam Lập + Cam Phước Đông...", "BC Nam Cam Ranh (MỚI) & BC Cam Linh", "Tách mới BC cover 6 xã phía Nam"]
        ]
    },
    {
        "province": "TỈNH NINH THUẬN",
        "rows": [
            ["Phường Ninh Chử", "TT. Khánh Hải + P. Văn Hải", "BC (NTH) Ninh Chử", "Di dời kho về trung tâm ĐVHC mới"],
            ["Phường Phan Rang", "P. Kinh Dinh + P. Phủ Hà + P. Đạo Long + P. Đài Sơn", "BC (NTH) Phan Rang", "Gộp phân vùng về BC Phan Rang"],
            ["Xã Ninh Hải", "X. Phương Hải + X. Tri Hải + X. Bắc Sơn", "BC (NTH) Ninh Chử", "BC Ninh Chử cover 75.7% sản lượng"],
            ["Xã Phước Dinh", "X. An Hải + X. Phước Dinh + P. Đông Hải", "BC Đông Hải (MỚI) & BC Phước Dinh", "Mở BC Đông Hải gánh đơn hải sản"]
        ]
    },
    {
        "province": "TỈNH BÌNH THUẬN",
        "rows": [
            ["Phường La Gi", "P. Tân Thiện + P. Tân An + P. Bình Tân + X. Tân Bình", "BC (BTH) Phước Hội & BC (BTH) Tân Hải", "Phước Hội cover 3 phường, Tân Hải cover X. Tân Bình"],
            ["Phường Bình Thuận", "P. Phú Tài + X. Phong Nẫm + X. Hàm Hiệp", "BC (BTH) Hàm Thắng & BC (BTH) Hàm Liêm", "Gộp địa bàn về BC Hàm Thắng & Hàm Liêm"],
            ["Phường Phan Thiết", "P. Phú Trinh + P. Lạc Đạo + P. Bình Hưng", "BC (BTH) Hàm Thắng & BC (BTH) Phú Thủy", "Gộp địa bàn về BC Hàm Thắng & Phú Thủy"],
            ["Phường Hàm Thắng", "P. Xuân An + TT. Phú Long + X. Hàm Thắng", "BC (BTH) Phú Thủy & BC (BTH) Hàm Liêm", "Gộp về BC Phú Thủy & BC Hàm Liêm"],
            ["Xã Phan Rí Cửa", "TT. Phan Rí Cửa + X. Chí Công + X. Hòa Minh", "BC Phan Rí Cửa & BC Liên Hương", "Giữ 2 BC cover song song"],
            ["Xã Tân Thành", "X. Tân Thuận + X. Tân Thành + X. Thuận Quý", "BC Tân Hải & BC Hàm Thuận Nam", "Giữ 2 BC (trục bờ biển dài >20km)"],
            ["Cụm Nam Thành", "X. Nam Thành + X. Nghị Đức", "BC (BTH) Nam Thành (MỚI)", "Mở mới xóa chạy chéo tuyến xa"]
        ]
    },
    {
        "province": "TỈNH ĐẮK NÔNG",
        "rows": [
            ["Phường Bắc Gia Nghĩa", "P. Nghĩa Đức + P. Nghĩa Thành + P. Quảng Thành + X. Đắc Ha", "BC (DNO) Bắc Gia Nghĩa", "Gộp địa bàn về BC Bắc Gia Nghĩa"],
            ["Xã Đắc Sắc", "X. Đắc Sắc + X. Nam Xuân + X. Long Sơn", "BC Đức Lập & BC Krông Nô", "Giữ 2 BC (NV Krông Nô không đi)"],
            ["Xã Đức An", "TT. Đức An + X. Đắc N'Drung + X. Nam Bình", "BC Đức An & BC Trường Xuân", "Giữ 2 BC (Đắk N'Drung gần TX hơn)"],
            ["Xã Tà Đùng", "X. Đắc Som + X. Đắc R'Măng", "BC Quảng Khê & BC Quảng Sơn", "Giữ 2 BC (đường đèo dốc >35km)"],
            ["Xã Quảng Tân", "X. Quảng Tân + X. Đắc Ngo", "BC Kiến Đức & BC Quảng Tín", "Giữ 2 BC cover song song"]
        ]
    }
]

current_row = 1

for t_data in province_tables:
    # Header Row
    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=h_text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    
    ws.row_dimensions[current_row].height = 26
    current_row += 1
    
    # Data Rows
    for r_data in t_data["rows"]:
        for col_idx, val in enumerate(r_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = align_left
            if col_idx == 1:
                cell.font = font_dvhc
            elif col_idx == 3:
                cell.font = font_buucuc
            else:
                cell.font = font_normal
        ws.row_dimensions[current_row].height = 22
        current_row += 1
    
    # Blank separator row
    current_row += 1

# Column Widths
ws.column_dimensions['A'].width = 24
ws.column_dimensions['B'].width = 44
ws.column_dimensions['C'].width = 38
ws.column_dimensions['D'].width = 38

out_xlsx = r'C:\Users\lap4all\Downloads\Quy_Hoach_Mang_Luoi_NTB_2026_Theo_Tinh.xlsx'
wb.save(out_xlsx)
print(f'Successfully generated Excel file matching screenshot: {out_xlsx}')
