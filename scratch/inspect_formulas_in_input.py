import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['NTB – Input']

print("=== Formulas in NTB - Input ===")
for r in range(1, 48):
    for c in range(1, 11):
        val = sheet.cell(r, c).value
        if isinstance(val, str) and val.startswith('='):
            col_letter = openpyxl.utils.get_column_letter(c)
            print(f"Cell {col_letter}{r}: {val}")

wb.close()
