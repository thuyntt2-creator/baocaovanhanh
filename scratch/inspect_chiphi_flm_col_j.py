import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['Chi phí FLM']

print("=== Sheet 'Chi phí FLM' col J (T7) values ===")
for r in range(15, 37):
    label = sheet.cell(r, 2).value or sheet.cell(r, 1).value or f"Row {r}"
    val = sheet.cell(r, 10).value  # col J is 10
    print(f"Row {r:2d} | {str(label)[:35]:<35} | J (T7): {val}")

wb.close()
