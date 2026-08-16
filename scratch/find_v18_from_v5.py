import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_v18 = r"C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v18.xlsx"
file_v5 = r"C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v5.xlsx"

wb_v18 = openpyxl.load_workbook(file_v18, data_only=True)
wb_v5 = openpyxl.load_workbook(file_v5, data_only=True)

for m in [7, 8, 9, 10, 11, 12]:
    print(f"\n==========================================")
    print(f"MONTH: T{m}")
    print(f"==========================================")
    
    sheet_18 = wb_v18[f'Forecast T{m}']
    sheet_5 = wb_v5[f'Forecast T{m}']
    
    for r in range(4, 8):
        name = sheet_18.cell(r, 1).value
        print(f"BC: {name}")
        
        mismatches_round = []
        mismatches_floor = []
        mismatches_ceil = []
        
        for c in range(2, sheet_18.max_column + 1):
            val_18 = sheet_18.cell(r, c).value
            val_5 = sheet_5.cell(r, c).value
            if val_18 is not None and val_5 is not None:
                # Test round
                v_round = round(val_5 / 55) * 55
                v_floor = (val_5 // 55) * 55
                v_ceil = -(-val_5 // 55) * 55 # ceil division
                
                if v_round != val_18:
                    mismatches_round.append((c - 1, val_5, val_18, v_round))
                if v_floor != val_18:
                    mismatches_floor.append((c - 1, val_5, val_18, v_floor))
                if v_ceil != val_18:
                    mismatches_ceil.append((c - 1, val_5, val_18, v_ceil))
                    
        print(f"  Round method: mismatches = {len(mismatches_round)}")
        print(f"  Floor method: mismatches = {len(mismatches_floor)}")
        print(f"  Ceil method:  mismatches = {len(mismatches_ceil)}")
        
        if len(mismatches_round) > 0 and len(mismatches_round) < 10:
            print("  Mismatches for Round:")
            for day, v5, v18, vr in mismatches_round:
                print(f"    Day {day}: v5={v5} | v18={v18} | round={vr}")
