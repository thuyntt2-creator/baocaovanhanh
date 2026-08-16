import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

filepath = r"C:\Users\lap4all\Downloads\Telegram Desktop\7. NTB_2026.xlsx"
wb = openpyxl.load_workbook(filepath, data_only=True)
sheet = wb['7. NTB_2026']

cols_to_print = [sheet.cell(1, c).value or sheet.cell(2, c).value or f"Col {c}" for c in range(1, sheet.max_column + 1)]
print("=== Column Headers in 7. NTB_2026 ===")
for c_idx, c_name in enumerate(cols_to_print):
    print(f"Col {c_idx+1}: {c_name}")

print("\n=== Values ===")
for r in [1, 2, 3, 4, 5, 10, 32, 33, 34, 35]:
    row_vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
    print(f"Row {r:3d} | {str(row_vals[0])[:35]:<35} | {row_vals[1:18]}")
wb.close()
