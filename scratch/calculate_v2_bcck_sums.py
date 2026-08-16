import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

v2_path = r"C:\Users\lap4all\Downloads\[V2] AOP_Hang_NTB_T7-T12_2026.xlsx"
wb = openpyxl.load_workbook(v2_path, data_only=True)
sheet = wb['0.3 Bưu cục Detail']

print("=== Summing V2 Bưu Cục volumes for BCCK ===")
# List of all bưu cục rows
bcs = []
for r in range(3, sheet.max_row + 1):
    bc_name = sheet.cell(r, 2).value
    t12_vol = sheet.cell(r, 9).value # Column I is T12
    if bc_name is not None and t12_vol is not None:
        daily_t12 = round(t12_vol / 30.0)
        bcs.append({'row': r, 'name': bc_name.strip(), 't12_monthly': t12_vol, 't12_daily': daily_t12})

# Let's write a grouping function based on keywords
groups = {
    'BCCK Nha Trang': [],
    'BCCK Di Linh': [],
    'BCCK Đơn Dương': [],
    'BCKK Đức Linh-Bình Thuận': []
}

# Nha Trang: contains "Nha Trang" or "Yết Kiêu" or "Lê Hồng Phong" (if TP.Nha Trang) or "Phước Đồng" or "Diên Khánh"
# Di Linh: contains "Di Linh" or "Lâm Hà" or "Đinh Văn" or "Tân Hà" or "Hòa Ninh" or "Ba Đình" or "Lộc Thắng" or "Bảo Lâm" or "Bảo Lộc"
# Đơn Dương: contains "Đơn Dương" or "Đức Trọng" or "Thạnh Mỹ" or "Liên Nghĩa" or "Hiệp Thạnh"
# Đức Linh: contains "Đức Linh" or "Nam Chính" or "Đức Tài"

for bc in bcs:
    name = bc['name']
    if any(k in name for k in ["Nha Trang", "Yết Kiêu", "Phước Đồng", "Hà Quang"]):
        groups['BCCK Nha Trang'].append(bc)
    elif "Lê Hồng Phong" in name and "Nha Trang" in name:
        groups['BCCK Nha Trang'].append(bc)
    elif any(k in name for k in ["Di Linh", "Lâm Hà", "Đinh Văn", "Tân Hà", "Hòa Ninh", "Ba Đình"]):
        groups['BCCK Di Linh'].append(bc)
    elif any(k in name for k in ["Đơn Dương", "Đức Trọng", "Thạnh Mỹ", "Liên Nghĩa", "Hiệp Thạnh"]):
        groups['BCCK Đơn Dương'].append(bc)
    elif any(k in name for k in ["Đức Linh", "Nam Chính", "Đức Tài"]):
        groups['BCKK Đức Linh-Bình Thuận'].append(bc)

for gname, items in groups.items():
    print(f"\n{gname}:")
    total_monthly = sum(i['t12_monthly'] for i in items)
    total_daily = sum(i['t12_daily'] for i in items)
    formula_cells = "+".join(f"'0.3 Bưu cục Detail'!J{i['row']}" for i in items)
    print(f"  Count: {len(items)}")
    print(f"  Total Monthly T12: {total_monthly:.1f}")
    print(f"  Total Daily T12: {total_daily:.1f} (Average/30: {total_monthly/30.0:.1f})")
    print(f"  Formula: {formula_cells}")
    for i in items:
        print(f"    - Row {i['row']}: {i['name']} | Monthly: {i['t12_monthly']:.1f} | Daily: {i['t12_daily']}")

