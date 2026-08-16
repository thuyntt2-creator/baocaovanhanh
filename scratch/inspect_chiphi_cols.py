import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['Chi phí FLM']

print("=== Chi phí FLM Header (Row 5) ===")
row_5 = [sheet.cell(5, c).value for c in range(1, sheet.max_column + 1)]
for idx, val in enumerate(row_5, 1):
    print(f"Col {idx} ({openpyxl.utils.get_column_letter(idx)}): {val}")

print("\n=== Chi phí FLM Row 7 (Đơn GIAO trong tháng) ===")
row_7 = [sheet.cell(7, c).value for c in range(1, sheet.max_column + 1)]
for idx, val in enumerate(row_7, 1):
    print(f"Col {idx} ({openpyxl.utils.get_column_letter(idx)}): {val}")

wb.close()
