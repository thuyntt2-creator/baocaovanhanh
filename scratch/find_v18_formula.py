import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v18.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

# Let's test the formula for T9:
# Trucks_9[i] = round(Trucks_8[i] * factor) or ceil(Trucks_8[i] * factor)
# We will search for a factor in range [1.5, 1.8] with step 0.001

sheet_prev = wb['Forecast T8']
sheet_curr = wb['Forecast T9']

for r in range(4, 8):
    name = sheet_curr.cell(r, 1).value
    prev_vals = [sheet_prev.cell(r, c).value for c in range(2, 32) if sheet_prev.cell(r, c).value is not None]
    curr_vals = [sheet_curr.cell(r, c).value for c in range(2, 32) if sheet_curr.cell(r, c).value is not None]
    
    prev_trucks = [v / 55 for v in prev_vals]
    curr_trucks = [v / 55 for v in curr_vals]
    
    print(f"\nEvaluating BC: {name}")
    best_factor = None
    best_matches = 0
    # Try different factors
    for f in [x * 0.0001 for x in range(15000, 18000)]:
        matches_round = sum(1 for p, c in zip(prev_trucks, curr_trucks) if round(p * f) == c)
        if matches_round > best_matches:
            best_matches = matches_round
            best_factor = f
            
    print(f"  Best round factor: {best_factor:.4f} (Matches: {best_matches}/{len(curr_vals)})")
    
    # Let's see the differences for best round factor
    if best_factor:
        for idx, (p, c) in enumerate(zip(prev_trucks, curr_trucks)):
            calc = round(p * best_factor)
            if calc != c:
                print(f"    Day {idx+1}: Aug_trucks={p} | calc={calc} | actual={c} | diff={c - calc}")
