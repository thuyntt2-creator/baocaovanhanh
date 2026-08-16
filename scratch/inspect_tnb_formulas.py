import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

downloads_dir = r"C:\Users\lap4all\Downloads"
tnb_path = os.path.join(downloads_dir, "TNB_Input_FLM_CRC.xlsx")

if os.path.exists(tnb_path):
    print("Found TNB_Input_FLM_CRC.xlsx")
    wb = openpyxl.load_workbook(tnb_path, data_only=False)
    # List sheets
    print("Sheets:", wb.sheetnames)
    # Check if there is an input sheet
    sheet_name = [s for s in wb.sheetnames if "Input" in s or "input" in s or "TNB" in s]
    if sheet_name:
        sheet = wb[sheet_name[0]]
        print(f"\n=== Formulas in {sheet_name[0]} rows 41-45 ===")
        cols = ['D', 'E', 'F', 'G', 'H', 'I']
        for r in range(41, 46):
            label = sheet.cell(r, 2).value or sheet.cell(r, 1).value
            print(f"Row {r:2d} | {label}")
            for c in cols:
                print(f"  {c}: {sheet[f'{c}{r}'].value}")
    wb.close()
else:
    print("TNB_Input_FLM_CRC.xlsx not found in Downloads")
