import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['Chi phí FLM']

print("=== Formulas/Values in 'Chi phí FLM' sheet ===")
for r in range(25, 45):
    row_vals = []
    for c in range(1, 17):
        cell = sheet.cell(r, c)
        val = cell.value
        if isinstance(val, str) and val.startswith('='):
            row_vals.append(f"F:{val}")
        else:
            row_vals.append(val)
    if any(v is not None for v in row_vals):
        row_lbl = sheet.cell(r, 1).value or sheet.cell(r, 2).value or f"Row {r}"
        print(f"Row {r:2d} | {str(row_lbl)[:45]:<45} | {row_vals[2:]}")

wb.close()
