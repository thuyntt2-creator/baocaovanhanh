import openpyxl, sys, shutil
from openpyxl.styles import Font, PatternFill, Alignment
sys.stdout.reconfigure(encoding='utf-8')

in_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v6.xlsx'
out_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v8.xlsx'

print("Đang tạo Bảng Định biên & Bảng Chi Phí (V8)...")
shutil.copy(in_path, out_path)
wb = openpyxl.load_workbook(out_path)
sheet = wb['Nguồn lực & chi phí']

# Styles
bold_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
center_align = Alignment(horizontal='center', vertical='center')
money_format = '#,##0'

months = ['T7', 'T8', 'T9', 'T10', 'T11', 'T12']
cols = ['B', 'C', 'D', 'E', 'F', 'G']
hub_refs = ["='Kế hoạch T7'!$A$11", "='Kế hoạch T7'!$A$21", "='Kế hoạch T7'!$A$31", "='Kế hoạch T7'!$A$41"]

r = 21

# --- BLOCK 1: NGƯỜI GIAO ---
sheet.cell(r, 1).value = "I. ĐỊNH BIÊN NGƯỜI GIAO (TÀI XẾ + PHỤ XE) THEO KHO/BC"
sheet.cell(r, 1).font = Font(bold=True, color="FF0000")
r += 1

sheet.cell(r, 1).value = "Chỉ tiêu"
for i, m in enumerate(months):
    c = sheet.cell(r, i + 2)
    c.value = m; c.font = bold_font; c.fill = header_fill; c.alignment = center_align
sheet.cell(r, 1).font = bold_font; sheet.cell(r, 1).fill = header_fill

r += 1
for i in range(4):
    sheet.cell(r, 1).value = hub_refs[i]
    for c_idx in range(6):
        col_letter = cols[c_idx]
        formula = f"=ROUNDUP('Tổng hợp xe'!{col_letter}{i+6}*'1. Thông số'!$B$20,0)"
        sheet.cell(r, c_idx + 2).value = formula
        sheet.cell(r, c_idx + 2).alignment = center_align
    r += 1


# --- BLOCK 2: XỬ LÝ KHO ---
r += 1
sheet.cell(r, 1).value = "II. ĐỊNH BIÊN NHÂN VIÊN XỬ LÝ KHO (SOX) THEO KHO/BC"
sheet.cell(r, 1).font = Font(bold=True, color="FF0000")
r += 1

sheet.cell(r, 1).value = "Chỉ tiêu"
for i, m in enumerate(months):
    c = sheet.cell(r, i + 2)
    c.value = m; c.font = bold_font; c.fill = header_fill; c.alignment = center_align
sheet.cell(r, 1).font = bold_font; sheet.cell(r, 1).fill = header_fill

r += 1
block2_start_r = r
for i in range(4):
    sheet.cell(r, 1).value = hub_refs[i]
    for c_idx, m in enumerate(months):
        end_col = 'AF' if m in ['T7', 'T8', 'T10', 'T12'] else 'AE'
        fc_row = i + 4 
        formula = f"=ROUNDUP(AVERAGE('Forecast {m}'!B{fc_row}:{end_col}{fc_row})/'1. Thông số'!$B$19,0)"
        sheet.cell(r, c_idx + 2).value = formula
        sheet.cell(r, c_idx + 2).alignment = center_align
    r += 1


# --- BLOCK 3: CHI PHÍ ---
r += 1
sheet.cell(r, 1).value = "III. TỔNG CHI PHÍ (THUÊ XE + LƯƠNG KHO) THEO KHO/BC (VNĐ)"
sheet.cell(r, 1).font = Font(bold=True, color="FF0000")
r += 1

sheet.cell(r, 1).value = "Chỉ tiêu"
for i, m in enumerate(months):
    c = sheet.cell(r, i + 2)
    c.value = m; c.font = bold_font; c.fill = header_fill; c.alignment = center_align
sheet.cell(r, 1).font = bold_font; sheet.cell(r, 1).fill = header_fill

r += 1
for i in range(4):
    sheet.cell(r, 1).value = hub_refs[i]
    kh_row = 18 + i*10 # 18, 28, 38, 48
    sox_r = block2_start_r + i # Dòng chứa nhân sự kho
    
    for c_idx, m in enumerate(months):
        col_letter = cols[c_idx]
        end_col = 'AF' if m in ['T7', 'T8', 'T10', 'T12'] else 'AE'
        
        # Tiền xe: = ROUND(AVERAGE(xe xếp thực tế),0) * 1.2M * số ngày (dòng 4)
        xe_logic = f"ROUND(AVERAGE('Kế hoạch {m}'!B{kh_row}:{end_col}{kh_row}),0)*'1. Thông số'!$B$23*{col_letter}$4"
        
        # Tiền lương: = Người xử lý kho (ô ở block 2) * 15M
        luong_logic = f"{col_letter}{sox_r}*'1. Thông số'!$B$24"
        
        formula = f"={xe_logic} + {luong_logic}"
        
        c = sheet.cell(r, c_idx + 2)
        c.value = formula
        c.number_format = money_format
    r += 1

wb.save(out_path)
print(f"Xong! Đã lưu: {out_path}")
