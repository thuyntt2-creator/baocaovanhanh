import openpyxl
import os
import sys
import glob

sys.stdout.reconfigure(encoding='utf-8')

downloads_dir = r"C:\Users\lap4all\Downloads"
all_files = glob.glob(os.path.join(downloads_dir, "*.xlsx"))

# Lọc các file aop hoặc ntb và dung lượng < 10MB
excel_files = []
for p in all_files:
    fname = os.path.basename(p)
    if fname.startswith("~$"):
        continue
    size_mb = os.path.getsize(p) / (1024 * 1024)
    name_lower = fname.lower()
    if ("aop" in name_lower or "ntb" in name_lower) and size_mb < 10:
        excel_files.append((p, size_mb))

# Thêm cả các file trong thư mục làm việc chính nếu có
workspace_dir = r"c:\Users\lap4all\Documents\Auto report"
for p in glob.glob(os.path.join(workspace_dir, "*.xlsx")):
    fname = os.path.basename(p)
    if not fname.startswith("~$"):
        excel_files.append((p, os.path.getsize(p) / (1024 * 1024)))

print(f"=== ĐANG QUÉT NHANH {len(excel_files)} FILE EXCEL ===")

targets = [715, 690, 1746, 126, 76]

for path, size in excel_files:
    fname = os.path.basename(path)
    # print(f"Quét {fname} ({size:.2f} MB)...")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        continue
        
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        max_r = min(sheet.max_row, 150)
        max_c = min(sheet.max_column, 25)
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                val = sheet.cell(r, c).value
                if val is not None:
                    # Kiểm tra khớp target
                    for target in targets:
                        # khớp số nguyên chính xác
                        if (isinstance(val, (int, float)) and abs(val - target) < 0.1) or \
                           (isinstance(val, (int, float)) and abs(val/1e6 - target) < 0.1) or \
                           (isinstance(val, (int, float)) and abs(val/1e3 - target) < 0.1):
                            print(f"[KHỚP] File: {fname} | Sheet: {sheet_name} | Ô {openpyxl.utils.get_column_letter(c)}{r}: Giá trị = {val} (Khớp với target {target})")
                        elif str(target) in str(val):
                            # kiểm tra chuỗi
                            if len(str(val)) < 15:
                                print(f"[KHỚP CHUỖI] File: {fname} | Sheet: {sheet_name} | Ô {openpyxl.utils.get_column_letter(c)}{r}: Giá trị = {val} (Chứa target {target})")
