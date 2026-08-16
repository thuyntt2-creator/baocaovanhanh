import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\V3_AOP_Hang_NTB_T7-T12_2026 mới.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

targets = [198219, 196204, 204379, 225298, 234987, 261430]

print("=== Searching for targets in V3 ===")
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(r, c).value
            if val is not None:
                # check if val is close to any target
                try:
                    num = float(val)
                    for t in targets:
                        if abs(num - t) < 5:
                            print(f"Sheet '{sheet_name}' cell {openpyxl.utils.get_column_letter(c)}{r}: {val} (target: {t})")
                except (ValueError, TypeError):
                    pass

wb.close()
