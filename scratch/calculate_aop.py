import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

template_path = r"C:\Users\lap4all\Downloads\AOP_Hang_NTB_T7-T12_2026.xlsx"
v2_path = r"C:\Users\lap4all\Downloads\[V2] AOP_Hang_NTB_T7-T12_2026.xlsx"
output_path = r"C:\Users\lap4all\Downloads\AOP_Hang_NTB_T7-T12_2026_calculated.xlsx"

def run_calculation():
    print("📖 Loading workbooks...")
    wb = openpyxl.load_workbook(template_path, data_only=False)
    wb_v2 = openpyxl.load_workbook(v2_path, data_only=True)

    print("✅ Copying new Volume Giao and Volume Lấy from [V2]...")
    for sheet_name in ['Volume Giao', 'Volume Lấy']:
        ws_template = wb[sheet_name]
        ws_v2 = wb_v2[sheet_name]
        
        # Clear the template sheet
        ws_template.delete_rows(1, ws_template.max_row + 1)
        
        # Copy values from [V2]
        for r in range(1, ws_v2.max_row + 1):
            for c in range(1, ws_v2.max_column + 1):
                val = ws_v2.cell(r, c).value
                ws_template.cell(r, c).value = val
        print(f"  - Copied {ws_v2.max_row} rows for {sheet_name}")

    print("✅ Copying Kênh & nhu cầu monthly totals...")
    ws_k_t = wb['Kênh & nhu cầu']
    ws_k_v2 = wb_v2['Kênh & nhu cầu']
    # Copy T7-T12 (Cols B to G) for rows 4 (Hàng vừa), 5 (Hàng nặng), and 6 (B2B)
    for r in [4, 5, 6]:
        for c in range(2, 8): # B to G
            val = ws_k_v2.cell(r, c).value
            ws_k_t.cell(r, c).value = val
    print("  - Copied Kênh & nhu cầu values.")

    print("✅ Loading [V2] 0.3 Bưu cục Detail data...")
    ws_det_v2 = wb_v2['0.3 Bưu cục Detail']
    # Read rows 3 to 84 (data rows)
    v2_rows_data = []
    for r in range(3, 85):
        stt = ws_det_v2.cell(r, 1).value
        bc_name = ws_det_v2.cell(r, 2).value
        bc_type = ws_det_v2.cell(r, 3).value
        t7 = ws_det_v2.cell(r, 4).value
        t8 = ws_det_v2.cell(r, 5).value
        t9 = ws_det_v2.cell(r, 6).value
        t10 = ws_det_v2.cell(r, 7).value
        t11 = ws_det_v2.cell(r, 8).value
        t12 = ws_det_v2.cell(r, 9).value
        if bc_name:
            v2_rows_data.append([stt, bc_name, bc_type, t7, t8, t9, t10, t11, t12])

    print(f"  - Found {len(v2_rows_data)} bưu cục in [V2].")

    print("✅ Overwriting 0.3 Bưu cục Detail sheet in Template...")
    ws_det_t = wb['0.3 Bưu cục Detail']
    # Clean up template data rows (delete rows from 3 to 1000 to clear any residual data)
    ws_det_t.delete_rows(3, ws_det_t.max_row + 10)
    
    # Write the new bưu cục data and formulas
    for i, row_data in enumerate(v2_rows_data):
        r_idx = 3 + i
        # Write A to I
        for c_idx, val in enumerate(row_data, start=1):
            ws_det_t.cell(r_idx, c_idx).value = val
            
        # Write formulas for J, K, L, M
        ws_det_t.cell(r_idx, 10).value = f"=IF(I{r_idx}=\"\",\"\",ROUND(I{r_idx}/'1. Thông số'!$B$17,0))"
        ws_det_t.cell(r_idx, 11).value = f"=IF(J{r_idx}=\"\",\"\",IF(J{r_idx}>='1. Thông số'!$B$18,\"Đủ mật độ — nên mở BCCK\",\"Không đủ — giao gộp\"))"
        ws_det_t.cell(r_idx, 12).value = f"=IF(J{r_idx}=\"\",\"\",ROUNDUP(J{r_idx}/'1. Thông số'!$B$7,0))"
        ws_det_t.cell(r_idx, 13).value = f"=IF(J{r_idx}=\"\",\"\",IF(J{r_idx}>='1. Thông số'!$B$18,\"Mở/giữ BCCK chuyên: kho ~\"&L{r_idx}&\" m² + xe van + xe tải; điều phối tập trung\",\"Không mở BCCK: gộp về BCCK gần nhất / BC thường + xe máy gánh hàng vừa + van gom tuyến liên xã; ≥20kg dồn GXT\"))"

    # Add the footer note at row 85 (last_row + 1)
    last_row = 3 + len(v2_rows_data) - 1
    ws_det_t.cell(last_row + 1, 1).value = "Nguồn BC-level: PSBBA 3_KQ_BC_Detail. Ô D..I = nạp sẵn; cột phân loại tự tính theo ngưỡng ở '1. Thông số'."
    print(f"  - Wrote {len(v2_rows_data)} data rows to 0.3 Bưu cục Detail sheet (up to row {last_row}).")

    print("✅ Updating sheet Mật độ & phương án summary formulas...")
    ws_md = wb['Mật độ & phương án']
    ws_md['B4'].value = f"=COUNTIF('0.3 Bưu cục Detail'!$K$3:$K${last_row},\"Đủ*\")"
    ws_md['B5'].value = f"=COUNTIF('0.3 Bưu cục Detail'!$K$3:$K${last_row},\"Không*\")"
    ws_md['B6'].value = f"=SUMIF('0.3 Bưu cục Detail'!$K$3:$K${last_row},\"Đủ*\",'0.3 Bưu cục Detail'!$I$3:$I${last_row})"
    ws_md['B7'].value = f"=SUMIF('0.3 Bưu cục Detail'!$K$3:$K${last_row},\"Không*\",'0.3 Bưu cục Detail'!$I$3:$I${last_row})"
    print("  - Updated Mật độ & phương án summary formulas.")

    print("✅ Building dynamic sum formulas for Mặt bằng...")
    # Bưu cục groups for summing
    # Let's map row indexes dynamically based on written bưu cục names
    bc_to_row = {}
    for i, row_data in enumerate(v2_rows_data):
        r_idx = 3 + i
        bc_name = row_data[1].strip()
        bc_to_row[bc_name.lower()] = r_idx

    # Group bưu cục lists
    nt_bcs = [
        "Bưu Cục 06 Lê Hồng Phong-TP.Nha Trang-Khánh Hòa",
        "Bưu Cục 195 Đường 2/4-Nha Trang-Khánh Hòa",
        "Bưu Cục 229 Phước Long-Nam Nha Trang-Khánh Hòa",
        "Bưu Cục 40A Yết Kiêu-Nha Trang-Khánh Hòa",
        "Bưu Cục 466 Đường 23/10-Nha Trang-Khánh Hòa",
        "Bưu Cục Đường 35 Hà Quang 1-Xã Nam Nha Trang-Khánh Hòa",
        "Bưu Cục Phước Đồng-Nha Trang-Khánh Hoà"
    ]
    dl_bcs = [
        "Bưu Cục 1322 Hùng Vương-Di Linh-Lâm Đồng",
        "(LDO) Đinh Văn Lâm Hà",
        "(LDO) Tân Hà Lâm Hà",
        "Bưu Cục 231 Thôn 1-Xã Hòa Ninh-Lâm Đồng",
        "(LDO) Nam Ban Lâm Hà"
    ]
    dd_bcs = [
        "(LDO) Đơn Dương",
        "(LDO) Hiệp Thạnh"
    ]
    dk_bcs = [
        "(BTH) Đức Linh"
    ]

    def build_sum_formula(bcs_list):
        rows = []
        for name in bcs_list:
            r = bc_to_row.get(name.strip().lower())
            if r is not None:
                rows.append(r)
            else:
                print(f"  ⚠️ Warning: Bưu cục '{name}' not found in data rows!")
        if not rows:
            return "0"
        return "+".join(f"'0.3 Bưu cục Detail'!J{r}" for r in rows)

    nt_formula = build_sum_formula(nt_bcs)
    dl_formula = build_sum_formula(dl_bcs)
    dd_formula = build_sum_formula(dd_bcs)
    dk_formula = build_sum_formula(dk_bcs)

    ws_mb = wb['Mặt bằng']
    
    # Row 3: BCCK Nha Trang
    ws_mb.cell(3, 2).value = "BCCK Nha Trang"
    ws_mb.cell(3, 3).value = f"={nt_formula}"
    ws_mb.cell(3, 4).value = "=IFERROR(ROUNDUP(C3/'1. Thông số'!$B$7,0),0)"
    ws_mb.cell(3, 5).value = "=D3*'1. Thông số'!$B$14"
    ws_mb.cell(3, 6).value = None # Diện tích hiện tại
    ws_mb.cell(3, 7).value = "=IFERROR(D3-F3,0)"
    ws_mb.cell(3, 8).value = "Mở thêm BCCK / Mở rộng BC  hiện tại"
    ws_mb.cell(3, 9).value = "Gộp cả toàn NT"
    ws_mb.cell(3, 10).value = 13.0
    ws_mb.cell(3, 11).value = 1.0
    ws_mb.cell(3, 12).value = 1.0
    ws_mb.cell(3, 13).value = "=SUM(J3:L3)"

    # Row 4: BCCK Di Linh
    ws_mb.cell(4, 2).value = "BCCK Di Linh"
    ws_mb.cell(4, 3).value = f"={dl_formula}"
    ws_mb.cell(4, 4).value = "=IFERROR(ROUNDUP(C4/'1. Thông số'!$B$7,0),0)"
    ws_mb.cell(4, 5).value = "=D4*'1. Thông số'!$B$14"
    ws_mb.cell(4, 6).value = None
    ws_mb.cell(4, 7).value = "=IFERROR(D4-F4,0)"
    ws_mb.cell(4, 8).value = "Mở thêm BCCK / Mở rộng BC  hiện tại"
    ws_mb.cell(4, 9).value = "Gộp 3 BC Di Linh - Lâm Hà - Hòa Ninh"
    ws_mb.cell(4, 10).value = 12.0
    ws_mb.cell(4, 11).value = 1.0
    ws_mb.cell(4, 12).value = 1.0
    ws_mb.cell(4, 13).value = "=SUM(J4:L4)"

    # Row 5: BCCK Đơn Dương
    ws_mb.cell(5, 2).value = "BCCK Đơn Dương"
    ws_mb.cell(5, 3).value = f"={dd_formula}"
    ws_mb.cell(5, 4).value = "=IFERROR(ROUNDUP(C5/'1. Thông số'!$B$7,0),0)"
    ws_mb.cell(5, 5).value = "=D5*'1. Thông số'!$B$14"
    ws_mb.cell(5, 6).value = None
    ws_mb.cell(5, 7).value = "=IFERROR(D5-F5,0)"
    ws_mb.cell(5, 8).value = "Mở thêm BCCK / Mở rộng BC  hiện tại"
    ws_mb.cell(5, 9).value = "Gộp Hiệp Thạnh và Đơn Dương"
    ws_mb.cell(5, 10).value = 8.0
    ws_mb.cell(5, 11).value = 1.0
    ws_mb.cell(5, 12).value = 1.0
    ws_mb.cell(5, 13).value = "=SUM(J5:L5)"

    # Row 6: BCKK Đức Linh-Bình Thuận
    ws_mb.cell(6, 2).value = "BCKK Đức Linh-Bình Thuận"
    ws_mb.cell(6, 3).value = f"={dk_formula}"
    ws_mb.cell(6, 4).value = "=IFERROR(ROUNDUP(C6/'1. Thông số'!$B$7,0),0)"
    ws_mb.cell(6, 5).value = 6500000.0 # Hardcoded rent as old
    ws_mb.cell(6, 6).value = None
    ws_mb.cell(6, 7).value = "=IFERROR(D6-F6,0)"
    ws_mb.cell(6, 8).value = "Mở thêm BCCK / Mở rộng BC  hiện tại"
    ws_mb.cell(6, 9).value = None
    ws_mb.cell(6, 10).value = 5.0
    ws_mb.cell(6, 11).value = 1.0
    ws_mb.cell(6, 12).value = 1.0
    ws_mb.cell(6, 13).value = "=SUM(J6:L6)"

    # Clear rows 7 to 17 bưu cục names and details
    for r in range(7, 18):
        ws_mb.cell(r, 2).value = None
        ws_mb.cell(r, 3).value = None
        ws_mb.cell(r, 4).value = f"=IFERROR(ROUNDUP(C{r}/'1. Thông số'!$B$7,0),0)"
        ws_mb.cell(r, 5).value = None
        ws_mb.cell(r, 6).value = None
        ws_mb.cell(r, 7).value = f"=IFERROR(D{r}-F{r},0)"
        ws_mb.cell(r, 8).value = None
        ws_mb.cell(r, 9).value = None
        ws_mb.cell(r, 10).value = None
        ws_mb.cell(r, 11).value = None
        ws_mb.cell(r, 12).value = None
        ws_mb.cell(r, 13).value = f"=SUM(J{r}:L{r})"

    # Row 18: TỔNG CỘNG
    ws_mb.cell(18, 10).value = "=SUM(J3:J17)"
    ws_mb.cell(18, 11).value = "=SUM(K3:K17)"
    ws_mb.cell(18, 12).value = "=SUM(L3:L17)"
    ws_mb.cell(18, 13).value = "=SUM(M3:M17)"

    print("  - Updated Mặt bằng sheet with new formulas and staff allocations.")

    print(f"💾 Saving workbooks to {output_path}...")
    wb.save(output_path)
    print("🎉 Done!")

if __name__ == "__main__":
    run_calculation()

