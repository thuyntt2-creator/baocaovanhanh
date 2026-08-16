# -*- coding: utf-8 -*-
import sys, os, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

ntb_file = r'C:\Users\lap4all\Downloads\config_psbba_NTB.xlsx'
out_check_file = r'C:\Users\lap4all\Downloads\Kiem_Tra_Cong_Thuc_Event_8.8_NTB_Corrected.xlsx'
workspace_check_file = r'c:\Users\lap4all\Documents\Auto report\Kiem_Tra_Cong_Thuc_Event_8.8_NTB_Corrected.xlsx'

wb = openpyxl.load_workbook(ntb_file)

sheet_name = 'CHECK_FORMULAS_EVENT_88'
if sheet_name in wb.sheetnames:
    del wb[sheet_name]

ws = wb.create_sheet(title=sheet_name, index=0)

# Styles
font_title = Font(name='Arial', size=16, bold=True, color='1F497D')
font_subtitle = Font(name='Arial', size=11, italic=True, color='595959')
font_sec_hdr = Font(name='Arial', size=13, bold=True, color='1F497D')
font_tbl_hdr = Font(name='Arial', size=10, bold=True, color='FFFFFF')
font_data = Font(name='Arial', size=10, bold=False, color='000000')
font_bold = Font(name='Arial', size=10, bold=True, color='000000')
font_red = Font(name='Arial', size=10, bold=True, color='C00000')

fill_tbl_hdr = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
fill_gt_row = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
fill_subtot_row = PatternFill(start_color='E9EDF4', end_color='E9EDF4', fill_type='solid')

thin_border_side = Side(border_style='thin', color='D3D3D3')
border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
border_header = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

align_left = Alignment(horizontal='left', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')
align_center = Alignment(horizontal='center', vertical='center')

target_dates = ['06/08', '07/08', '08/08', '09/08', '10/08', '11/08', '12/08', '13/08', '14/08', '15/08']
date_col_letters = ['AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK']
sort_date_col_letters = ['Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH']

ws['A1'] = "BẢNG KIỂM TRA HÀM CÔNG THỨC EXCEL - EVENT 8.8 (NTB)"
ws['A1'].font = font_title
ws['A2'] = "File đối chiếu công thức Excel từ sheet 6_FC_Lay_Daily, 7_FC_Giao_Daily & FC Sorting 60d"
ws['A2'].font = font_subtitle

curr_row = 4

# --- SECTION 1: VOLUME LẤY THEO SÀN ---
ws.cell(curr_row, 1, "1. VOLUME LẤY THEO SÀN (Event 8.8)").font = font_sec_hdr
curr_row += 1

headers_t1 = ['Sàn / Kênh'] + target_dates + ['Tổng 10 ngày', 'Trung bình']
for c_i, h in enumerate(headers_t1, 1):
    cell = ws.cell(curr_row, c_i, h)
    cell.font = font_tbl_hdr
    cell.fill = fill_tbl_hdr
    cell.alignment = align_center
    cell.border = border_header

curr_row += 1
t1_start_row = curr_row

san_list = ['Shopee', 'Shopee-Bulky', 'Shopee-Bulky (10-15kg)', 'SME', 'SME-Bulky', 'TTS', 'TTS-Bulky']
for s in san_list:
    ws.cell(curr_row, 1, s).font = font_data
    ws.cell(curr_row, 1).alignment = align_left
    ws.cell(curr_row, 1).border = border_all
    
    for d_i, col_let in enumerate(date_col_letters, 2):
        cell = ws.cell(curr_row, d_i)
        cell.value = f"=SUMIFS('6_FC_Lay_Daily'!{col_let}:{col_let}, '6_FC_Lay_Daily'!$E:$E, $A{curr_row})"
        cell.font = font_data
        cell.alignment = align_right
        cell.number_format = '#,##0'
        cell.border = border_all
    
    cell_tot = ws.cell(curr_row, 12, f"=SUM(B{curr_row}:K{curr_row})")
    cell_tot.font = font_bold
    cell_tot.alignment = align_right
    cell_tot.number_format = '#,##0'
    cell_tot.border = border_all
    
    cell_avg = ws.cell(curr_row, 13, f"=AVERAGE(B{curr_row}:K{curr_row})")
    cell_avg.font = font_bold
    cell_avg.alignment = align_right
    cell_avg.number_format = '#,##0'
    cell_avg.border = border_all
    
    curr_row += 1

t1_end_row = curr_row - 1

ws.cell(curr_row, 1, "Grand Total").font = font_bold
ws.cell(curr_row, 1).fill = fill_gt_row
ws.cell(curr_row, 1).alignment = align_left
ws.cell(curr_row, 1).border = border_all

for d_i in range(2, 14):
    c_let = get_column_letter(d_i)
    cell = ws.cell(curr_row, d_i, f"=SUM({c_let}{t1_start_row}:{c_let}{t1_end_row})")
    cell.font = font_bold
    cell.fill = fill_gt_row
    cell.alignment = align_right
    cell.number_format = '#,##0'
    cell.border = border_all

t1_gt_row = curr_row
curr_row += 3

# --- SECTION 2: VOLUME LẤY THEO TỈNH ---
ws.cell(curr_row, 1, "2. VOLUME LẤY THEO TỈNH/QUẬN (Event 8.8)").font = font_sec_hdr
curr_row += 1

headers_t2 = ['Tỉnh / Quận'] + target_dates + ['Tổng 10 ngày', 'Trung bình']
for c_i, h in enumerate(headers_t2, 1):
    cell = ws.cell(curr_row, c_i, h)
    cell.font = font_tbl_hdr
    cell.fill = fill_tbl_hdr
    cell.alignment = align_center
    cell.border = border_header

curr_row += 1
t2_start_row = curr_row

tinh_list = ['Bình Thuận', 'Khánh Hòa', 'Lâm Đồng', 'Ninh Thuận', 'Đắk Nông']
for t in tinh_list:
    ws.cell(curr_row, 1, t).font = font_data
    ws.cell(curr_row, 1).alignment = align_left
    ws.cell(curr_row, 1).border = border_all
    
    for d_i, col_let in enumerate(date_col_letters, 2):
        cell = ws.cell(curr_row, d_i)
        cell.value = f"=SUMIFS('6_FC_Lay_Daily'!{col_let}:{col_let}, '6_FC_Lay_Daily'!$B:$B, $A{curr_row})"
        cell.font = font_data
        cell.alignment = align_right
        cell.number_format = '#,##0'
        cell.border = border_all
    
    cell_tot = ws.cell(curr_row, 12, f"=SUM(B{curr_row}:K{curr_row})")
    cell_tot.font = font_bold
    cell_tot.alignment = align_right
    cell_tot.number_format = '#,##0'
    cell_tot.border = border_all
    
    cell_avg = ws.cell(curr_row, 13, f"=AVERAGE(B{curr_row}:K{curr_row})")
    cell_avg.font = font_bold
    cell_avg.alignment = align_right
    cell_avg.number_format = '#,##0'
    cell_avg.border = border_all
    
    curr_row += 1

t2_end_row = curr_row - 1

ws.cell(curr_row, 1, "Grand Total").font = font_bold
ws.cell(curr_row, 1).fill = fill_gt_row
ws.cell(curr_row, 1).alignment = align_left
ws.cell(curr_row, 1).border = border_all

for d_i in range(2, 14):
    c_let = get_column_letter(d_i)
    cell = ws.cell(curr_row, d_i, f"=SUM({c_let}{t2_start_row}:{c_let}{t2_end_row})")
    cell.font = font_bold
    cell.fill = fill_gt_row
    cell.alignment = align_right
    cell.number_format = '#,##0'
    cell.border = border_all

t2_gt_row = curr_row
curr_row += 3

# --- SECTION 3: VOLUME GIAO THEO SÀN ---
ws.cell(curr_row, 1, "3. VOLUME GIAO THEO SÀN (Event 8.8)").font = font_sec_hdr
curr_row += 1

headers_t3 = ['Sàn / Kênh'] + target_dates + ['Tổng 10 ngày', 'Trung bình']
for c_i, h in enumerate(headers_t3, 1):
    cell = ws.cell(curr_row, c_i, h)
    cell.font = font_tbl_hdr
    cell.fill = fill_tbl_hdr
    cell.alignment = align_center
    cell.border = border_header

curr_row += 1
t3_start_row = curr_row

for s in san_list:
    ws.cell(curr_row, 1, s).font = font_data
    ws.cell(curr_row, 1).alignment = align_left
    ws.cell(curr_row, 1).border = border_all
    
    for d_i, col_let in enumerate(date_col_letters, 2):
        cell = ws.cell(curr_row, d_i)
        cell.value = f"=SUMIFS('7_FC_Giao_Daily'!{col_let}:{col_let}, '7_FC_Giao_Daily'!$E:$E, $A{curr_row})"
        cell.font = font_data
        cell.alignment = align_right
        cell.number_format = '#,##0'
        cell.border = border_all
    
    cell_tot = ws.cell(curr_row, 12, f"=SUM(B{curr_row}:K{curr_row})")
    cell_tot.font = font_bold
    cell_tot.alignment = align_right
    cell_tot.number_format = '#,##0'
    cell_tot.border = border_all
    
    cell_avg = ws.cell(curr_row, 13, f"=AVERAGE(B{curr_row}:K{curr_row})")
    cell_avg.font = font_bold
    cell_avg.alignment = align_right
    cell_avg.number_format = '#,##0'
    cell_avg.border = border_all
    
    curr_row += 1

t3_end_row = curr_row - 1

ws.cell(curr_row, 1, "Grand Total").font = font_bold
ws.cell(curr_row, 1).fill = fill_gt_row
ws.cell(curr_row, 1).alignment = align_left
ws.cell(curr_row, 1).border = border_all

for d_i in range(2, 14):
    c_let = get_column_letter(d_i)
    cell = ws.cell(curr_row, d_i, f"=SUM({c_let}{t3_start_row}:{c_let}{t3_end_row})")
    cell.font = font_bold
    cell.fill = fill_gt_row
    cell.alignment = align_right
    cell.number_format = '#,##0'
    cell.border = border_all

t3_gt_row = curr_row
curr_row += 3

# --- SECTION 4: VOLUME GIAO THEO TỈNH ---
ws.cell(curr_row, 1, "4. VOLUME GIAO THEO TỈNH/QUẬN (Event 8.8)").font = font_sec_hdr
curr_row += 1

headers_t4 = ['Tỉnh / Quận'] + target_dates + ['Tổng 10 ngày', 'Trung bình']
for c_i, h in enumerate(headers_t4, 1):
    cell = ws.cell(curr_row, c_i, h)
    cell.font = font_tbl_hdr
    cell.fill = fill_tbl_hdr
    cell.alignment = align_center
    cell.border = border_header

curr_row += 1
t4_start_row = curr_row

for t in tinh_list:
    ws.cell(curr_row, 1, t).font = font_data
    ws.cell(curr_row, 1).alignment = align_left
    ws.cell(curr_row, 1).border = border_all
    
    for d_i, col_let in enumerate(date_col_letters, 2):
        cell = ws.cell(curr_row, d_i)
        cell.value = f"=SUMIFS('7_FC_Giao_Daily'!{col_let}:{col_let}, '7_FC_Giao_Daily'!$B:$B, $A{curr_row})"
        cell.font = font_data
        cell.alignment = align_right
        cell.number_format = '#,##0'
        cell.border = border_all
    
    cell_tot = ws.cell(curr_row, 12, f"=SUM(B{curr_row}:K{curr_row})")
    cell_tot.font = font_bold
    cell_tot.alignment = align_right
    cell_tot.number_format = '#,##0'
    cell_tot.border = border_all
    
    cell_avg = ws.cell(curr_row, 13, f"=AVERAGE(B{curr_row}:K{curr_row})")
    cell_avg.font = font_bold
    cell_avg.alignment = align_right
    cell_avg.number_format = '#,##0'
    cell_avg.border = border_all
    
    curr_row += 1

t4_end_row = curr_row - 1

ws.cell(curr_row, 1, "Grand Total").font = font_bold
ws.cell(curr_row, 1).fill = fill_gt_row
ws.cell(curr_row, 1).alignment = align_left
ws.cell(curr_row, 1).border = border_all

for d_i in range(2, 14):
    c_let = get_column_letter(d_i)
    cell = ws.cell(curr_row, d_i, f"=SUM({c_let}{t4_start_row}:{c_let}{t4_end_row})")
    cell.font = font_bold
    cell.fill = fill_gt_row
    cell.alignment = align_right
    cell.number_format = '#,##0'
    cell.border = border_all

t4_gt_row = curr_row
curr_row += 3

# --- SECTION 5: ĐỐI CHIẾU COMBO CHART ---
ws.cell(curr_row, 1, "5. BẢNG TÍNH COMBO CHART (VOLUME, TRUNG BÌNH, % TĂNG/GIẢM FC)").font = font_sec_hdr
curr_row += 1

ws.cell(curr_row, 1, "A. FC Volume Lấy event 08.08 (Style Combo Chart)").font = font_bold
curr_row += 1

headers_combo = ['Chỉ tiêu'] + target_dates + ['Trung bình toàn đợt']
for c_i, h in enumerate(headers_combo, 1):
    cell = ws.cell(curr_row, c_i, h)
    cell.font = font_tbl_hdr
    cell.fill = fill_tbl_hdr
    cell.alignment = align_center
    cell.border = border_header

curr_row += 1
ws.cell(curr_row, 1, "Volume lấy").font = font_bold
ws.cell(curr_row, 1).alignment = align_left
ws.cell(curr_row, 1).border = border_all
for d_i in range(2, 12):
    c_let = get_column_letter(d_i)
    cell = ws.cell(curr_row, d_i, f"={c_let}{t1_gt_row}")
    cell.font = font_bold
    cell.alignment = align_right
    cell.number_format = '#,##0'
    cell.border = border_all

cell_avg_l = ws.cell(curr_row, 12, f"=AVERAGE(B{curr_row}:K{curr_row})")
cell_avg_l.font = font_red
cell_avg_l.alignment = align_right
cell_avg_l.number_format = '#,##0'
cell_avg_l.border = border_all
r_lay_vol = curr_row
curr_row += 1

ws.cell(curr_row, 1, "Trung bình").font = font_bold
ws.cell(curr_row, 1).alignment = align_left
ws.cell(curr_row, 1).border = border_all
for d_i in range(2, 12):
    cell = ws.cell(curr_row, d_i, f"=$L${r_lay_vol}")
    cell.font = font_red
    cell.alignment = align_right
    cell.number_format = '#,##0'
    cell.border = border_all
ws.cell(curr_row, 12, f"=$L${r_lay_vol}").font = font_red
ws.cell(curr_row, 12).alignment = align_right
ws.cell(curr_row, 12).number_format = '#,##0'
ws.cell(curr_row, 12).border = border_all
curr_row += 1

ws.cell(curr_row, 1, "Tăng/ giảm FC").font = font_bold
ws.cell(curr_row, 1).alignment = align_left
ws.cell(curr_row, 1).border = border_all
for d_i in range(2, 12):
    c_let = get_column_letter(d_i)
    cell = ws.cell(curr_row, d_i, f"=({c_let}{r_lay_vol}-$L${r_lay_vol})/$L${r_lay_vol}")
    cell.font = font_bold
    cell.alignment = align_right
    cell.number_format = '+0.00%;-0.00%;0.00%'
    cell.border = border_all
ws.cell(curr_row, 12, 0).font = font_bold
ws.cell(curr_row, 12).alignment = align_right
ws.cell(curr_row, 12).number_format = '0.00%'
ws.cell(curr_row, 12).border = border_all

curr_row += 3

ws.cell(curr_row, 1, "B. FC Volume Giao event 08.08 (Style Combo Chart)").font = font_bold
curr_row += 1

for c_i, h in enumerate(headers_combo, 1):
    cell = ws.cell(curr_row, c_i, h)
    cell.font = font_tbl_hdr
    cell.fill = fill_tbl_hdr
    cell.alignment = align_center
    cell.border = border_header

curr_row += 1
ws.cell(curr_row, 1, "Volume giao").font = font_bold
ws.cell(curr_row, 1).alignment = align_left
ws.cell(curr_row, 1).border = border_all
for d_i in range(2, 12):
    c_let = get_column_letter(d_i)
    cell = ws.cell(curr_row, d_i, f"={c_let}{t3_gt_row}")
    cell.font = font_bold
    cell.alignment = align_right
    cell.number_format = '#,##0'
    cell.border = border_all

cell_avg_g = ws.cell(curr_row, 12, f"=AVERAGE(B{curr_row}:K{curr_row})")
cell_avg_g.font = font_red
cell_avg_g.alignment = align_right
cell_avg_g.number_format = '#,##0'
cell_avg_g.border = border_all
r_giao_vol = curr_row
curr_row += 1

ws.cell(curr_row, 1, "Trung bình").font = font_bold
ws.cell(curr_row, 1).alignment = align_left
ws.cell(curr_row, 1).border = border_all
for d_i in range(2, 12):
    cell = ws.cell(curr_row, d_i, f"=$L${r_giao_vol}")
    cell.font = font_red
    cell.alignment = align_right
    cell.number_format = '#,##0'
    cell.border = border_all
ws.cell(curr_row, 12, f"=$L${r_giao_vol}").font = font_red
ws.cell(curr_row, 12).alignment = align_right
ws.cell(curr_row, 12).number_format = '#,##0'
ws.cell(curr_row, 12).border = border_all
curr_row += 1

ws.cell(curr_row, 1, "Tăng/ giảm FC").font = font_bold
ws.cell(curr_row, 1).alignment = align_left
ws.cell(curr_row, 1).border = border_all
for d_i in range(2, 12):
    c_let = get_column_letter(d_i)
    cell = ws.cell(curr_row, d_i, f"=({c_let}{r_giao_vol}-$L${r_giao_vol})/$L${r_giao_vol}")
    cell.font = font_bold
    cell.alignment = align_right
    cell.number_format = '+0.00%;-0.00%;0.00%'
    cell.border = border_all
ws.cell(curr_row, 12, 0).font = font_bold
ws.cell(curr_row, 12).alignment = align_right
ws.cell(curr_row, 12).number_format = '0.00%'
ws.cell(curr_row, 12).border = border_all

curr_row += 3

# --- SECTION 6: KHO SORTING / HUB BREAKDOWN (PULLING 100% DIRECTLY FROM FC SORTING 60D SHEET) ---
ws.cell(curr_row, 1, "6. PHÂN BỔ VOLUME SORTING THEO 5 KHO KTC / CHUYỂN TIẾP (TABLE 8 - PULL TRỰC TIẾP TỪ FC SORTING 60D)").font = font_sec_hdr
curr_row += 1

headers_t8 = ['Kho / Hub', 'Nhóm hàng'] + target_dates + ['Tổng 10 ngày', 'Trung bình']
for c_i, h in enumerate(headers_t8, 1):
    cell = ws.cell(curr_row, c_i, h)
    cell.font = font_tbl_hdr
    cell.fill = fill_tbl_hdr
    cell.alignment = align_center
    cell.border = border_header

curr_row += 1
t8_start_row = curr_row

hubs_def = [
    'Kho Trung Chuyển Khánh Hòa',
    'Kho Chuyển Tiếp Bình Thuận',
    'Kho Chuyển Tiếp Đức Trọng-Lâm Đồng',
    'Kho Chuyển Tiếp Bảo Lộc-Lâm Đồng',
    'Kho Chuyển Tiếp Đắk Nông'
]

total_row_indices = []

for h_title in hubs_def:
    for cat in ['Normal', 'Bulky', 'Freight', 'Total']:
        is_tot = (cat == 'Total')
        bg = fill_subtot_row if is_tot else None
        if is_tot:
            total_row_indices.append(curr_row)
        
        ws.cell(curr_row, 1, h_title).font = font_bold if is_tot else font_data
        ws.cell(curr_row, 1).alignment = align_left
        ws.cell(curr_row, 1).border = border_all
        if bg: ws.cell(curr_row, 1).fill = bg
        
        ws.cell(curr_row, 2, cat).font = font_bold if is_tot else font_data
        ws.cell(curr_row, 2).alignment = align_left
        ws.cell(curr_row, 2).border = border_all
        if bg: ws.cell(curr_row, 2).fill = bg
        
        for d_i, col_let in enumerate(sort_date_col_letters, 3):
            cell = ws.cell(curr_row, d_i)
            cell.value = f"=SUMIFS('FC Sorting 60d'!{col_let}:{col_let}, 'FC Sorting 60d'!$A:$A, $A{curr_row}, 'FC Sorting 60d'!$B:$B, $B{curr_row})"
            cell.font = font_bold if is_tot else font_data
            cell.alignment = align_right
            cell.number_format = '#,##0'
            cell.border = border_all
            if bg: cell.fill = bg
            
        cell_tot = ws.cell(curr_row, 13, f"=SUM(C{curr_row}:L{curr_row})")
        cell_tot.font = font_bold
        cell_tot.alignment = align_right
        cell_tot.number_format = '#,##0'
        cell_tot.border = border_all
        if bg: cell_tot.fill = bg
        
        cell_avg = ws.cell(curr_row, 14, f"=AVERAGE(C{curr_row}:L{curr_row})")
        cell_avg.font = font_bold
        cell_avg.alignment = align_right
        cell_avg.number_format = '#,##0'
        cell_avg.border = border_all
        if bg: cell_avg.fill = bg
        
        curr_row += 1

# Total NTB Sorting Row
ws.cell(curr_row, 1, "TỔNG NTB SORTING").font = font_bold
ws.cell(curr_row, 1).fill = fill_gt_row
ws.cell(curr_row, 1).border = border_all

ws.cell(curr_row, 2, "TỔNG KTC").font = font_bold
ws.cell(curr_row, 2).fill = fill_gt_row
ws.cell(curr_row, 2).border = border_all

sum_terms = "+".join([f"{{col}}{r}" for r in total_row_indices])

for d_i in range(3, 15):
    c_let = get_column_letter(d_i)
    cell = ws.cell(curr_row, d_i, f"={sum_terms.format(col=c_let)}")
    cell.font = font_bold
    cell.fill = fill_gt_row
    cell.alignment = align_right
    cell.number_format = '#,##0'
    cell.border = border_all

ws.column_dimensions['A'].width = 32
ws.column_dimensions['B'].width = 14
for c_i in range(3, 15):
    c_let = get_column_letter(c_i)
    ws.column_dimensions[c_let].width = 15

# Save to standalone check file
for p in [out_check_file, workspace_check_file]:
    try:
        wb.save(p)
        print(f"Successfully saved Excel check workbook: {p}")
    except Exception as e:
        print(f"Could not save {p}: {e}")

print("Excel formula check workbook updated with 100% FC Sorting 60d links successfully!")
