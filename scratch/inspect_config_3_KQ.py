import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

config_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB.xlsx"
wb = openpyxl.load_workbook(config_path, data_only=True)

print("=== Sheet: 3_KQ_BC_Detail ===")
sheet = wb['3_KQ_BC_Detail']
print(f"Max row: {sheet.max_row}, Max col: {sheet.max_column}")
for r in range(1, 25):
    print(f"Row {r:2d}: {[sheet.cell(r, c).value for c in range(1, 15)]}")

print("\n=== Sheet: 1_Config_Chuan ===")
sheet_cfg = wb['1_Config_Chuan']
print(f"Max row: {sheet_cfg.max_row}, Max col: {sheet_cfg.max_column}")
for r in range(1, 10):
    print(f"Row {r:2d}: {[sheet_cfg.cell(r, c).value for c in range(1, 15)]}")

print("\n=== Sheet: Topline_Giao_NTB ===")
sheet_tl = wb['Topline_Giao_NTB']
print(f"Max row: {sheet_tl.max_row}, Max col: {sheet_tl.max_column}")
for r in range(1, 10):
    print(f"Row {r:2d}: {[sheet_tl.cell(r, c).value for c in range(1, 10)]}")

