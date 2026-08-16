import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v5.xlsx"
if not os.path.exists(file_path):
    print("v5 not found")
    sys.exit(1)

wb_value = openpyxl.load_workbook(file_path, data_only=True)

# For each month T7-T12, get the list of values for 'BCCK Nha Trang'
data = {}
for m in [7, 8, 9, 10, 11, 12]:
    sheet_name = f"Forecast T{m}"
    sheet = wb_value[sheet_name]
    
    # Dates and values
    vols = []
    for c in range(2, sheet.max_column + 1):
        d_val = sheet.cell(2, c).value
        vol_val = sheet.cell(4, c).value
        if d_val is not None:
            vols.append(vol_val)
    data[m] = vols
    print(f"Month {m}: Total = {sum(vols)}, length = {len(vols)}, values = {vols[:10]}")

# Let's check the ratio of T9 to T8 in v5:
# Is T9[i] = T8[i] * (Total_T9 / Total_T8) ?
# Or is T9[i] related to T7[i]?
t8_vols = data[8]
t9_vols = data[9]
ratio_t9_t8 = sum(t9_vols) / sum(t8_vols)
print(f"\nT9 sum / T8 sum = {ratio_t9_t8:.6f}")

for i in range(min(len(t8_vols), len(t9_vols))):
    expected = t8_vols[i] * ratio_t9_t8
    actual = t9_vols[i]
    diff = actual - expected
    print(f"Day {i+1}: T8={t8_vols[i]} | Expected T9={expected:.2f} | Actual T9={actual} | Diff={diff:.2f}")

# Let's also check if T9 is related to T7
t7_vols = data[7]
ratio_t9_t7 = sum(t9_vols) / sum(t7_vols)
print(f"\nT9 sum / T7 sum = {ratio_t9_t7:.6f}")
for i in range(min(len(t7_vols), len(t9_vols))):
    expected = t7_vols[i] * ratio_t9_t7
    actual = t9_vols[i]
    diff = actual - expected
    print(f"Day {i+1}: T7={t7_vols[i]} | Expected T9={expected:.2f} | Actual T9={actual} | Diff={diff:.2f}")

