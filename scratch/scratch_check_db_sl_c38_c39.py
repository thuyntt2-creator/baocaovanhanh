import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['Định biên & Sản lượng']

wb_eval = openpyxl.load_workbook(file_path, data_only=True)
sheet_eval = wb_eval['Định biên & Sản lượng']

print("=== Cell values/formulas in 'Định biên & Sản lượng' ===")
for r in [38, 39]:
    cell_formula = sheet.cell(r, 3).value
    cell_val = sheet_eval.cell(r, 3).value
    label = sheet.cell(r, 2).value or sheet.cell(r, 1).value
    print(f"Row {r:2d} | Label: '{label}' | Formula: '{cell_formula}' | Value: '{cell_val}'")

wb_eval.close()
wb.close()
