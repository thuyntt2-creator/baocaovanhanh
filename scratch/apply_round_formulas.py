import openpyxl
import win32com.client
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

downloads_dir = r"C:\Users\lap4all\Downloads"
files_to_fix = [
    os.path.join(downloads_dir, "NTB_Input_Da_Dien_FLM_CRC_2.xlsx"),
    os.path.join(downloads_dir, "NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx")
]

cols = ['D', 'E', 'F', 'G', 'H', 'I']
flm_cols = ['J', 'K', 'L', 'M', 'N', 'O']

for filepath in files_to_fix:
    if not os.path.exists(filepath):
        continue
        
    print(f"Applying ROUND to formulas in: {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath, data_only=False)
    sheet = wb['NTB – Input']
    
    for idx, col in enumerate(cols):
        f_col = flm_cols[idx]
        
        # Round Giao and Nhận to 2 decimals
        sheet[f"{col}41"] = f"=ROUND('Chi phí FLM'!{f_col}37*4/1000000, 2)"
        sheet[f"{col}42"] = f"=ROUND('Chi phí FLM'!{f_col}38*4/1000000, 2)"
        
        # Round Total FLM to 2 decimals
        sheet[f"{col}43"] = f"=ROUND({col}41+{col}42, 2)"
        
        # Round Chi phí / đơn to 0 decimals (integer VNĐ/đơn)
        sheet[f"{col}44"] = f"=ROUND(IF({col}7=0,0,{col}43*1000000/{col}7), 0)"
        
        # Round Cost / kg to 2 decimals
        sheet[f"{col}45"] = f"=ROUND('Chi phí FLM'!{f_col}46, 2)"
        
    wb.save(filepath)
    wb.close()
    
    # Recalculate via Excel COM
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        workbook = excel.Workbooks.Open(filepath)
        workbook.Save()
        workbook.Close()
        print("Success!")
    except Exception as e:
        print(f"Error during recalculation: {e}")
    finally:
        excel.Quit()
        
# Verify the rounded values
for filepath in files_to_fix:
    if not os.path.exists(filepath):
        continue
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb['NTB – Input']
    print(f"\n=== ROUNDED VALUES IN: {os.path.basename(filepath)} ===")
    for r in [41, 42, 43, 44, 45]:
        label = sheet.cell(r, 2).value or sheet.cell(r, 1).value
        vals = [sheet.cell(r, openpyxl.utils.column_index_from_string(c)).value for c in cols]
        print(f"Row {r:2d} | {str(label)[:35]:<35} | {vals}")
    wb.close()
