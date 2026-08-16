import openpyxl, os, sys, glob
sys.stdout.reconfigure(encoding='utf-8')

for f in glob.glob(r'C:\Users\lap4all\Downloads\*.xlsx'):
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        for s in wb.sheetnames:
            sheet = wb[s]
            for r in range(1, 100):
                for c in range(1, 50):
                    val = sheet.cell(r, c).value
                    if isinstance(val, (int, float)) and round(val) == 624:
                        print(f'Found 624 in {os.path.basename(f)}!{s}!{sheet.cell(r, c).coordinate} -> {val}')
                    if isinstance(val, (int, float)) and round(val) == 211:
                        print(f'Found 211 in {os.path.basename(f)}!{s}!{sheet.cell(r, c).coordinate} -> {val}')
    except Exception as e:
        pass
print('Finished.')
