import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\V3_AOP_Hang_NTB_T7-T12_2026 mới.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['Nguồn lực & chi phí']

print("=== Formulas in V3 Nguồn lực & chi phí ===")
for r in range(1, 25):
    for c in range(1, 8):
        val = sheet.cell(r, c).value
        if isinstance(val, str) and val.startswith('='):
            col = openpyxl.utils.get_column_letter(c)
            print(f"Cell {col}{r}: {val}")

wb.close()
