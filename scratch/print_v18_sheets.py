import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v18.xlsx"

if not os.path.exists(excel_path):
    print("File không tồn tại")
    sys.exit(1)

wb = openpyxl.load_workbook(excel_path, read_only=True)
print("Các sheet thực tế:", wb.sheetnames)
for sname in wb.sheetnames:
    print(f"Sheet name representation: {repr(sname)}")
