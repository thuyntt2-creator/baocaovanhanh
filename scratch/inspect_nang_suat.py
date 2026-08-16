import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['Định biên & Sản lượng']

print("=== Định biên & Sản lượng - Năng suất rows (raw formulas/values) ===")
for r in range(1, 20):
    row_vals = [sheet.cell(r, c).value for c in range(1, 6)]
    if any(v is not None for v in row_vals):
        print(f"Row {r:2d}: {row_vals}")

wb.close()

# Also check NTB – Input formulas for rows 17, 22, 27
wb2 = openpyxl.load_workbook(file_path, data_only=False)
sheet2 = wb2['NTB – Input']
print("\n=== NTB – Input - Năng suất formulas ===")
for r in [17, 22, 27]:
    label = sheet2.cell(r, 2).value or sheet2.cell(r, 1).value
    vals = [sheet2[f"{c}{r}"].value for c in ['D','E','F','G','H','I']]
    print(f"Row {r:2d} ({label}): {vals}")
wb2.close()
