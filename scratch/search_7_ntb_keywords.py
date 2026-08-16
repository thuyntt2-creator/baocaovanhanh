import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

filepath = r"C:\Users\lap4all\Downloads\Telegram Desktop\7. NTB_2026.xlsx"
wb = openpyxl.load_workbook(filepath, data_only=True)
sheet = wb['7. NTB_2026']

print("=== Searching for keywords in 7. NTB_2026 sheet ===")
for r in range(1, min(sheet.max_row + 1, 200)):
    row_vals = [sheet.cell(r, c).value for c in range(1, min(sheet.max_column + 1, 15))]
    row_str = " | ".join(str(v) for v in row_vals if v is not None)
    if any(k in row_str.lower() for k in ["utility", "utilities", "điện", "nước", "bc", "bưu cục"]):
        print(f"Row {r:3d}: {row_str[:120]}")
wb.close()
