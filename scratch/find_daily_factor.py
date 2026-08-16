import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v5.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

for m in [9, 10, 11, 12]:
    print(f"\n--- Checking Month T{m} vs T8 ---")
    sheet_prev = wb['Forecast T8']
    sheet_curr = wb[f'Forecast T{m}']
    
    # We will check rows 4 to 7
    for r in range(4, 8):
        name = sheet_curr.cell(r, 1).value
        prev_vals = []
        curr_vals = []
        for c in range(2, 32): # first 30 days
            v_prev = sheet_prev.cell(r, c).value
            v_curr = sheet_curr.cell(r, c).value
            if v_prev is not None and v_curr is not None:
                prev_vals.append(v_prev)
                curr_vals.append(v_curr)
                
        # Calculate ratios
        ratios = [c / p for p, c in zip(prev_vals, curr_vals) if p > 0]
        avg_ratio = sum(ratios) / len(ratios) if ratios else 0
        print(f"  BC: {name} | Avg Ratio: {avg_ratio:.6f} | Min Ratio: {min(ratios):.6f} | Max Ratio: {max(ratios):.6f}")

