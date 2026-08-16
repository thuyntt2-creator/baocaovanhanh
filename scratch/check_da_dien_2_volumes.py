import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

filepath = r"C:\Users\lap4all\Downloads\NTB_Input_Da_Dien_FLM_CRC_2.xlsx"
if os.path.exists(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb['NTB – Input']
    print("=== Volumes in NTB_Input_Da_Dien_FLM_CRC_2.xlsx ===")
    cols = ['D', 'E', 'F', 'G', 'H', 'I']
    for r in range(7, 14):
        label = sheet.cell(r, 2).value or sheet.cell(r, 1).value
        vals = [sheet.cell(r, openpyxl.utils.column_index_from_string(c)).value for c in cols]
        print(f"Row {r:2d} | {str(label)[:35]:<35} | {vals}")
    wb.close()
