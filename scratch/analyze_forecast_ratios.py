import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v18.xlsx"
wb_value = openpyxl.load_workbook(file_path, data_only=True)

# For each month T7-T12, get the list of values for 'BCCK Nha Trang'
for m in [7, 8, 9, 10, 11, 12]:
    sheet_name = f"Forecast T{m}"
    sheet = wb_value[sheet_name]
    
    # Dates and values
    dates = []
    days = []
    vols = []
    for c in range(2, sheet.max_column + 1):
        d_val = sheet.cell(2, c).value
        day_val = sheet.cell(3, c).value
        vol_val = sheet.cell(4, c).value
        if d_val is not None:
            dates.append(d_val)
            days.append(day_val)
            vols.append(vol_val)
            
    print(f"\n--- Month {m} (Total Vol: {sum(vols)}, Days: {len(vols)}) ---")
    print("Days of week:", days[:15])
    print("Volumes:     ", vols[:15])
    
    # Calculate day of week averages
    dow_sums = {}
    dow_counts = {}
    for d, v in zip(days, vols):
        if v is not None:
            dow_sums[d] = dow_sums.get(d, 0.0) + v
            dow_counts[d] = dow_counts.get(d, 0) + 1
            
    print("Averages by Day of Week:")
    for d in ['T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'CN']:
        if d in dow_sums:
            avg = dow_sums[d] / dow_counts[d]
            print(f"  {d}: Avg={avg:.2f} (Count={dow_counts[d]}, Sum={dow_sums[d]})")
            
    # Check if they are exact multiples of some base pattern
    # For example, does T9 equal T8 multiplied by some factor? Or is T9 equal to another month?
