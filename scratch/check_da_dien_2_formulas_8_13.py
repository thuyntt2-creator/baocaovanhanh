import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

filepath = r"C:\Users\lap4all\Downloads\NTB_Input_Da_Dien_FLM_CRC_2.xlsx"
wb = openpyxl.load_workbook(filepath, data_only=False)
sheet = wb['NTB – Input']

cols = ['D', 'E', 'F', 'G', 'H', 'I']
print("=== NTB_Input_Da_Dien_FLM_CRC_2.xlsx - NTB – Input rows 8-13 formulas ===")
for r in range(8, 14):
    label = sheet.cell(r, 2).value or sheet.cell(r, 1).value
    print(f"Row {r:2d} | {str(label)[:35]:<35}")
    for c in cols:
        print(f"  {c}: {sheet[f'{c}{r}'].value}")
wb.close()
