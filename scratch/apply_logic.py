import openpyxl, sys, os
sys.stdout.reconfigure(encoding='utf-8')

# 1. Load config
print("Loading config...")
wb_cfg = openpyxl.load_workbook(r'C:\Users\lap4all\Downloads\config_psbba_NTB.xlsx', data_only=True)
sheet_cfg = wb_cfg['1_Config_Chuan']

# Mapping logic
ward_map = {} # (District, Ward) -> {id, bc_id, bc_name, grp}
bc_ward_count = {} # bc_id -> count
bc_id_to_name = {}

for r in range(2, sheet_cfg.max_row+1):
    dist = sheet_cfg.cell(r, 3).value # Quận_Tên
    ward_id = sheet_cfg.cell(r, 4).value # Phường_ID
    ward_name = sheet_cfg.cell(r, 5).value # Phường_Tên
    bc_id = sheet_cfg.cell(r, 8).value # BC_Giao_ID
    bc_name = sheet_cfg.cell(r, 9).value # BC_Giao_Tên
    grp = sheet_cfg.cell(r, 11).value # B2B_Nhóm
    
    if dist and ward_name and ward_id and bc_id:
        dist = str(dist).strip()
        ward_name = str(ward_name).strip()
        ward_id = str(ward_id).strip()
        bc_id = str(bc_id).strip()
        bc_name = str(bc_name).strip() if bc_name else ""
        
        ward_map[(dist, ward_name)] = {
            'id': ward_id,
            'bc_id': bc_id,
            'bc_name': bc_name,
            'grp': grp
        }
        bc_ward_count[bc_id] = bc_ward_count.get(bc_id, 0) + 1
        bc_id_to_name[bc_id] = bc_name

# 2. Parse 3_KQ_BC_Detail
print("Parsing 3_KQ_BC_Detail...")
sheet_3 = wb_cfg['3_KQ_BC_Detail']
t7_col_3 = 6
bc_reg_vol = {} # bc_id -> vol
for r in range(2, sheet_3.max_row+1):
    bc_id = sheet_3.cell(r, 2).value # routed_bc_id
    weight = str(sheet_3.cell(r, 5).value or '')
    if bc_id and weight in ['10-15kg', '15-30kg', '≥30kg']:
        vol = sheet_3.cell(r, t7_col_3).value or 0
        bc_id = str(bc_id).strip()
        bc_reg_vol[bc_id] = bc_reg_vol.get(bc_id, 0) + vol

# 3. Parse 4_KQ_B2B
print("Parsing 4_KQ_B2B...")
sheet_4 = wb_cfg['4_KQ_B2B']
t7_col_4 = 12
ward_b2b_vol = {} # ward_id -> vol
for r in range(2, sheet_4.max_row+1):
    ward_id = sheet_4.cell(r, 4).value # Phường_ID
    weight = str(sheet_4.cell(r, 11).value or '')
    if ward_id and weight in ['15to30', 'gte30']:
        vol = sheet_4.cell(r, t7_col_4).value or 0
        ward_id = str(ward_id).strip()
        ward_b2b_vol[ward_id] = ward_b2b_vol.get(ward_id, 0) + vol

# 4. Update AOP File
print("Updating V2_AOP_van_fixed...")
aop_path = r'C:\Users\lap4all\Downloads\V2_AOP_van_fixed.xlsx'
wb_aop = openpyxl.load_workbook(aop_path)
sheet_aop = wb_aop['Timeline tiếp nhận']

updated_count = 0
for r in range(3, sheet_aop.max_row+1):
    ward = sheet_aop.cell(r, 1).value
    dist = sheet_aop.cell(r, 2).value
    if ward and dist:
        ward = str(ward).strip()
        dist = str(dist).strip()
        key = (dist, ward)
        
        if key in ward_map:
            w_info = ward_map[key]
            w_id = w_info['id']
            bc_id = w_info['bc_id']
            grp = w_info['grp']
            
            # Calcs
            w_count = bc_ward_count.get(bc_id, 1)
            tot_bc_reg = bc_reg_vol.get(bc_id, 0)
            avg_reg = tot_bc_reg / w_count
            b2b_vol = ward_b2b_vol.get(w_id, 0)
            
            if grp in ['A', 'B']:
                final_vol = b2b_vol + avg_reg
            else:
                final_vol = avg_reg
                
            sheet_aop.cell(r, 3).value = round(final_vol)
            updated_count += 1
            print(f"Updated {ward} ({dist}) - Nhóm {grp}: {round(final_vol)}")
        else:
            print(f"NOT FOUND IN CONFIG: {ward} - {dist}")

out_path = r'C:\Users\lap4all\Downloads\V2_AOP_van_fixed_final.xlsx'
wb_aop.save(out_path)
print(f"\nSaved to {out_path}. Total updated: {updated_count}")
