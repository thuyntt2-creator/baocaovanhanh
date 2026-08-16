import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\V3_AOP_Hang_NTB_T7-T12_2026 mới.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['Hướng dẫn']

print("=== Sheet: Hướng dẫn ===")
for r in range(1, sheet.max_row + 1):
    row_vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
    if any(v is not None for v in row_vals):
        non_empty = [f"Col{c}:{v}" for c, v in enumerate(row_vals, 1) if v is not None]
        print(f"Row {r:2d}: {non_empty}")

wb.close()
