import openpyxl
import os
import glob
import sys

sys.stdout.reconfigure(encoding='utf-8')

downloads_dir = r"C:\Users\lap4all\Downloads"
xlsx_files = glob.glob(os.path.join(downloads_dir, "*.xlsx"))

targets = [68578, 74620, 77680, 86427, 89586, 99054, 198219, 196204, 204379, 225298, 234987, 261430]

print(f"Searching in {len(xlsx_files)} excel files in Downloads:")

for f in xlsx_files:
    # Skip temporary files
    if "~$" in f:
        continue
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # only check first 100 rows and 20 columns to keep it fast
            for r in range(1, min(100, sheet.max_row + 1)):
                for c in range(1, min(20, sheet.max_column + 1)):
                    val = sheet.cell(r, c).value
                    if val is not None:
                        try:
                            num = float(val)
                            for t in targets:
                                if abs(num - t) < 5:
                                    print(f"Found in {os.path.basename(f)} | Sheet '{sheet_name}' cell {openpyxl.utils.get_column_letter(c)}{r}: {val} (target: {t})")
                        except (ValueError, TypeError):
                            pass
        wb.close()
    except Exception as e:
        print(f"Error reading {os.path.basename(f)}: {e}")

print("Search completed.")
