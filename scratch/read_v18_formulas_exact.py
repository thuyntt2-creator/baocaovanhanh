import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v18.xlsx"

if not os.path.exists(excel_path):
    print("File không tồn tại")
    sys.exit(1)

# Đọc file để lấy cả công thức và giá trị chính xác
wb_formula = openpyxl.load_workbook(excel_path, data_only=False)
wb_value = openpyxl.load_workbook(excel_path, data_only=True)

sheet_f = wb_formula['Nguồn lực & chi phí']
sheet_v = wb_value['Nguồn lực & chi phí']

print("=== SO SÁNH CÔNG THỨC VÀ GIÁ TRỊ CHÍNH XÁC TRONG EXCEL V18 ===")
for r in range(1, 23):
    f_vals = [sheet_f.cell(r, c).value for c in range(1, 9)]
    v_vals = [sheet_v.cell(r, c).value for c in range(1, 9)]
    
    if any(x is not None for x in v_vals):
        print(f"\nRow {r:02d} | Title: {v_vals[0]}")
        print(f"  Formulas: {f_vals[1:7]}")
        print(f"  Values  : {v_vals[1:7]}")
