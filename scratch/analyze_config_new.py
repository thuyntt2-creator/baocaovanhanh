import openpyxl
import sys
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB new.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

# Check 3_KQ_BC_Detail - unique weight groups and BC types
print("=== 3_KQ_BC_Detail ===")
sheet = wb['3_KQ_BC_Detail']
headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
print("Headers:", headers)

weight_groups = set()
bc_types = set()
for r in range(2, sheet.max_row + 1):
    wg = sheet.cell(r, 5).value  # weight_group column
    bt = sheet.cell(r, 4).value  # bc_type column
    if wg: weight_groups.add(wg)
    if bt: bc_types.add(bt)

print(f"Weight groups: {weight_groups}")
print(f"BC types: {bc_types}")

# Sum by weight group and month
months_cols = {'T07': 6, 'T08': 7, 'T09': 8, 'T10': 9, 'T11': 10, 'T12': 11}
totals_by_wg = defaultdict(lambda: defaultdict(float))

for r in range(2, sheet.max_row + 1):
    wg = sheet.cell(r, 5).value
    if not wg:
        continue
    for m, col in months_cols.items():
        val = sheet.cell(r, col).value or 0
        totals_by_wg[wg][m] += val

print("\n=== TỔNG VOLUME theo Weight Group và Tháng ===")
for wg in sorted(totals_by_wg.keys()):
    print(f"\n  {wg}:")
    for m in ['T07', 'T08', 'T09', 'T10', 'T11', 'T12']:
        print(f"    {m}: {totals_by_wg[wg][m]:,.1f}")

# Check 4_KQ_B2B - unique weight bands
print("\n\n=== 4_KQ_B2B ===")
sheet_b2b = wb['4_KQ_B2B']
headers_b2b = [sheet_b2b.cell(1, c).value for c in range(1, sheet_b2b.max_column + 1)]
print("Headers:", headers_b2b)

wb.close()
