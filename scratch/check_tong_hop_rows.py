import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"

# Check formulas (data_only=False)
wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['NTB – Input']

print("=== FORMULAS trong dòng 41-45 ===\n")
cols = ['D', 'E', 'F', 'G', 'H', 'I']
for r in range(41, 46):
    label = sheet.cell(r, 2).value or sheet.cell(r, 1).value or f"Row {r}"
    print(f"Row {r} | {label}")
    for col in cols:
        val = sheet[f"{col}{r}"].value
        if val is not None:
            print(f"  {col}: {val}")
        else:
            print(f"  {col}: [TRỐNG]")
wb.close()

# Check calculated values (data_only=True)
wb2 = openpyxl.load_workbook(file_path, data_only=True)
sheet2 = wb2['NTB – Input']
print("\n=== GIÁ TRỊ TÍNH ĐƯỢC trong dòng 41-45 ===\n")
months = ['T7','T8','T9','T10','T11','T12']
for r in range(41, 46):
    label = sheet2.cell(r, 2).value or sheet2.cell(r, 1).value or f"Row {r}"
    vals = [sheet2[f"{c}{r}"].value for c in cols]
    print(f"Row {r} | {label}: {vals}")
wb2.close()
