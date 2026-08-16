import openpyxl
import os
import glob
import sys

sys.stdout.reconfigure(encoding='utf-8')

downloads_dir = r"C:\Users\lap4all\Downloads"
template_files = glob.glob(os.path.join(downloads_dir, "*MAU*.xlsx"))

print("=== Checking templates for setup/di dời costs ===")
for f in template_files:
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        for sname in wb.sheetnames:
            if "input" in sname.lower() or "aop" in sname.lower():
                sheet = wb[sname]
                print(f"\nFile: {os.path.basename(f)} | Sheet: {sname}")
                # check rows 30 to 45
                for r in range(30, min(46, sheet.max_row + 1)):
                    row_lbl = sheet.cell(r, 2).value or sheet.cell(r, 1).value
                    if row_lbl and any(term in str(row_lbl).lower() for term in ["setup", "di dời", "dời", "tiện ích", "utilities"]):
                        vals = [sheet.cell(r, c).value for c in range(4, 10)]
                        print(f"  Row {r}: {row_lbl} -> {vals}")
        wb.close()
    except Exception as e:
        print(f"  Error reading {os.path.basename(f)}: {e}")
