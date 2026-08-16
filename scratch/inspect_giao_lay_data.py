import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

xlsx_aop_path = r"C:\Users\lap4all\Downloads\AOP_Hang_NTB_T7-T12_2026.xlsx"
wb = openpyxl.load_workbook(xlsx_aop_path, data_only=True)

for sheet_name in ['Volume Giao', 'Volume Lấy']:
    sheet = wb[sheet_name]
    print(f"\n--- Sheet: {sheet_name} (max_row: {sheet.max_row}, max_column: {sheet.max_column}) ---")
    for r_idx in range(1, min(sheet.max_row + 1, 30)):
        row_vals = [sheet.cell(r_idx, c_idx).value for c_idx in range(1, min(sheet.max_column + 1, 15))]
        if any(v is not None for v in row_vals):
            print(f"  Row {r_idx:2d}: {row_vals}")

