import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['NTB – Input']

cols = ['D', 'E', 'F', 'G', 'H', 'I']
months = ['T7', 'T8', 'T9', 'T10', 'T11', 'T12']

print("=== Kiểm tra giá trị thực tế trong file (có decimal không?) ===\n")
rows_to_check = [8, 9, 10, 11, 12, 13, 17, 20, 22, 27, 32, 37, 38, 39]

for r in rows_to_check:
    row_label = sheet.cell(r, 2).value or sheet.cell(r, 1).value or f"Row {r}"
    vals = [sheet[f"{c}{r}"].value for c in cols]
    # Check if any has decimal
    has_decimal = any(isinstance(v, float) and v % 1 != 0 for v in vals)
    flag = " ⚠️ CÓ DECIMAL" if has_decimal else ""
    print(f"Row {r:2d} | {str(row_label)[:35]:<35} | {vals}{flag}")

wb.close()
