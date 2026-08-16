import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

downloads_dir = r"C:\Users\lap4all\Downloads"

excel_files = []
for f in os.listdir(downloads_dir):
    if f.endswith('.xlsx') and not f.startswith('~$'):
        excel_files.append(os.path.join(downloads_dir, f))

targets = {
    'Tổng CP T9': 2408.6,
    'Tổng CP T10': 2522.1,
    'Tổng CP T11': 2803.9,
    'Tổng CP T12': 2914.5
}

print(f"=== QUÉT NHANH {len(excel_files)} FILE EXCEL BẰNG OPENPYXL ===")

for path in excel_files:
    fname = os.path.basename(path)
    try:
        # Load file excel (chỉ lấy value)
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        for sname in wb.sheetnames:
            sheet = wb[sname]
            # Quét các ô trong sheet
            # openpyxl read_only sheet hỗ trợ iter_rows
            try:
                for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    for c_idx, val in enumerate(row):
                        if val is not None and isinstance(val, (int, float)):
                            for tname, target in targets.items():
                                if abs(val - target) < 0.1 or abs(val/1e6 - target) < 0.1 or abs(val/1e3 - target) < 0.1:
                                    print(f"[KHỚP] File: {fname} | Sheet: {sname} | Ô {r_idx+1},{c_idx+1}: Giá trị = {val} ({tname})")
            except Exception as e:
                # Một số file có thể bị lỗi khi đọc dòng
                pass
    except Exception as e:
        print(f"Lỗi đọc {fname}: {e}")
print("=== HOÀN THÀNH QUÉT ===")
