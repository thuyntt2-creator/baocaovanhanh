import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"

wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['Chi phí FLM']

print("=== Sheet 'Chi phí FLM' - formulas rows 30-50 ===\n")
cols_check = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O']
for r in range(30, 50):
    row_vals = [sheet.cell(r, c).value for c in range(1, 16)]
    if any(v is not None for v in row_vals):
        print(f"Row {r:2d}: {row_vals[:6]}...")

wb.close()

# Also check values
wb2 = openpyxl.load_workbook(file_path, data_only=True)
sheet2 = wb2['Chi phí FLM']
print("\n=== VALUES rows 30-50 (cols J-O = T7-T12) ===\n")
for r in range(30, 50):
    label = sheet2.cell(r, 1).value or sheet2.cell(r, 2).value
    vals = [sheet2.cell(r, c).value for c in range(10, 16)]  # J=10 to O=15
    if any(v is not None for v in vals):
        print(f"Row {r:2d} | {str(label)[:40]}: {vals}")
wb2.close()
