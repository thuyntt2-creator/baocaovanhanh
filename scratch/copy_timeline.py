import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

src_path = r"C:\Users\lap4all\Downloads\AOP_V2_updated_2.xlsx"
dst_path = r"C:\Users\lap4all\Downloads\V2 AOP_Hang_NTB_T7-T12_2026.xlsx"

if not os.path.exists(src_path):
    print(f"Error: Source file {src_path} does not exist!")
    sys.exit(1)

if not os.path.exists(dst_path):
    print(f"Error: Target file {dst_path} does not exist!")
    sys.exit(1)

print(f"Loading source: {src_path}...")
wb_src = openpyxl.load_workbook(src_path, data_only=False)
if "Timeline tiếp nhận" not in wb_src.sheetnames:
    print("Error: 'Timeline tiếp nhận' sheet not found in source workbook!")
    sys.exit(1)
sheet_src = wb_src["Timeline tiếp nhận"]

print(f"Loading target: {dst_path}...")
try:
    wb_dst = openpyxl.load_workbook(dst_path, data_only=False)
except PermissionError:
    print("\n⚠️ HÃY ĐÓNG FILE Excel 'V2 AOP_Hang_NTB_T7-T12_2026.xlsx' để script có thể ghi dữ liệu nhé!")
    sys.exit(1)

# Get or create target sheet
if "Timeline tiếp nhận" in wb_dst.sheetnames:
    # Clear existing target sheet
    sheet_dst = wb_dst["Timeline tiếp nhận"]
    wb_dst.remove(sheet_dst)
sheet_dst = wb_dst.create_sheet("Timeline tiếp nhận")

# Copy all values, formulas, and basic properties
print("Copying sheet content...")
for r in range(1, sheet_src.max_row + 1):
    for c in range(1, sheet_src.max_column + 1):
        cell_src = sheet_src.cell(r, c)
        cell_dst = sheet_dst.cell(r, c)
        cell_dst.value = cell_src.value
        
        # Copy styles if any (simple font, alignment, border copy)
        if cell_src.has_style:
            cell_dst.font = openpyxl.styles.Font(
                name=cell_src.font.name,
                size=cell_src.font.size,
                bold=cell_src.font.bold,
                italic=cell_src.font.italic,
                color=cell_src.font.color
            )
            cell_dst.alignment = openpyxl.styles.Alignment(
                horizontal=cell_src.alignment.horizontal,
                vertical=cell_src.alignment.vertical,
                wrap_text=cell_src.alignment.wrap_text
            )

# Maintain sheet tab position
# Find original position of Timeline tiếp nhận in target if it was there
sheet_names = wb_dst.sheetnames
# Move Timeline tiếp nhận to the end or back to its place (usually last is fine, or we can move it)
# By default, openpyxl appends. We can move it to index 10 (the 11th tab)
if "Timeline tiếp nhận" in sheet_names:
    target_idx = sheet_names.index("Timeline tiếp nhận")
    current_idx = len(wb_dst._sheets) - 1
    if current_idx != target_idx:
        wb_dst._sheets.insert(target_idx, wb_dst._sheets.pop(current_idx))

print("Saving target file...")
try:
    wb_dst.save(dst_path)
    print("✅ Đã copy sheet 'Timeline tiếp nhận' thành công!")
except PermissionError:
    print("\n⚠️ KHÔNG THỂ LƯU FILE! Vui lòng đóng file Excel 'V2 AOP_Hang_NTB_T7-T12_2026.xlsx' và chạy lại script!")
    sys.exit(1)
