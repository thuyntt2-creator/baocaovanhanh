import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

downloads_dir = r"C:\Users\lap4all\Downloads"
files = [
    os.path.join(downloads_dir, "AOP_MAU_NTB_T7-T12_2026_v4.xlsx"),
    os.path.join(downloads_dir, "AOP_MAU_NTB_T7-T12_2026_v3.xlsx"),
    os.path.join(downloads_dir, "V2 AOP_Hang_NTB_T7-T12_2026.xlsx"),
    os.path.join(downloads_dir, "[V2] AOP_Hang_NTB_T7-T12_2026.xlsx"),
]
files = [f for f in files if os.path.exists(f)]

for f in files:
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        for sname in wb.sheetnames:
            sheet = wb[sname]
            for r in range(1, min(100, sheet.max_row + 1)):
                for c in range(1, min(20, sheet.max_column + 1)):
                    val = sheet.cell(r, c).value
                    if val is not None and any(term in str(val).lower() for term in ["setup mở mới", "di dời"]):
                        print(f"File: {os.path.basename(f)} | Sheet: {sname} | Cell {openpyxl.utils.get_column_letter(c)}{r}: {val}")
                        # print the row values
                        row_vals = [sheet.cell(r, col_idx).value for col_idx in range(1, 15)]
                        print(f"  Row: {row_vals}")
        wb.close()
    except Exception as e:
        print(f"Error {os.path.basename(f)}: {e}")
