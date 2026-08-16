import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\V3 AOP_NTB_T70-T12_2026.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

search_terms = ["setup", "mở mới", "mở", "di dời", "dời", "di doi", "tiện ích", "utilities"]

print("=== Searching in V3 AOP_NTB_T70-T12_2026.xlsx ===")
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(r, c).value
            if val is not None:
                val_str = str(val).lower()
                for term in search_terms:
                    if term in val_str:
                        print(f"Sheet '{sheet_name}' cell {openpyxl.utils.get_column_letter(c)}{r}: {val}")
                        break

wb.close()
