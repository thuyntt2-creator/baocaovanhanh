import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\V3 AOP_NTB_T70-T12_2026.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['Nguồn lực & chi phí']

print("=== Formulas in Nguồn lực & chi phí row 6 and 9 ===")
for r in [6, 9]:
    for c in range(1, 9):
        val = sheet.cell(r, c).value
        col = openpyxl.utils.get_column_letter(c)
        print(f"  Cell {col}{r}: {val}")

wb.close()
