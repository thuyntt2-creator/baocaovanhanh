import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

tg_dir = r"C:\Users\lap4all\Downloads\Telegram Desktop"

def inspect_file(filepath):
    print(f"\n======================================")
    print(f"FILE: {os.path.basename(filepath)}")
    print(f"======================================")
    
    wb = openpyxl.load_workbook(filepath, data_only=False)
    sheetnames = wb.sheetnames
    print("Sheets:", sheetnames)
    
    # Try to find input sheet
    input_sheet = None
    for name in sheetnames:
        if "Input" in name or "input" in name or "DCL" in name or "TNB" in name:
            input_sheet = wb[name]
            print(f"Using sheet: {name}")
            break
            
    if not input_sheet:
        input_sheet = wb.active
        print(f"Using active sheet: {input_sheet.title}")
        
    cols = ['D', 'E', 'F', 'G', 'H', 'I']
    print("\n--- Formulas (data_only=False) ---")
    for r in range(41, 46):
        label = input_sheet.cell(r, 2).value or input_sheet.cell(r, 1).value or f"Row {r}"
        print(f"Row {r:2d} | {str(label)[:35]:<35}")
        for col in cols:
            print(f"  {col}: {input_sheet[f'{col}{r}'].value}")
            
    wb.close()
    
    # Read values
    wb_val = openpyxl.load_workbook(filepath, data_only=True)
    sheet_val = wb_val[input_sheet.title]
    print("\n--- Values (data_only=True) ---")
    for r in range(41, 46):
        label = sheet_val.cell(r, 2).value or sheet_val.cell(r, 1).value or f"Row {r}"
        vals = [sheet_val[f"{c}{r}"].value for c in cols]
        print(f"Row {r:2d} | {str(label)[:35]:<35} | {vals}")
    wb_val.close()

inspect_file(os.path.join(tg_dir, "CRC_DCL_Input_FLM_2026.xlsx"))
inspect_file(os.path.join(tg_dir, "TNB_Input_FLM_CRC.xlsx"))
