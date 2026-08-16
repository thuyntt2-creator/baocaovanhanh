import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_files = [
    r"C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v18.xlsx",
    r"C:\Users\lap4all\Downloads\V2 AOP_Hang_NTB_T7-T12_2026.xlsx",
    r"C:\Users\lap4all\Downloads\V2 AOP_Hang_NTB_T7-T12_2026_new.xlsx",
    r"C:\Users\lap4all\Downloads\[V2] AOP_Hang_NTB_T7-T12_2026.xlsx",
    r"C:\Users\lap4all\Downloads\Copy of AOP_Hang_NTB_T7-T12_2026.xlsx",
]

for path in excel_files:
    if os.path.exists(path):
        print(f"\n=== FILE: {os.path.basename(path)} ===")
        wb = openpyxl.load_workbook(path, data_only=True)
        print("Các sheet:", wb.sheetnames)
        
        # Tìm sheet nguồn lực
        nlcp_sheet = None
        for sname in wb.sheetnames:
            if "nguồn lực" in sname.lower() or "nlcp" in sname.lower():
                nlcp_sheet = wb[sname]
                break
        
        if nlcp_sheet:
            print(f"Đọc sheet: {nlcp_sheet.title}")
            # Đọc dòng tiêu đề (tháng T7 -> T12)
            # In thử 30 dòng đầu, 8 cột đầu
            for r in range(1, 35):
                row_vals = [nlcp_sheet.cell(r, c).value for c in range(1, 9)]
                # Bỏ qua dòng trống
                if any(x is not None for x in row_vals):
                    print(f"  Row {r}: {row_vals}")
        else:
            print("Không tìm thấy sheet liên quan đến Nguồn lực & Chi phí")
    else:
        print(f"File không tồn tại: {path}")
