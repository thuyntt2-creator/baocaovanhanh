import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['Chi phí FLM']

print("=== Evaluated Values in 'Chi phí FLM' sheet ===")
for r in range(25, 38):
    row_vals = [sheet.cell(r, c).value for c in range(1, 16)]
    if any(v is not None for v in row_vals):
        row_lbl = sheet.cell(r, 1).value or sheet.cell(r, 2).value or f"Row {r}"
        print(f"Row {r:2d} | {str(row_lbl)[:45]:<45} | {row_vals[2:]}")

wb.close()
