import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r'C:\Users\lap4all\Downloads\BaoCao_Tuan_NTB_W32_2026.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

def fmt_val(v, is_pct=False, is_int=False, diff_pct=False, diff_int=False):
    if v is None or v == "":
        return "-"
    try:
        val = float(v)
    except (ValueError, TypeError):
        return str(v)
        
    if is_pct:
        pct_val = val * 100 if abs(val) <= 1.0 and val != 0 else val
        if diff_pct:
            sign = "+" if pct_val > 0 else ""
            return f"{sign}{pct_val:.1f}%"
        return f"{pct_val:.1f}%"
    elif is_int:
        if diff_int:
            sign = "+" if val > 0 else ""
            return f"{sign}{int(round(val)):,}"
        return f"{int(round(val)):,}"
    else:
        return f"{val:,.1f}"

print("=== 00_Tong quan ===")
ws = wb['00_Tong quan']
# Table 2 data (R19 to R30)
t2_data = []
for r in range(19, 31):
    metric = ws.cell(r, 1).value
    w29 = ws.cell(r, 2).value
    w30 = ws.cell(r, 3).value
    w31 = ws.cell(r, 4).value
    w32 = ws.cell(r, 5).value
    diff = ws.cell(r, 6).value
    t2_data.append((metric, w29, w30, w31, w32, diff))
    print(f"{metric}: W29={w29}, W30={w30}, W31={w31}, W32={w32}, Δ={diff}")

print("\n=== 01_San luong ===")
ws = wb['01_San luong']
print("Row 5:", [ws.cell(5, c).value for c in range(1, 7)])
print("Row 6 (Full):", [ws.cell(6, c).value for c in range(1, 7)])
print("Row 7 (TTS):", [ws.cell(7, c).value for c in range(1, 7)])
