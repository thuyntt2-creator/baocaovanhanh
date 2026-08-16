import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

# Read v5 Forecast sums for each warehouse
v5_path = r"C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v5.xlsx"
wb_v5 = openpyxl.load_workbook(v5_path, data_only=True)

v5_sums = {}
for m in [7, 8, 9, 10, 11, 12]:
    sheet = wb_v5[f"Forecast T{m}"]
    v5_sums[m] = {}
    for r in range(4, 8):
        name = sheet.cell(r, 1).value
        # Sum of cols
        vals = []
        for c in range(2, sheet.max_column + 1):
            val = sheet.cell(r, c).value
            if val is not None:
                vals.append(val)
        v5_sums[m][name] = sum(vals)

print("=== v5 Forecast Monthly Sums ===")
for m in [7, 8, 9, 10, 11, 12]:
    print(f"Month T{m}: {v5_sums[m]}")

print("\n=== v5 Ratios of Month X to Month 8 ===")
for m in [9, 10, 11, 12]:
    print(f"Month T{m} to T8 Ratios:")
    for name in v5_sums[8].keys():
        ratio_of_sums = v5_sums[m][name] / v5_sums[8][name]
        print(f"  {name}: sum_ratio = {ratio_of_sums:.6f}")
