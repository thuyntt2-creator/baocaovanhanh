import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\V3_AOP_Hang_NTB_T7-T12_2026 mới.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

print("=== External or worksheet references in V3 formulas ===")
for name in wb.sheetnames:
    sheet = wb[name]
    for r in range(1, min(100, sheet.max_row + 1)):
        for c in range(1, min(20, sheet.max_column + 1)):
            val = sheet.cell(r, c).value
            if isinstance(val, str) and val.startswith('='):
                val_str = val.lower()
                # If formula contains brackets (which often means external reference or cell reference)
                if "[" in val_str or "xlsx" in val_str:
                    col_letter = openpyxl.utils.get_column_letter(c)
                    print(f"Sheet '{name}' cell {col_letter}{r}: {val}")

wb.close()
