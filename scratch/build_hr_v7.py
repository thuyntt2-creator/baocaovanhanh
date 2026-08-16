import openpyxl, sys, shutil
from openpyxl.styles import Font, PatternFill, Alignment
sys.stdout.reconfigure(encoding='utf-8')

in_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v6.xlsx'
out_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v7.xlsx'

print("Đang tạo bảng Định biên Nhân sự chi tiết (V7)...")
shutil.copy(in_path, out_path)
wb = openpyxl.load_workbook(out_path)

sheet = wb['Nguồn lực & chi phí']

# Styles
bold_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
sub_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
center_align = Alignment(horizontal='center', vertical='center')

# --- BLOCK 1: NGƯỜI GIAO ---
r = 16
sheet.cell(r, 1).value = "I. ĐỊNH BIÊN NGƯỜI GIAO (TÀI XẾ + PHỤ XE) THEO KHO/BC"
sheet.cell(r, 1).font = Font(bold=True, color="FF0000")
r += 1

# Header
sheet.cell(r, 1).value = "Chỉ tiêu"
months = ['T7', 'T8', 'T9', 'T10', 'T11', 'T12']
cols = ['B', 'C', 'D', 'E', 'F', 'G']
for i, m in enumerate(months):
    c = sheet.cell(r, i + 2)
    c.value = m
    c.font = bold_font
    c.fill = header_fill
    c.alignment = center_align
sheet.cell(r, 1).font = bold_font
sheet.cell(r, 1).fill = header_fill

r += 1
# Rows for 4 Hubs
hub_refs = ["='Kế hoạch T7'!$A$11", "='Kế hoạch T7'!$A$21", "='Kế hoạch T7'!$A$31", "='Kế hoạch T7'!$A$41"]
for i in range(4):
    sheet.cell(r, 1).value = hub_refs[i]
    for c_idx in range(6):
        col_letter = cols[c_idx]
        # Tổng hợp xe row is 6, 7, 8, 9 for the 4 hubs.
        # So it's i + 6
        formula = f"=ROUNDUP('Tổng hợp xe'!{col_letter}{i+6}*'1. Thông số'!$B$20,0)"
        sheet.cell(r, c_idx + 2).value = formula
        sheet.cell(r, c_idx + 2).alignment = center_align
    r += 1

# --- BLOCK 2: XỬ LÝ KHO ---
r += 1
sheet.cell(r, 1).value = "II. ĐỊNH BIÊN NHÂN VIÊN XỬ LÝ KHO (SOX) THEO KHO/BC"
sheet.cell(r, 1).font = Font(bold=True, color="FF0000")
r += 1

# Header
sheet.cell(r, 1).value = "Chỉ tiêu"
for i, m in enumerate(months):
    c = sheet.cell(r, i + 2)
    c.value = m
    c.font = bold_font
    c.fill = header_fill
    c.alignment = center_align
sheet.cell(r, 1).font = bold_font
sheet.cell(r, 1).fill = header_fill

r += 1
# Rows for 4 Hubs
for i in range(4):
    sheet.cell(r, 1).value = hub_refs[i]
    for c_idx, m in enumerate(months):
        end_col = 'AF' if m in ['T7', 'T8', 'T10', 'T12'] else 'AE'
        fc_row = i + 4 # Forecast rows: 4, 5, 6, 7
        formula = f"=ROUNDUP(AVERAGE('Forecast {m}'!B{fc_row}:{end_col}{fc_row})/'1. Thông số'!$B$19,0)"
        sheet.cell(r, c_idx + 2).value = formula
        sheet.cell(r, c_idx + 2).alignment = center_align
    r += 1

wb.save(out_path)
print(f"Xong! Đã lưu: {out_path}")
