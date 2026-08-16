import openpyxl, sys, shutil
from openpyxl.styles import Font, PatternFill, Alignment
sys.stdout.reconfigure(encoding='utf-8')

in_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v6.xlsx'
out_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v16.xlsx'

print("Đang tạo Báo cáo V16 (Fix đơn giá mặt bằng)...")
shutil.copy(in_path, out_path)
wb = openpyxl.load_workbook(out_path)

# Cập nhật đơn giá mặt bằng để khớp số 82.5 triệu của sếp
sheet_ts = wb['1. Thông số']
sheet_ts['B22'] = 150000 

sheet = wb['Nguồn lực & chi phí']

# Styles
bold_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
total_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
center_align = Alignment(horizontal='center', vertical='center')
money_format = '#,##0'
red_font = Font(bold=True, color="FF0000")

months = ['T7', 'T8', 'T9', 'T10', 'T11', 'T12']
cols = ['B', 'C', 'D', 'E', 'F', 'G']
hub_refs = ["='Kế hoạch T7'!$A$11", "='Kế hoạch T7'!$A$21", "='Kế hoạch T7'!$A$31", "='Kế hoạch T7'!$A$41"]

r = 21

# --- BLOCK 1: XỬ LÝ KHO ---
sheet.cell(r, 1).value = "I. ĐỊNH BIÊN NHÂN VIÊN XỬ LÝ KHO (SOX) THEO KHO/BC (Đã cộng 3 NVXL cứng mỗi kho)"
sheet.cell(r, 1).font = red_font
r += 1

sheet.cell(r, 1).value = "Chỉ tiêu"
for i, m in enumerate(months):
    c = sheet.cell(r, i + 2)
    c.value = m; c.font = bold_font; c.fill = header_fill; c.alignment = center_align
sheet.cell(r, 1).font = bold_font; sheet.cell(r, 1).fill = header_fill

r += 1
block1_start_r = r # Dòng 23
for i in range(4):
    sheet.cell(r, 1).value = hub_refs[i]
    for c_idx, m in enumerate(months):
        end_col = 'AF' if m in ['T7', 'T8', 'T10', 'T12'] else 'AE'
        fc_row = i + 4 
        formula = f"=3 + ROUNDUP(AVERAGE('Forecast {m}'!B{fc_row}:{end_col}{fc_row})/'1. Thông số'!$B$19,0)"
        sheet.cell(r, c_idx + 2).value = formula
        sheet.cell(r, c_idx + 2).alignment = center_align
    r += 1

# Tổng cộng Block 1
block1_total_r = r # Dòng 27
sheet.cell(r, 1).value = "TỔNG CỘNG NHÂN SỰ"
sheet.cell(r, 1).font = bold_font; sheet.cell(r, 1).fill = total_fill
for c_idx in range(6):
    c = sheet.cell(r, c_idx + 2)
    c.value = f"=SUM({cols[c_idx]}{block1_start_r}:{cols[c_idx]}{r-1})"
    c.font = bold_font; c.fill = total_fill; c.alignment = center_align
r += 2

# --- BLOCK 2: CHI PHÍ ---
sheet.cell(r, 1).value = "II. TỔNG CHI PHÍ (THUÊ XE + LƯƠNG KHO) THEO KHO/BC (VNĐ)"
sheet.cell(r, 1).font = red_font
r += 1

sheet.cell(r, 1).value = "Chỉ tiêu"
for i, m in enumerate(months):
    c = sheet.cell(r, i + 2)
    c.value = m; c.font = bold_font; c.fill = header_fill; c.alignment = center_align
sheet.cell(r, 1).font = bold_font; sheet.cell(r, 1).fill = header_fill

r += 1
block2_start_r = r # Dòng 31
for i in range(4):
    sheet.cell(r, 1).value = hub_refs[i]
    kh_row = 18 + i*10 
    sox_r = block1_start_r + i 
    
    for c_idx, m in enumerate(months):
        col_letter = cols[c_idx]
        end_col = 'AF' if m in ['T7', 'T8', 'T10', 'T12'] else 'AE'
        
        xe_logic = f"ROUND(AVERAGE('Kế hoạch {m}'!B{kh_row}:{end_col}{kh_row}),0)*'1. Thông số'!$B$23*{col_letter}$4"
        luong_logic = f"{col_letter}{sox_r}*'1. Thông số'!$B$24"
        
        formula = f"={xe_logic} + {luong_logic}"
        
        c = sheet.cell(r, c_idx + 2)
        c.value = formula
        c.number_format = money_format
    r += 1

# Tổng cộng Block 2
sheet.cell(r, 1).value = "TỔNG CỘNG CHI PHÍ 4 HUB"
sheet.cell(r, 1).font = bold_font; sheet.cell(r, 1).fill = total_fill
for c_idx in range(6):
    c = sheet.cell(r, c_idx + 2)
    c.value = f"=SUM({cols[c_idx]}{block2_start_r}:{cols[c_idx]}{r-1})"
    c.font = bold_font; c.fill = total_fill; c.number_format = money_format

# --- OVERWRITE TOPLINE ROWS ---
# Không chèn dòng, xóa công thức cty, ghi đè số liệu 4 hub
for c_idx in range(6):
    # Dòng 11: Người xử lý kho (Thay bằng số tổng từ Bảng Định Biên)
    sheet.cell(11, c_idx+2).value = f"={cols[c_idx]}{block1_total_r}"
    
    # Dòng 17: Chi phí nhân sự xử lý (Tính theo lượng người 4 hub)
    sheet.cell(17, c_idx+2).value = f"={cols[c_idx]}11*'1. Thông số'!$B$24"
    
    # Dòng 18: TỔNG CHI PHÍ (Mặt bằng + Xe + NS 4 Hub)
    sheet.cell(18, c_idx+2).value = f"={cols[c_idx]}15+{cols[c_idx]}16+{cols[c_idx]}17"

wb.save(out_path)
print(f"Xong! Đã lưu: {out_path}")
