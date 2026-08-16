import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

mau_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v5.xlsx'
wb = openpyxl.load_workbook(mau_path, data_only=True)

# Test on BCCK Di Linh in Kế hoạch T7
sheet_kh = wb['Kế hoạch T7']

# Find Di Linh row in Kế hoạch T7
r_di_linh = None
for r in range(4, 50):
    if sheet_kh.cell(r, 1).value == 'BCCK Di Linh':
        r_di_linh = r
        break

if r_di_linh:
    print("--- TRƯỚC KHI TỐI ƯU (Kế hoạch T7 - BCCK Di Linh) ---")
    total_gap = 0
    total_trucks = 0
    total_demand = 0
    
    # Gap is at r_di_linh + 9 (Gap thiếu thừa)
    # Trucks is at r_di_linh + 7
    # Demand is at r_di_linh + 4 (Tổng nhu cầu thực tế)
    r_gap = r_di_linh + 9
    r_truck = r_di_linh + 7
    r_demand = r_di_linh + 4
    
    for day in range(1, 32):
        col = day + 1
        gap = sheet_kh.cell(r_gap, col).value or 0
        trucks = sheet_kh.cell(r_truck, col).value or 0
        demand = sheet_kh.cell(r_demand, col).value or 0
        
        total_gap += gap
        total_trucks += trucks
        total_demand += demand
        # print(f"Ngày {day}: Demand={demand} -> Trucks={trucks}, GAP={gap}")
        
    print(f"Tổng Demand T7: {total_demand}")
    print(f"Tổng Trucks T7: {total_trucks}")
    print(f"Tổng GAP T7: {total_gap}")
