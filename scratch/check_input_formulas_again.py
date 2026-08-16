import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['NTB – Input']

print("=== Formulas in NTB - Input row 41, 42, 43 ===")
for r in [41, 42, 43]:
    for c in range(4, 10):
        val = sheet.cell(r, c).value
        col = openpyxl.utils.get_column_letter(c)
        print(f"  Cell {col}{r}: {val}")

wb.close()
