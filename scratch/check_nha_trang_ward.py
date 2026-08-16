import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\lap4all\Documents\Auto report\NTB_Phan_Tuyen_Hanh_Chinh_Quy_Hoach_Chi_Tiet.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Sheet1']

headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column+1)]

for r in range(2, sheet.max_row+1):
    row = {headers[i]: sheet.cell(r, i+1).value for i in range(len(headers)) if headers[i]}
    xa_moi = str(row.get('Tên Xã mới') or '').strip()
    if 'Nha Trang' in xa_moi:
        print(f"Row {r}: Xã mới='{xa_moi}', BC đề xuất='{row.get('Tên Bưu cục giao mới đề xuất')}', Đề xuất='{row.get('Đánh giá & Phương án đề xuất')}'")
