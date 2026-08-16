import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\V3_AOP_Hang_NTB_T7-T12_2026 mới.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['Bưu cục detai;']

# Group definitions from calculate_test_sums.py
groups = {
    'BCCK Nha Trang': [
        "Bưu Cục 06 Lê Hồng Phong-TP.Nha Trang-Khánh Hòa",
        "Bưu Cục 195 Đường 2/4-Nha Trang-Khánh Hòa",
        "Bưu Cục 229 Phước Long-Nam Nha Trang-Khánh Hòa",
        "Bưu Cục 40A Yết Kiêu-Nha Trang-Khánh Hòa",
        "Bưu Cục 466 Đường 23/10-Nha Trang-Khánh Hòa",
        "Bưu Cục Đường 35 Hà Quang 1-Xã Nam Nha Trang-Khánh Hòa",
        "Bưu Cục Phước Đồng-Nha Trang-Khánh Hoà"
    ],
    'BCCK Di Linh': [
        "Bưu Cục 1322 Hùng Vương-Di Linh-Lâm Đồng",
        "(LDO) Đinh Văn Lâm Hà",
        "(LDO) Tân Hà Lâm Hà",
        "Bưu Cục 231 Thôn 1-Xã Hòa Ninh-Lâm Đồng",
        "(LDO) Nam Ban Lâm Hà"
    ],
    'BCCK Đơn Dương': [
        "(LDO) Đơn Dương",
        "(LDO) Hiệp Thạnh"
    ],
    'BCCK Đức Linh': [
        "(BTH) Đức Linh"
    ]
}

# Clean group names
group_names_set = {}
for g, items in groups.items():
    group_names_set[g] = [name.strip().lower() for name in items]

# Read bưu cục detail sheet
bc_vols = {}
for r in range(3, sheet.max_row + 1):
    bc_name = sheet.cell(r, 2).value
    if bc_name is not None:
        bc_vols[bc_name.strip().lower()] = {
            'T7': sheet.cell(r, 4).value or 0.0,
            'T8': sheet.cell(r, 5).value or 0.0,
            'T9': sheet.cell(r, 6).value or 0.0,
            'T10': sheet.cell(r, 7).value or 0.0,
            'T11': sheet.cell(r, 8).value or 0.0,
            'T12': sheet.cell(r, 9).value or 0.0
        }

print("=== Summing Volumes of the 4 Hubs in V3 ===")
monthly_totals = {m: 0.0 for m in ['T7', 'T8', 'T9', 'T10', 'T11', 'T12']}
for gname, items in group_names_set.items():
    print(f"\n{gname}:")
    g_totals = {m: 0.0 for m in ['T7', 'T8', 'T9', 'T10', 'T11', 'T12']}
    for name in items:
        if name in bc_vols:
            for m in g_totals:
                g_totals[m] += bc_vols[name][m]
        else:
            print(f"  ⚠️ Warning: {name} not found in sheet!")
    
    # print monthly values
    print("  " + " | ".join(f"{m}: {g_totals[m]:.1f}" for m in ['T7', 'T8', 'T9', 'T10', 'T11', 'T12']))
    for m in monthly_totals:
        monthly_totals[m] += g_totals[m]

print("\n=== TOTAL SUM of 4 Hubs ===")
print("  " + " | ".join(f"{m}: {monthly_totals[m]:.1f}" for m in ['T7', 'T8', 'T9', 'T10', 'T11', 'T12']))

wb.close()
