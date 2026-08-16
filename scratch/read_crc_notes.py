import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['NTB – Input']

print("=== Tất cả nội dung trong cột J (Note/Nhận xét CRC) ===\n")
for r in range(1, sheet.max_row + 1):
    val_j = sheet.cell(r, 10).value  # Cột J
    if val_j is not None:
        row_label = sheet.cell(r, 2).value or sheet.cell(r, 1).value or f"Row {r}"
        print(f"Row {r:2d} | {str(row_label)[:40]} | {val_j}")

print("\n\n=== Tất cả nội dung cột K (nếu có note thêm) ===\n")
for r in range(1, sheet.max_row + 1):
    val_k = sheet.cell(r, 11).value  # Cột K
    if val_k is not None:
        row_label = sheet.cell(r, 2).value or sheet.cell(r, 1).value or f"Row {r}"
        print(f"Row {r:2d} | {str(row_label)[:40]} | {val_k}")

wb.close()
