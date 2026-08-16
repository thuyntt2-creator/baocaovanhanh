import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB new.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
print("All sheets:", wb.sheetnames)

for sname in wb.sheetnames:
    sheet = wb[sname]
    print(f"\n=== {sname} (row={sheet.max_row}, col={sheet.max_column}) ===")
    for r in range(1, 5):
        row_vals = [sheet.cell(r, c).value for c in range(1, min(18, sheet.max_column + 1))]
        if any(v is not None for v in row_vals):
            print(f"  Row {r}: {row_vals}")
wb.close()
