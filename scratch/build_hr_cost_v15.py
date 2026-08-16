import openpyxl, sys, shutil
from openpyxl.styles import Font, PatternFill, Alignment
sys.stdout.reconfigure(encoding='utf-8')

in_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v6.xlsx'
out_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v15.xlsx'

print("Đang tạo Báo cáo V15 (Xé lẻ 3 dòng nhân sự)...")
shutil.copy(in_path, out_path)
wb = openpyxl.load_workbook(out_path)
sheet = wb['Nguồn lực & chi phí']

# Chèn thêm 2 dòng để đủ không gian cho 3 dòng nhân sự (Dòng 17, 18, 19)
sheet.insert_rows(18, 2)

r = 23
months = ['T7', 'T8', 'T9', 'T10', 'T11', 'T12']
cols = ['B', 'C', 'D', 'E', 'F', 'G']
hub_refs = ["='Kế hoạch T7'!$A$11", "='Kế hoạch T7'!$A$21", "='Kế hoạch T7'!$A$31", "='Kế hoạch T7'!$A$41"]

# Styles
bold_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
total_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
center_align = Alignment(horizontal='center', vertical='center')
money_format = '#,##0'
red_font = Font(bold=True, color="FF0000")

# --- BLOCK 1: XỬ LÝ KHO ---
sheet.cell(r, 1).value = "I. ĐỊNH BIÊN NHÂN VIÊN XỬ LÝ KHO (SOX) THEO KHO/BC (Đã cộng 3 NVXL cứng mỗi kho)"
sheet.cell(r, 1).font = red_font
r += 1

sheet.cell(r, 1).value = "Chỉ tiêu"
for i, m in enumerate(months):
    c = sheet.cell(r, i + 2)
    c.value = m; c.font = bold_font; c.fill = header_fill; c.alignment = center_align
sheet.cell(r, 1).font = bold_font; sheet.cell(r, 1).fill = header_fill

r += 1
block1_start_r = r # Dòng 25
for i in range(4):
    sheet.cell(r, 1).value = hub_refs[i]
    for c_idx, m in enumerate(months):
        end_col = 'AF' if m in ['T7', 'T8', 'T10', 'T12'] else 'AE'
        fc_row = i + 4 
        formula = f"=3 + ROUNDUP(AVERAGE('Forecast {m}'!B{fc_row}:{end_col}{fc_row})/'1. Thông số'!$B$19,0)"
        sheet.cell(r, c_idx + 2).value = formula
        sheet.cell(r, c_idx + 2).alignment = center_align
    r += 1

# Tổng cộng Block 1
block1_total_r = r # Dòng 29
sheet.cell(r, 1).value = "TỔNG CỘNG NHÂN SỰ"
sheet.cell(r, 1).font = bold_font; sheet.cell(r, 1).fill = total_fill
for c_idx in range(6):
    c = sheet.cell(r, c_idx + 2)
    c.value = f"=SUM({cols[c_idx]}{block1_start_r}:{cols[c_idx]}{r-1})"
    c.font = bold_font; c.fill = total_fill; c.alignment = center_align
r += 2

# --- BLOCK 2: CHI PHÍ ---
sheet.cell(r, 1).value = "II. TỔNG CHI PHÍ (THUÊ XE + LƯƠNG KHO) THEO KHO/BC (VNĐ)"
sheet.cell(r, 1).font = red_font
r += 1

sheet.cell(r, 1).value = "Chỉ tiêu"
for i, m in enumerate(months):
    c = sheet.cell(r, i + 2)
    c.value = m; c.font = bold_font; c.fill = header_fill; c.alignment = center_align
sheet.cell(r, 1).font = bold_font; sheet.cell(r, 1).fill = header_fill

r += 1
block2_start_r = r # Dòng 33
for i in range(4):
    sheet.cell(r, 1).value = hub_refs[i]
    kh_row = 18 + i*10 
    sox_r = block1_start_r + i 
    
    for c_idx, m in enumerate(months):
        col_letter = cols[c_idx]
        end_col = 'AF' if m in ['T7', 'T8', 'T10', 'T12'] else 'AE'
        
        xe_logic = f"ROUND(AVERAGE('Kế hoạch {m}'!B{kh_row}:{end_col}{kh_row}),0)*'1. Thông số'!$B$23*{col_letter}$4"
        luong_logic = f"{col_letter}{sox_r}*'1. Thông số'!$B$24"
        
        formula = f"={xe_logic} + {luong_logic}"
        
        c = sheet.cell(r, c_idx + 2)
        c.value = formula
        c.number_format = money_format
    r += 1

# Tổng cộng Block 2
sheet.cell(r, 1).value = "TỔNG CỘNG CHI PHÍ 4 HUB"
sheet.cell(r, 1).font = bold_font; sheet.cell(r, 1).fill = total_fill
for c_idx in range(6):
    c = sheet.cell(r, c_idx + 2)
    c.value = f"=SUM({cols[c_idx]}{block2_start_r}:{cols[c_idx]}{r-1})"
    c.font = bold_font; c.fill = total_fill; c.number_format = money_format

# --- OVERWRITE TOPLINE ROWS ---

# Dòng 11: Người xử lý kho (Thay bằng tổng quân số thực tế)
for c_idx in range(6):
    sheet.cell(11, c_idx+2).value = f"={cols[c_idx]}{block1_total_r}"

# Dòng 17: Chi phí NV xử lý kho (4 người / 4 BC)
sheet.cell(17, 1).value = "Chi phí NV xử lý kho (4 người / 4 BC)"
for c_idx in range(6):
    sheet.cell(17, c_idx+2).value = f"=4*'1. Thông số'!$B$24"
    sheet.cell(17, c_idx+2).number_format = money_format

# Dòng 18: Chi phí NV quản lý & backup (4 người / 4 BC)
sheet.cell(18, 1).value = "Chi phí NV quản lý & backup (4 người / 4 BC)"
for c_idx in range(6):
    sheet.cell(18, c_idx+2).value = f"=4*'1. Thông số'!$B$24"
    sheet.cell(18, c_idx+2).number_format = money_format

# Dòng 19: Chi phí NV giao hàng (x người / 4 BC)
# Dùng hàm Excel nối chuỗi để chữ trên cột A tự động thay đổi theo quân số Tháng 7 (Cột B)
sheet.cell(19, 1).value = f'="Chi phí NV giao hàng (" & (B11-8) & " người / 4 BC)"'
for c_idx in range(6):
    # Cost formula
    sheet.cell(19, c_idx+2).value = f"=({cols[c_idx]}11-8)*'1. Thông số'!$B$24"
    sheet.cell(19, c_idx+2).number_format = money_format

# Dòng 20: TỔNG CHI PHÍ
for c_idx in range(6):
    sheet.cell(20, c_idx+2).value = f"={cols[c_idx]}15+{cols[c_idx]}16+{cols[c_idx]}17+{cols[c_idx]}18+{cols[c_idx]}19"
    sheet.cell(20, c_idx+2).number_format = money_format

wb.save(out_path)
print(f"Xong! Đã lưu: {out_path}")
