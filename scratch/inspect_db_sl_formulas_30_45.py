import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['Định biên & Sản lượng']

print("=== Formulas in 'Định biên & Sản lượng' rows 30-45 ===")
for r in range(30, 46):
    label = sheet.cell(r, 1).value or sheet.cell(r, 2).value or f"Row {r}"
    c_val = sheet[f"C{r}"].value
    print(f"Row {r:2d} | {str(label)[:35]:<35} | C: {c_val}")
wb.close()
