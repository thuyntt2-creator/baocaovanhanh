import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

for name in ['Định biên & Sản lượng', 'Chi phí FLM']:
    sheet = wb[name]
    print(f"\n=== Sheet: {name} (max_row: {sheet.max_row}, max_col: {sheet.max_column}) ===")
    for r in range(1, 40):
        row_vals = [sheet.cell(r, c).value for c in range(1, 15)]
        if any(v is not None for v in row_vals):
            print(f"Row {r:2d}: {row_vals}")

wb.close()
