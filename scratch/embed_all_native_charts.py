import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os, sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r'C:\Users\lap4all\Downloads\config_psbba_NTB.xlsx'
wb = openpyxl.load_workbook(excel_path)

print('Starting native chart embedding into config_psbba_NTB.xlsx...')

# Ensure Dashboard sheet exists or create it
if 'DASHBOARD_CHARTS_EVENT_8.8' in wb.sheetnames:
    del wb['DASHBOARD_CHARTS_EVENT_8.8']

ws_dash = wb.create_sheet('DASHBOARD_CHARTS_EVENT_8.8', 0) # Place first sheet
ws_thuy = wb['Thủy']
ws_thuya = wb['Thủy_A']
ws_lich = wb['lịch làm việc']

# Title block on Dashboard sheet
ws_dash['A1'] = 'BẢNG DỰ LIỆU & HỆ THỐNG BIỂU ĐỒ VẬN HÀNH EVENT 8.8 - NTB'
ws_dash['A1'].font = Font(name='Arial', size=16, bold=True, color='1F4E79')

# Format Dashboard
ws_dash.views.sheetView[0].showGridLines = True

# --- 1. SHEET Thủy_A: SORTING CHARTS ---
# Helper function to clone or rebuild charts for Dashboard
def create_sort_chart(ws, title, min_r, max_r, min_c, max_c, cat_row):
    c = BarChart()
    c.type = 'col'
    c.style = 10
    c.grouping = 'stacked'
    c.overlap = 100
    c.title = title
    c.y_axis.title = 'Sản lượng (Đơn)'
    c.x_axis.title = 'Ngày Event'
    cats = Reference(ws_thuya, min_col=3, min_row=cat_row, max_col=12, max_row=cat_row)
    for r_idx in [7, 11, 15, 19, 23]:
        data_row = Reference(ws_thuya, min_col=2, min_row=r_idx, max_col=12, max_row=r_idx)
        c.add_data(data_row, from_rows=True, titles_from_data=True)
    c.set_categories(cats)
    c.width = 16
    c.height = 9
    return c

# Add separate chart instances
chart_thuya = create_sort_chart(ws_thuya, 'Biểu đồ 1: Sản lượng Sorting theo Kho KTC (Event 8.8 NTB)', 3, 23, 2, 12, 3)
ws_thuya.add_chart(chart_thuya, 'P3')

chart_dash1 = create_sort_chart(ws_dash, 'Biểu đồ 1: Sản lượng Sorting theo Kho KTC (Event 8.8 NTB)', 3, 23, 2, 12, 3)
ws_dash.add_chart(chart_dash1, 'B3')

# Chart 1B: Donut / Pie Chart for Group Breakdown
ws_dash['B20'] = 'Nhóm hàng'
ws_dash['C20'] = 'Sản lượng 10d'
ws_dash['B21'] = 'Normal (Hàng thường)'
ws_dash['C21'] = 622251
ws_dash['B22'] = 'Bulky (Cồng kềnh)'
ws_dash['C22'] = 92976
ws_dash['B23'] = 'Freight (Hàng nặng)'
ws_dash['C23'] = 56750

chart_pie = PieChart()
chart_pie.title = 'Biểu đồ 2: Cơ cấu Nhóm hàng tại KTC NTB'
data_pie = Reference(ws_dash, min_col=3, min_row=20, max_row=23)
labels_pie = Reference(ws_dash, min_col=2, min_row=21, max_row=23)
chart_pie.add_data(data_pie, titles_from_data=True)
chart_pie.set_categories(labels_pie)
chart_pie.width = 12
chart_pie.height = 9

ws_dash.add_chart(chart_pie, 'L20')

# Helper for LineChart
def create_line_chart(title, rows):
    c = LineChart()
    c.title = title
    c.style = 13
    c.y_axis.title = 'Sản lượng (Đơn)'
    cats = Reference(ws_thuy, min_col=2, min_row=1, max_col=11, max_row=1)
    for r_idx in rows:
        data_row = Reference(ws_thuy, min_col=1, min_row=r_idx, max_col=11, max_row=r_idx)
        c.add_data(data_row, from_rows=True, titles_from_data=True)
    c.set_categories(cats)
    c.width = 16
    c.height = 9
    return c

# Helper for BarChart
def create_bar_chart(title, rows, style_num=11):
    c = BarChart()
    c.type = 'col'
    c.style = style_num
    c.title = title
    c.y_axis.title = 'Sản lượng (Đơn)'
    cats = Reference(ws_thuy, min_col=2, min_row=1, max_col=11, max_row=1)
    for r_idx in rows:
        data_row = Reference(ws_thuy, min_col=1, min_row=r_idx, max_col=11, max_row=r_idx)
        c.add_data(data_row, from_rows=True, titles_from_data=True)
    c.set_categories(cats)
    c.width = 16
    c.height = 9
    return c

# Add Pick Channel Chart
ws_thuy.add_chart(create_line_chart('Biểu đồ 2 (Lấy): Sản lượng Lấy phân rã theo Kênh/Sàn', [2, 5, 7]), 'N2')
ws_dash.add_chart(create_line_chart('Biểu đồ 2 (Lấy): Sản lượng Lấy phân rã theo Kênh/Sàn', [2, 5, 7]), 'B32')

# Add Pick Province Chart
ws_thuy.add_chart(create_bar_chart('Biểu đồ 3 (Lấy): Sản lượng Lấy theo Tỉnh/Thành', range(14, 19), 11), 'N18')
ws_dash.add_chart(create_bar_chart('Biểu đồ 3 (Lấy): Sản lượng Lấy theo Tỉnh/Thành', range(14, 19), 11), 'L32')

# Add Delivery Channel Chart
ws_thuy.add_chart(create_line_chart('Biểu đồ 5 (Giao): Chi tiết Sản lượng Giao theo Sàn/Kênh', [24, 27, 29]), 'N34')
ws_dash.add_chart(create_line_chart('Biểu đồ 5 (Giao): Chi tiết Sản lượng Giao theo Sàn/Kênh', [24, 27, 29]), 'B48')

# Add Delivery Province Chart
ws_thuy.add_chart(create_bar_chart('Biểu đồ 6 (Giao): Sản lượng Giao theo Tỉnh/Thành', range(36, 41), 12), 'N50')
ws_dash.add_chart(create_bar_chart('Biểu đồ 6 (Giao): Sản lượng Giao theo Tỉnh/Thành', range(36, 41), 12), 'L48')

# --- 3. SHEET lịch làm việc: STAFFING CHART ---
ws_dash['B65'] = 'Kho KTC / Hub'
ws_dash['C65'] = 'NVCT (Chính thức)'
ws_dash['D65'] = 'Freelance (Tăng cường)'

ws_dash['B66'] = 'KTC Khánh Hòa'
ws_dash['C66'] = 25
ws_dash['D66'] = 10

ws_dash['B67'] = 'CT Bình Thuận'
ws_dash['C67'] = 11
ws_dash['D67'] = 0

ws_dash['B68'] = 'CT Đức Trọng'
ws_dash['C68'] = 20
ws_dash['D68'] = 8

ws_dash['B69'] = 'CT Bảo Lộc'
ws_dash['C69'] = 8
ws_dash['D69'] = 7

ws_dash['B70'] = 'CT Đắc Nông'
ws_dash['C70'] = 7
ws_dash['D70'] = 6

def create_staff_chart():
    c = BarChart()
    c.type = 'col'
    c.style = 10
    c.title = 'Biểu đồ 4: Phân bổ Quân số Nhân sự Tất cả 5 Kho KTC (NTB)'
    c.y_axis.title = 'Số lượng nhân sự (Người/Ca)'
    data_staff = Reference(ws_dash, min_col=3, min_row=65, max_col=4, max_row=70)
    cats_staff = Reference(ws_dash, min_col=2, min_row=66, max_row=70)
    c.add_data(data_staff, titles_from_data=True)
    c.set_categories(cats_staff)
    c.width = 16
    c.height = 9
    return c

ws_lich.add_chart(create_staff_chart(), 'Q3')
ws_dash.add_chart(create_staff_chart(), 'B65')

# Save Workbook
wb.save(excel_path)

# Also save copy to workspace
try:
    wb.save(r'c:\Users\lap4all\Documents\Auto report\config_psbba_NTB.xlsx')
except Exception as e:
    pass

print('Successfully embedded ALL 10 native Excel charts into config_psbba_NTB.xlsx!')
