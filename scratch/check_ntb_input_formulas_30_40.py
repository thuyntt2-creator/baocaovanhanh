import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['NTB – Input']

print("=== Formulas in NTB – Input rows 30-40 ===")
cols = ['D', 'E', 'F', 'G', 'H', 'I']
for r in range(30, 40):
    label = sheet.cell(r, 2).value or sheet.cell(r, 1).value or f"Row {r}"
    print(f"Row {r:2d} | {str(label)[:35]:<35}")
    for c in cols:
        val = sheet[f"{c}{r}"].value
        print(f"  {c}: {val}")
wb.close()
