import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

v2_path = r"C:\Users\lap4all\Downloads\[V2] AOP_Hang_NTB_T7-T12_2026.xlsx"
wb = openpyxl.load_workbook(v2_path, data_only=False)
sheet = wb['1. Thông số']

print("=== [V2] 1. Thông số ===")
for r in range(1, sheet.max_row + 1):
    vals = [sheet.cell(r, c).value for c in range(1, 10)]
    if any(v is not None for v in vals):
        print(f"Row {r:2d}: {vals}")

