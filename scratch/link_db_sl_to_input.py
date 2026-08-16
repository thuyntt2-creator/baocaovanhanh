import openpyxl
import win32com.client
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

downloads_dir = r"C:\Users\lap4all\Downloads"
files_to_fix = [
    os.path.join(downloads_dir, "NTB_Input_Da_Dien_FLM_CRC_2.xlsx"),
    os.path.join(downloads_dir, "NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx")
]

cols_map = {
    'I': 'D',
    'J': 'E',
    'K': 'F',
    'L': 'G',
    'M': 'H',
    'N': 'I'
}

for filepath in files_to_fix:
    if not os.path.exists(filepath):
        continue
        
    print(f"\nLinking 'Định biên & Sản lượng' to 'NTB – Input' in: {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath, data_only=False)
    sheet_db = wb['Định biên & Sản lượng']
    
    # Fill formulas in Định biên & Sản lượng row 72 to 77
    for db_col, input_col in cols_map.items():
        sheet_db[f"{db_col}72"] = f"='NTB – Input'!{input_col}8/4"
        sheet_db[f"{db_col}73"] = f"='NTB – Input'!{input_col}9/4"
        sheet_db[f"{db_col}74"] = f"='NTB – Input'!{input_col}10/4"
        sheet_db[f"{db_col}75"] = f"='NTB – Input'!{input_col}11/4"
        sheet_db[f"{db_col}76"] = f"='NTB – Input'!{input_col}12/4"
        sheet_db[f"{db_col}77"] = f"='NTB – Input'!{input_col}13/4"
        
    wb.save(filepath)
    wb.close()
    
    # Recalculate via Excel COM
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        workbook = excel.Workbooks.Open(filepath)
        workbook.Save()
        workbook.Close()
        print("Recalculation successful!")
    except Exception as e:
        print(f"Error during recalculation: {e}")
    finally:
        excel.Quit()
        
# Verify values in NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx
con_thieu_path = os.path.join(downloads_dir, "NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx")
if os.path.exists(con_thieu_path):
    wb = openpyxl.load_workbook(con_thieu_path, data_only=True)
    sheet = wb['NTB – Input']
    cols = ['D', 'E', 'F', 'G', 'H', 'I']
    print("\n=== VERIFIED VALUES in NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx ===")
    for r in [7, 41, 42, 43, 44, 45]:
        label = sheet.cell(r, 2).value or sheet.cell(r, 1).value
        vals = [sheet.cell(r, openpyxl.utils.column_index_from_string(c)).value for c in cols]
        print(f"Row {r:2d} | {str(label)[:35]:<35} | {vals}")
    wb.close()
