import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v18.xlsx"

wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Nguồn lực & chi phí']

print("=== IN 20 CỘT ĐẦU CỦA SHEET NGUỒN LỰC & CHI PHÍ ===")
for r in [2, 10, 15, 17, 19, 20]:
    row_text = f"Row {r:02d} ({sheet.cell(row=r, column=1).value}):\n  "
    for c in range(1, 25):
        val = sheet.cell(row=r, column=c).value
        if val is not None:
            row_text += f"Col_{c}={val} | "
    print(row_text)
wb.close()
