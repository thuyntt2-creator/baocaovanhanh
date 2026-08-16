import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

for sname in wb.sheetnames:
    sheet = wb[sname]
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(r, c).value
            if val == 905 or val == 905.0:
                print(f"Found 905 in Sheet: '{sname}', Cell: {openpyxl.utils.get_column_letter(c)}{r}")
wb.close()
