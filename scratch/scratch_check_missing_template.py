import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    sys.exit(1)

wb = openpyxl.load_workbook(file_path, data_only=True)
for sname in wb.sheetnames:
    sheet = wb[sname]
    print(f"\n--- Sheet: {sname} (max_row: {sheet.max_row}, max_col: {sheet.max_column}) ---")
    missing_cells = []
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(r, c).value
            if val is not None:
                val_str = str(val).strip()
                if "thiếu" in val_str.lower() or "thieu" in val_str.lower():
                    missing_cells.append((r, c, val))
    print(f"Found {len(missing_cells)} cells with 'Thiếu':")
    for r, c, val in missing_cells[:10]:
        header = sheet.cell(2, c).value or sheet.cell(1, c).value or f"Col {c}"
        row_label = sheet.cell(r, 2).value or sheet.cell(r, 1).value or f"Row {r}"
        print(f"  Cell {openpyxl.utils.get_column_letter(c)}{r}: RowLabel='{row_label}', Header='{header}', Value='{val}'")

wb.close()
