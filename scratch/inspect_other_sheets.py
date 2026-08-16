import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

xlsx_aop_path = r"C:\Users\lap4all\Downloads\AOP_Hang_NTB_T7-T12_2026.xlsx"

wb = openpyxl.load_workbook(xlsx_aop_path, data_only=False)

for sname in wb.sheetnames:
    if sname in ['Hướng dẫn', '0. Thông tin vùng', '1. Thông số']:
        continue
    sheet = wb[sname]
    print(f"\n==================== Sheet: {sname} (Rows: {sheet.max_row}, Cols: {sheet.max_column}) ====================")
    # Print the first 10 rows and up to 12 columns
    for r_idx in range(1, min(sheet.max_row + 1, 15)):
        row_vals = []
        for c_idx in range(1, min(sheet.max_column + 1, 15)):
            cell = sheet.cell(r_idx, c_idx)
            val = cell.value
            # If value is a formula, show both formula and if possible we'll know it's a formula
            if isinstance(val, str) and val.startswith('='):
                row_vals.append(f"F:{val}")
            else:
                row_vals.append(val)
        if any(v is not None for v in row_vals):
            print(f"Row {r_idx:2d}: {row_vals}")

