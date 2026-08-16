import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

filepath = r"C:\Users\lap4all\Downloads\Telegram Desktop\7. NTB_2026.xlsx"
wb = openpyxl.load_workbook(filepath, data_only=True)
sheet = wb['7. NTB_2026']

months = ['T7', 'T8', 'T9', 'T10', 'T11', 'T12']
cols = [14, 15, 16, 17, 18, 19]  # N to S

print("=== CALCULATIONS ===")
for idx, m in enumerate(months):
    col = cols[idx]
    col_letter = openpyxl.utils.get_column_letter(col)
    
    num_bc = sheet.cell(10, col).value
    utilities = sheet.cell(32, col).value
    dien = sheet.cell(33, col).value
    nuoc = sheet.cell(34, col).value
    internet = sheet.cell(35, col).value
    
    ratio = utilities / num_bc if num_bc else 0
    ratio_tr = ratio / 1000000
    
    print(f"Month {m} (Col {col_letter}):")
    print(f"  Số lượng BC (Row 10): {num_bc}")
    print(f"  Total Utilities (Row 32): {utilities:,.0f} VNĐ")
    print(f"    ↳ Điện: {dien:,.0f} VNĐ")
    print(f"    ↳ Nước: {nuoc:,.0f} VNĐ")
    print(f"    ↳ Internet: {internet:,.0f} VNĐ")
    print(f"  Ratio (Utilities/BC): {ratio:,.2f} VNĐ/BC")
    print(f"  Ratio in Triệu VNĐ: {ratio_tr:.6f}")
wb.close()
