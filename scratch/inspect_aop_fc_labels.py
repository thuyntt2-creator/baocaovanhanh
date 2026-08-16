import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v18.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['Nguồn lực & chi phí']

print("=== Nguồn lực & chi phí (v18) ===")
for r in range(1, sheet.max_row + 1):
    row_vals = [sheet.cell(r, c).value for c in range(1, 10)]
    if any(v is not None for v in row_vals):
        cols_str = []
        for c_idx, val in enumerate(row_vals):
            if val is not None:
                col_letter = openpyxl.utils.get_column_letter(c_idx + 1)
                cols_str.append(f"{col_letter}: {repr(val)}")
        print(f"Row {r:2d}: " + " | ".join(cols_str))
