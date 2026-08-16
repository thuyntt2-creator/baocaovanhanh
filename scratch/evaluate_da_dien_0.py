import win32com.client
import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

filepath = r"C:\Users\lap4all\Downloads\NTB_Input_Da_Dien_FLM_CRC.xlsx"

# Force recalculation via Excel
excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = False
excel.DisplayAlerts = False
try:
    workbook = excel.Workbooks.Open(filepath)
    workbook.Save()
    workbook.Close()
    print("Recalculation successful!")
except Exception as e:
    print(f"Error: {e}")
finally:
    excel.Quit()

# Now read the values
wb = openpyxl.load_workbook(filepath, data_only=True)
sheet = wb['NTB – Input']
cols = ['D', 'E', 'F', 'G', 'H', 'I']
print("\n=== Values after recalculation in NTB_Input_Da_Dien_FLM_CRC ===")
for r in [41, 42, 43, 44, 45]:
    label = sheet.cell(r, 2).value or sheet.cell(r, 1).value
    vals = [sheet.cell(r, openpyxl.utils.column_index_from_string(c)).value for c in cols]
    print(f"Row {r:2d} | {str(label)[:35]:<35} | {vals}")
wb.close()
