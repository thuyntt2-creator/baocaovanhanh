import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

for name in ['Định biên & Sản lượng', 'Chi phí FLM']:
    sheet = wb[name]
    print(f"\n=== Formulas in {name} ===")
    count = 0
    for r in range(1, 100):
        for c in range(1, 15):
            val = sheet.cell(r, c).value
            if isinstance(val, str) and val.startswith('='):
                col_letter = openpyxl.utils.get_column_letter(c)
                print(f"  Cell {col_letter}{r}: {val}")
                count += 1
                if count >= 30:
                    break
        if count >= 30:
            break

wb.close()
