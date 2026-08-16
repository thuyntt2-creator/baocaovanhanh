import openpyxl, sys, shutil
from openpyxl.styles import Font, PatternFill, Alignment
sys.stdout.reconfigure(encoding='utf-8')

in_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v6.xlsx'
out_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v20.xlsx'

print("Đang tạo Báo cáo V20 (Giữ nguyên form gốc, thay công thức dòng 17)...")
shutil.copy(in_path, out_path)
wb = openpyxl.load_workbook(out_path)

# Fix mặt bằng
sheet_ts = wb['1. Thông số']
sheet_ts['B22'] = 150000 

sheet = wb['Nguồn lực & chi phí']

months = ['T7', 'T8', 'T9', 'T10', 'T11', 'T12']
cols = ['B', 'C', 'D', 'E', 'F', 'G']

# Dòng 17: Chi phí nhân sự xử lý = Tổng chi phí 4 Hub (Xe + Lương kho)
for c_idx, m in enumerate(months):
    col_letter = cols[c_idx]
    end_col = 'AF' if m in ['T7', 'T8', 'T10', 'T12'] else 'AE'
    
    hub_formulas = []
    for i in range(4):
        fc_row = i + 4
        kh_row = 18 + i*10
        
        # Lương kho của 1 hub
        luong = f"(3 + ROUNDUP(AVERAGE('Forecast {m}'!B{fc_row}:{end_col}{fc_row})/'1. Thông số'!$B$19,0))*'1. Thông số'!$B$24"
        # Xe của 1 hub
        xe = f"ROUND(AVERAGE('Kế hoạch {m}'!B{kh_row}:{end_col}{kh_row}),0)*'1. Thông số'!$B$23*{col_letter}$4"
        
        hub_formulas.append(f"({luong}+{xe})")
    
    final_formula = "=" + "+".join(hub_formulas)
    sheet.cell(17, c_idx+2).value = final_formula

# Dòng 18 (Tổng chi phí) và 19 (Chi phí/đơn) giữ nguyên công thức gốc của V6

wb.save(out_path)
print(f"Xong! Đã lưu: {out_path}")
