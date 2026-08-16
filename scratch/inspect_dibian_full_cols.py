import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['Định biên & Sản lượng']

print("=== Row 17 ===")
for col in range(1, sheet.max_column + 1):
    val = sheet.cell(17, col).value
    if val is not None:
        print(f"Col {col} ({openpyxl.utils.get_column_letter(col)}): {val}")

print("=== Row 31 ===")
for col in range(1, sheet.max_column + 1):
    val = sheet.cell(31, col).value
    if val is not None:
        print(f"Col {col} ({openpyxl.utils.get_column_letter(col)}): {val}")

print("=== Row 33 ===")
for col in range(1, sheet.max_column + 1):
    val = sheet.cell(33, col).value
    if val is not None:
        print(f"Col {col} ({openpyxl.utils.get_column_letter(col)}): {val}")

wb.close()
