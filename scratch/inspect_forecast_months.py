import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v18.xlsx"

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    sys.exit(1)

wb_formula = openpyxl.load_workbook(file_path, data_only=False)
wb_value = openpyxl.load_workbook(file_path, data_only=True)

for m in [7, 8, 9, 10, 11, 12]:
    sheet_name = f"Forecast T{m}"
    if sheet_name in wb_formula.sheetnames:
        sheet_f = wb_formula[sheet_name]
        sheet_v = wb_value[sheet_name]
        print(f"\n==========================================")
        print(f"SHEET: {sheet_name} (Rows: {sheet_f.max_row}, Cols: {sheet_f.max_column})")
        print(f"==========================================")
        
        # Print row 1-7, columns A-L
        for r in range(1, min(sheet_f.max_row + 1, 9)):
            row_vals_f = [sheet_f.cell(r, c).value for c in range(1, min(sheet_f.max_column + 1, 15))]
            row_vals_v = [sheet_v.cell(r, c).value for c in range(1, min(sheet_v.max_column + 1, 15))]
            if any(v is not None for v in row_vals_f):
                print(f"Row {r}:")
                for c_idx, (vf, vv) in enumerate(zip(row_vals_f, row_vals_v)):
                    if vf is not None:
                        col_letter = openpyxl.utils.get_column_letter(c_idx + 1)
                        print(f"  {col_letter}: Formula = {repr(vf)} | Value = {repr(vv)}")
    else:
        print(f"\nSheet {sheet_name} not found!")
