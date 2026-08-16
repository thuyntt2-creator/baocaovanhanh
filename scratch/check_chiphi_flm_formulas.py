import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['Chi phí FLM']

print("=== Formulas in 'Chi phí FLM' sheet ===")
for r in range(30, 48):
    label = sheet.cell(r, 2).value or sheet.cell(r, 1).value or f"Row {r}"
    c_val = sheet[f"D{r}"].value  # Let's print the formula for col D (July)
    print(f"Row {r:2d} | {str(label)[:35]:<35} | D: {c_val}")
wb.close()
