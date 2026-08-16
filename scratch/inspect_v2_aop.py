import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

xlsx_path = r"C:\Users\lap4all\Downloads\[V2] AOP_Hang_NTB_T7-T12_2026.xlsx"
if not os.path.exists(xlsx_path):
    print(f"File not found: {xlsx_path}")
    sys.exit(1)

wb = openpyxl.load_workbook(xlsx_path, data_only=True)
print("Sheets in [V2] AOP_Hang_NTB_T7-T12_2026.xlsx:", wb.sheetnames)

for sname in wb.sheetnames:
    sheet = wb[sname]
    print(f"\n--- Sheet {sname} (Rows: {sheet.max_row}, Cols: {sheet.max_column}) ---")
    for r in range(1, min(sheet.max_row + 1, 20)):
        row_vals = [sheet.cell(r, c).value for c in range(1, min(sheet.max_column + 1, 12))]
        if any(v is not None for v in row_vals):
            print(f"  Row {r:2d}: {row_vals}")

