import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v18.xlsx"

wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Kế hoạch T7']

print("=== QUÉT TÌM CHỮ VIẾT TẮT NT, DL, DD, ĐL TRONG KẾ HOẠCH T7 ===")
abbrevs = ['NT', 'DL', 'DD', 'ĐL']

for r in range(1, 150):
    for c in range(1, 100):
        val = sheet.cell(row=r, column=c).value
        if val is not None:
            val_str = str(val).strip()
            if val_str in abbrevs or any(f" {a} " in f" {val_str} " for a in abbrevs):
                print(f"Khớp tại Row {r}, Col {c} ({sheet.cell(row=r, column=c).coordinate}): {val}")
wb.close()
