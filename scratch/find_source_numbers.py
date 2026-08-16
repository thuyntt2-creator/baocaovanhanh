import openpyxl
import os
import glob
import sys

sys.stdout.reconfigure(encoding='utf-8')

downloads_dir = r"C:\Users\lap4all\Downloads"
search_pattern = os.path.join(downloads_dir, "*.xlsx")

# Target sequence of values for BCCK Nha Trang (first few days of T7)
# 385, 385, 275, 220, 275, 220, 770, 880, 385, 440, 495, 220
target_seq = [385, 385, 275, 220, 275, 220, 770, 880]

print("Searching for target sequence in all xlsx files in Downloads...")

for fpath in glob.glob(search_pattern):
    # Skip temporary files
    if os.path.basename(fpath).startswith("~$"):
        continue
    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # Search rows
            for r_idx in range(1, min(sheet.max_row + 1, 100)):
                # Read row values
                row_vals = []
                for c_idx in range(1, min(sheet.max_column + 1, 40)):
                    row_vals.append(sheet.cell(r_idx, c_idx).value)
                
                # Check if target_seq is a subsequence of row_vals
                for i in range(len(row_vals) - len(target_seq) + 1):
                    sub = row_vals[i:i+len(target_seq)]
                    if sub == target_seq:
                        print(f"MATCH FOUND in file: {os.path.basename(fpath)} | Sheet: {sheet_name} | Row {r_idx} | Col index {i+1}")
                        # Print some labels in that row or nearby rows
                        print(f"  Row label (Col A/first non-empty): {row_vals[0]}")
    except Exception as e:
        # Ignore errors for password protected or corrupted files
        pass

print("\nDone searching.")
