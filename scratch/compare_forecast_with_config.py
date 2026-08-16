import openpyxl
import os
import sys
from datetime import datetime, date

sys.stdout.reconfigure(encoding='utf-8')

config_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB (1).xlsx"
if not os.path.exists(config_path):
    config_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB.xlsx"

v18_path = r"C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v18.xlsx"

wb_cfg = openpyxl.load_workbook(config_path, data_only=True)
wb_v18 = openpyxl.load_workbook(v18_path, data_only=True)

# Let's inspect '7_FC_Giao_Daily' in config
sheet_giao = wb_cfg['7_FC_Giao_Daily']
giao_headers = [cell.value for cell in sheet_giao[1]]

def parse_header_date(col_name):
    if not isinstance(col_name, str):
        return None
    parts = col_name.strip().split()
    if len(parts) == 2:
        try:
            day_str, month_str = parts[1].split('/')
            return date(2026, int(month_str), int(day_str))
        except:
            return None
    return None

date_cols = {}
for idx, col in enumerate(giao_headers):
    d = parse_header_date(col)
    if d:
        date_cols[d] = idx + 1

# Sum all Bulky rows for Nha Trang in config
# Nha Trang post office IDs:
nt_bc_ids = {20320000, 22704000, 2502, 22746000, 22774000, 22363000, 20495000, 2399}

print("Comparing daily values for July:")
# For each day in July 2026, sum from config and compare with v18 Forecast T7
for day in range(1, 32):
    d = date(2026, 7, day)
    col_idx = date_cols.get(d)
    
    config_sum = 0.0
    if col_idx:
        for r in range(3, sheet_giao.max_row + 1):
            bc_id = sheet_giao.cell(r, 3).value
            san = sheet_giao.cell(r, 5).value
            if bc_id in nt_bc_ids and san and 'Bulky' in str(san):
                val = sheet_giao.cell(r, col_idx).value
                config_sum += float(val) if val is not None else 0.0
                
    # Get from v18 Forecast T7
    sheet_v18 = wb_v18['Forecast T7']
    # Column for day is col (day + 1)
    v18_val = sheet_v18.cell(4, day + 1).value # Row 4 is BCCK Nha Trang
    
    print(f"Day {day:02d}/07/2026: Config Sum = {config_sum:.2f} | v18 = {v18_val}")
