import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\V3_AOP_Hang_NTB_T7-T12_2026 mới.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

print("=== Summing Forecast sheets ===")
for m in range(7, 13):
    sheet_name = f"Forecast T{m}"
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        total = 0.0
        # Column A is BC name. Row 4 is first BC, let's sum everything in columns B onwards
        for r in range(4, sheet.max_row + 1):
            bc_name = sheet.cell(r, 1).value
            if bc_name is None:
                continue
            # columns B to AE/AF/AG
            for c in range(2, sheet.max_column + 1):
                val = sheet.cell(r, c).value or 0.0
                total += float(val)
        print(f"  {sheet_name}: {total:.1f}")

wb.close()
