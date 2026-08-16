import openpyxl, sys, shutil
sys.stdout.reconfigure(encoding='utf-8')

# Paths
monthly_path = r'C:\Users\lap4all\Downloads\V2 AOP_Hang_NTB_T7-T12_2026.xlsx'
config_path = r'C:\Users\lap4all\Downloads\config_psbba_NTB.xlsx'
mau_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026.xlsx'
out_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final.xlsx'

print("1. Trích xuất danh sách BCCK...")
wb_mon = openpyxl.load_workbook(monthly_path, data_only=True)
sheet_bc = wb_mon['0.3 Bưu cục Detail']
bcck_list = []
for r in range(3, 30):
    bc_name = sheet_bc.cell(r, 2).value
    cls = sheet_bc.cell(r, 11).value
    if bc_name and cls and 'mở bcck' in str(cls).lower():
        bcck_list.append(str(bc_name).strip())

print(f"-> Tìm thấy {len(bcck_list)} BCCK: {bcck_list}")

print("2. Đọc data từ 7_FC_Giao_Daily...")
wb_cfg = openpyxl.load_workbook(config_path, data_only=True)
sheet_fc = wb_cfg['7_FC_Giao_Daily']
mc = sheet_fc.max_column

col_map = {}
for c in range(6, mc):
    header = sheet_fc.cell(1, c).value
    if header:
        parts = str(header).split()
        if len(parts) >= 2:
            col_map[parts[1]] = c

daily_vol = {bc: {} for bc in bcck_list}
for r in range(2, sheet_fc.max_row+1):
    bc = sheet_fc.cell(r, 4).value
    band = sheet_fc.cell(r, 5).value
    if bc and band:
        bc_str = str(bc).strip()
        band_str = str(band).strip()
        matched_bc = None
        for b in bcck_list:
            if b in bc_str or bc_str in b:
                matched_bc = b
                break
        
        if matched_bc and 'Bulky' in band_str:
            for date_str, c in col_map.items():
                vol = sheet_fc.cell(r, c).value or 0
                daily_vol[matched_bc][date_str] = daily_vol[matched_bc].get(date_str, 0) + vol

print("3. Đắp data vào file MAU...")
shutil.copy(mau_path, out_path)
wb_mau = openpyxl.load_workbook(out_path)

sheet_param = wb_mau['1. Thông số']
sheet_param['B2'] = 1.4
sheet_param['B5'] = 47
sheet_param['B6'] = 64
sheet_param['B20'] = 2
sheet_param['B21'] = 3.3
sheet_param['B22'] = 120000
sheet_param['B23'] = 1200000
sheet_param['B24'] = 15000000

months = ['T7', 'T8', 'T9', 'T10', 'T11', 'T12']
for m_idx, m_name in enumerate(months):
    sheet_f = wb_mau[f'Forecast {m_name}']
    
    bc_rows = []
    for r in range(4, 300):
        val = sheet_f.cell(r, 1).value
        if val and str(val).startswith('Kho/BC'):
            bc_rows.append(r)
    
    for i, bc in enumerate(bcck_list):
        if i >= 8: break
        if i < len(bc_rows):
            r_bc = bc_rows[i]
            sheet_f.cell(r_bc, 1).value = bc
            
            for day in range(1, 32):
                date_col = day + 1
                month_num = 7 + m_idx
                date_str = f"{day:02d}/{month_num:02d}"
                
                if m_name in ['T7', 'T8'] and date_str in daily_vol[bc]:
                    sheet_f.cell(r_bc + 1, date_col).value = round(daily_vol[bc][date_str])
                else:
                    if len(daily_vol[bc]) > 0:
                        avg = sum(daily_vol[bc].values()) / len(daily_vol[bc])
                        sheet_f.cell(r_bc + 1, date_col).value = round(avg)

wb_mau.save(out_path)
print("Hoàn thành!")
