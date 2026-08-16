import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

calculated_path = r"C:\Users\lap4all\Downloads\AOP_Hang_NTB_T7-T12_2026_calculated.xlsx"
if not os.path.exists(calculated_path):
    print(f"❌ Calculated file not found: {calculated_path}")
    sys.exit(1)

wb = openpyxl.load_workbook(calculated_path, data_only=False)

print("=== VERIFYING CALCULATED WORKBOOK ===")

# 1. Check 1. Thông số
print("\n1. Sheet: 1. Thông số parameters:")
sheet_ts = wb['1. Thông số']
params = {
    'Hệ số ngày cao điểm (X)': sheet_ts.cell(2, 2).value,
    'Năng suất GTC/chuyến (ngày thường)': sheet_ts.cell(4, 2).value,
    'Năng suất GTC/chuyến (cao điểm)': sheet_ts.cell(5, 2).value,
    'Chi phí xe (đ/xe/ngày)': sheet_ts.cell(8, 2).value,
    'Lương BQ nhân sự xử lý': sheet_ts.cell(15, 2).value,
    'Đơn giá thuê (đ/m²/tháng)': sheet_ts.cell(14, 2).value
}
for name, val in params.items():
    print(f"  - {name}: {val}")

# 2. Check 0.3 Bưu cục Detail row count and formulas
print("\n2. Sheet: 0.3 Bưu cục Detail structure:")
sheet_det = wb['0.3 Bưu cục Detail']
print(f"  - max_row = {sheet_det.max_row}")
# row 3 sample formulas
print(f"  - Row 3 name: {sheet_det.cell(3, 2).value}")
print(f"  - Row 3 formulas:")
print(f"    Col J (SL/ngày): {sheet_det.cell(3, 10).value}")
print(f"    Col K (Mật độ): {sheet_det.cell(3, 11).value}")
print(f"    Col L (m2 gợi ý): {sheet_det.cell(3, 12).value}")
print(f"    Col M (Phương án): {sheet_det.cell(3, 13).value}")
# row 85 sample note
print(f"  - Row 85 value: {sheet_det.cell(85, 1).value}")

# 3. Check Mật độ & phương án summary formulas
print("\n3. Sheet: Mật độ & phương án formulas:")
sheet_md = wb['Mật độ & phương án']
print(f"  - B4 (Số BC đủ): {sheet_md['B4'].value}")
print(f"  - B5 (Số BC không đủ): {sheet_md['B5'].value}")
print(f"  - B6 (Vol đủ): {sheet_md['B6'].value}")
print(f"  - B7 (Vol không đủ): {sheet_md['B7'].value}")

# 4. Check Mặt bằng bưu cục sums and formulas
print("\n4. Sheet: Mặt bằng details:")
sheet_mb = wb['Mặt bằng']
print(f"  - max_row = {sheet_mb.max_row}")
for r in range(3, 8):
    name = sheet_mb.cell(r, 2).value
    vol_formula = sheet_mb.cell(r, 3).value
    m2_formula = sheet_mb.cell(r, 4).value
    rent = sheet_mb.cell(r, 5).value
    shippers = sheet_mb.cell(r, 10).value
    boxep = sheet_mb.cell(r, 11).value
    quanly = sheet_mb.cell(r, 12).value
    total_staff = sheet_mb.cell(r, 13).value
    print(f"  - Row {r}: Name='{name}'")
    print(f"    Vol Formula: {vol_formula}")
    print(f"    m² Formula: {m2_formula}")
    print(f"    Rent: {rent}")
    print(f"    Shippers: {shippers} | Bốc xếp: {boxep} | Quản lý: {quanly} | Total staff formula: {total_staff}")

print(f"  - Row 18 (TỔNG CỘNG):")
print(f"    Shippers sum: {sheet_mb.cell(18, 10).value}")
print(f"    Bốc xếp sum: {sheet_mb.cell(18, 11).value}")
print(f"    Quản lý sum: {sheet_mb.cell(18, 12).value}")
print(f"    Total staff sum: {sheet_mb.cell(18, 13).value}")

# 5. Check Nguồn lực & chi phí formulas
print("\n5. Sheet: Nguồn lực & chi phí structure:")
sheet_nlcp = wb['Nguồn lực & chi phí']
for r in [4, 5, 6, 8, 9, 11, 12, 13, 14, 15, 16, 17, 18]:
    print(f"  - Row {r:2d} ({sheet_nlcp.cell(r, 1).value}): Col B (T7) formula = {sheet_nlcp.cell(r, 2).value}")

print("\n🎉 Verification script completed successfully!")

