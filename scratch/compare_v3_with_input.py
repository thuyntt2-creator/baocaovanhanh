import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

v3_path = r"C:\Users\lap4all\Downloads\V3_AOP_Hang_NTB_T7-T12_2026 mới.xlsx"
input_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"

wb_v3 = openpyxl.load_workbook(v3_path, data_only=True)
wb_in = openpyxl.load_workbook(input_path, data_only=True)

print("=== V3 Nguồn lực & chi phí ===")
sheet_v3 = wb_v3['Nguồn lực & chi phí']
for r in range(1, 23):
    row_vals = [sheet_v3.cell(r, c).value for c in range(1, 8)]
    print(f"Row {r:2d}: {row_vals}")

print("\n=== NTB - Input ===")
sheet_in = wb_in['NTB – Input']
for r in range(1, 48):
    row_vals = [sheet_in.cell(r, c).value for c in range(1, 11)]
    print(f"Row {r:2d}: {row_vals}")

wb_v3.close()
wb_in.close()
