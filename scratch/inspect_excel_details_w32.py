import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r'C:\Users\lap4all\Downloads\BaoCao_Tuan_NTB_W32_2026.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

def print_sheet_summary(sheetname, max_r=40, max_c=12):
    ws = wb[sheetname]
    print(f"\n==========================================")
    print(f"=== SHEET: {sheetname} (max_row={ws.max_row}, max_col={ws.max_column}) ===")
    print(f"==========================================")
    for r in range(1, min(max_r, ws.max_row+1)):
        row_vals = [ws.cell(r, c).value for c in range(1, min(max_c, ws.max_column+1))]
        if any(v is not None for v in row_vals):
            formatted_vals = []
            for v in row_vals:
                if v is None:
                    formatted_vals.append("")
                elif isinstance(v, float):
                    if 0 < abs(v) < 1:
                        formatted_vals.append(f"{v*100:.1f}%")
                    else:
                        formatted_vals.append(f"{v:,.1f}")
                else:
                    formatted_vals.append(str(v)[:25])
            print(f"R{r:2d}:", formatted_vals)

for s in wb.sheetnames:
    print_sheet_summary(s, max_r=35, max_c=10)
