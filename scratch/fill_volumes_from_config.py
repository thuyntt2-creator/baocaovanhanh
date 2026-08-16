import openpyxl
import win32com.client
import sys
import math

sys.stdout.reconfigure(encoding='utf-8')

config_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB new.xlsx"
input_path  = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"

# ===== ĐỌC DATA TỪ CONFIG =====
from collections import defaultdict

wb_cfg = openpyxl.load_workbook(config_path, data_only=True)

def sum_topline(sheet, months=[7,8,9,10,11,12]):
    """Sum volume by rank_weight and month"""
    result = defaultdict(lambda: defaultdict(float))
    for r in range(2, sheet.max_row + 1):
        rw  = sheet.cell(r, 4).value
        vol = sheet.cell(r, 6).value or 0
        m   = sheet.cell(r, 9).value
        if rw and m:
            result[str(rw)][int(m)] += vol
    return result

giao = sum_topline(wb_cfg['Topline_Giao_NTB'])
lay  = sum_topline(wb_cfg['Topline_Lay_NTB'])
wb_cfg.close()

months = [7, 8, 9, 10, 11, 12]

# Band mapping
B3 = '03.10-15kg'
B4 = '04.15-20kg'
B5 = '05.>=20kg'

print("=== DATA TỪ CONFIG ===")
print("GIAO Band3:", [round(giao[B3][m]) for m in months])
print("GIAO Band4:", [round(giao[B4][m]) for m in months])
print("GIAO Band5:", [round(giao[B5][m]) for m in months])
print("LAY  Band3:", [round(lay[B3][m]) for m in months])
print("LAY  Band4:", [round(lay[B4][m]) for m in months])
print("LAY  Band5:", [round(lay[B5][m]) for m in months])

# ===== ĐIỀN VÀO FILE =====
wb_in = openpyxl.load_workbook(input_path, data_only=False)
sheet = wb_in['NTB – Input']
cols  = ['D', 'E', 'F', 'G', 'H', 'I']  # T7 → T12

row_map = {
    8:  [round(giao[B3][m]) for m in months],  # GIAO Band 3
    9:  [round(giao[B4][m]) for m in months],  # GIAO Band 4
    # Row 10: GIAO Band 5 (≥20kg) - cũng cập nhật từ config
    10: [round(giao[B5][m]) for m in months],  # GIAO Band 5
    11: [round(lay[B3][m])  for m in months],  # LẤY Band 3
    12: [round(lay[B4][m])  for m in months],  # LẤY Band 4
    13: [round(lay[B5][m])  for m in months],  # LẤY Band 5
}

for row, values in row_map.items():
    for idx, col in enumerate(cols):
        sheet[f"{col}{row}"] = values[idx]

wb_in.save(input_path)
wb_in.close()
print("\nĐã lưu. Đang recalculate qua Excel COM...")

excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = False
excel.DisplayAlerts = False
try:
    workbook = excel.Workbooks.Open(input_path)
    workbook.Save()
    workbook.Close()
    print("Hoàn tất!")
except Exception as e:
    print(f"Lỗi COM: {e}")
finally:
    excel.Quit()

# Xác nhận lại
wb_verify = openpyxl.load_workbook(input_path, data_only=True)
s = wb_verify['NTB – Input']
print("\n=== Giá trị đã điền ===")
row_labels = {8:"GIAO Band3", 9:"GIAO Band4", 10:"GIAO Band5",
              11:"LAY Band3", 12:"LAY Band4", 13:"LAY Band5"}
for r, lbl in row_labels.items():
    vals = [s[f"{c}{r}"].value for c in cols]
    print(f"  Row {r:2d} ({lbl}): {vals}")
wb_verify.close()
