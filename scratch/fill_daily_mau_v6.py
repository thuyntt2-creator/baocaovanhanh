import openpyxl, sys, shutil
sys.stdout.reconfigure(encoding='utf-8')

mau_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v5.xlsx'
out_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v6.xlsx'

print("Đang áp dụng thuật toán Volume Smoothing để tạo V6...")
shutil.copy(mau_path, out_path)
wb = openpyxl.load_workbook(out_path)

months = ['T7', 'T8', 'T9', 'T10', 'T11', 'T12']
days_in_month = {'T7': 31, 'T8': 31, 'T9': 30, 'T10': 31, 'T11': 30, 'T12': 31}

# Năng suất lý tưởng cho 1 xe (47 đơn / Tỷ lệ GTC 85%)
OPTIMAL_TRUCK_CAPACITY = 55 

for m in months:
    sheet = wb[f'Forecast {m}']
    num_days = days_in_month[m]
    
    for i in range(4): # 4 Hubs in rows 4, 5, 6, 7
        r_bc = 4 + i
        hub_name = sheet.cell(r_bc, 1).value
        if not hub_name: continue
        
        excess = 0
        for day in range(1, num_days + 1):
            col = day + 1
            orig_vol = sheet.cell(r_bc, col).value
            if orig_vol is None: orig_vol = 0
            
            current_vol = orig_vol + excess
            
            # Tính số xe tối ưu
            trucks = round(current_vol / OPTIMAL_TRUCK_CAPACITY)
            
            # Đảm bảo mỗi ngày tối thiểu 1 xe nếu có phát sinh hàng, trừ phi hàng quá ít dồn qua ngày sau
            if trucks == 0 and current_vol > 0:
                if current_vol > (OPTIMAL_TRUCK_CAPACITY * 0.4): # Lớn hơn 40% xe thì cho chạy luôn 1 xe
                    trucks = 1
                else:
                    trucks = 0 # Ép dồn sang ngày sau
                    
            quantized_vol = trucks * OPTIMAL_TRUCK_CAPACITY
            
            # Trong ngày cuối cùng của tháng, ta ép quantized_vol = current_vol 
            # để đảm bảo KHÔNG BỊ RỚT HÀNG sang tháng sau, giữ nguyên tổng tháng.
            if day == num_days:
                quantized_vol = current_vol
                
            excess = current_vol - quantized_vol
            sheet.cell(r_bc, col).value = round(quantized_vol)

wb.save(out_path)
print(f"Hoàn thành tối ưu GAP! Đã lưu: {out_path}")
