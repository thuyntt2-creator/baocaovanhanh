import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\V3_AOP_Hang_NTB_T7-T12_2026 mới.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

if 'Nguồn lực & chi phí' in wb.sheetnames:
    sheet = wb['Nguồn lực & chi phí']
    print(f"=== Sheet: Nguồn lực & chi phí (max_row: {sheet.max_row}, max_col: {sheet.max_column}) ===")
    for r in range(1, 55):
        row_vals = []
        for c in range(1, 15):
            row_vals.append(sheet.cell(r, c).value)
        if any(v is not None for v in row_vals):
            print(f"Row {r:2d}: {row_vals}")
else:
    print("Sheet 'Nguồn lực & chi phí' not found!")

wb.close()
