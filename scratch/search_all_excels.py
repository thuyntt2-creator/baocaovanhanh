import openpyxl
import os
import sys
import glob

sys.stdout.reconfigure(encoding='utf-8')

downloads_dir = r"C:\Users\lap4all\Downloads"
excel_files = glob.glob(os.path.join(downloads_dir, "*.xlsx"))

targets = [715, 690, 1746, 126, 76]

print(f"=== ĐANG QUÉT TẤT CẢ FILE EXCEL TRONG DOWNLOADS ({len(excel_files)} file) ===")

for path in excel_files:
    fname = os.path.basename(path)
    if fname.startswith("~$"):
        continue  # Bỏ qua file tạm
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        # print(f"Không thể đọc {fname}: {e}")
        continue
        
    # print(f"Quét {fname}...")
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # Giới hạn quét để tránh quá chậm trên file lớn
        max_r = min(sheet.max_row, 300)
        max_c = min(sheet.max_column, 50)
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                val = sheet.cell(r, c).value
                if val is not None:
                    # Kiểm tra khớp target
                    for target in targets:
                        # khớp số nguyên chính xác
                        if (isinstance(val, (int, float)) and abs(val - target) < 0.1) or \
                           (isinstance(val, (int, float)) and abs(val/1e6 - target) < 0.1) or \
                           (isinstance(val, (int, float)) and abs(val/1e3 - target) < 0.1):
                            print(f"[KHỚP] File: {fname} | Sheet: {sheet_name} | Ô {openpyxl.utils.get_column_letter(c)}{r}: Giá trị = {val} (Khớp với target {target})")
                        elif str(target) in str(val):
                            # kiểm tra chuỗi
                            if len(str(val)) < 15: # tránh khớp quá nhiều chuỗi rác dài
                                print(f"[KHỚP CHUỖI] File: {fname} | Sheet: {sheet_name} | Ô {openpyxl.utils.get_column_letter(c)}{r}: Giá trị = {val} (Chứa target {target})")
