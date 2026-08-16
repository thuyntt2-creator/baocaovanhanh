import openpyxl
import os
import sys
import math

sys.stdout.reconfigure(encoding='utf-8')

v2_path = r"C:\Users\lap4all\Downloads\[V2] AOP_Hang_NTB_T7-T12_2026.xlsx"
wb = openpyxl.load_workbook(v2_path, data_only=True)
sheet = wb['0.3 Bưu cục Detail']

print("=== Summing V2 Bưu Cục volumes for BCCK with New Grouping ===")

# List of all bưu cục rows
bcs = {}
for r in range(3, sheet.max_row + 1):
    bc_name = sheet.cell(r, 2).value
    if bc_name is not None:
        bcs[bc_name.strip()] = {
            'row': r,
            't7': sheet.cell(r, 4).value or 0.0,
            't10': sheet.cell(r, 7).value or 0.0,
            't12': sheet.cell(r, 9).value or 0.0
        }

groups = {
    'BCCK Nha Trang': [
        "Bưu Cục 06 Lê Hồng Phong-TP.Nha Trang-Khánh Hòa",
        "Bưu Cục 195 Đường 2/4-Nha Trang-Khánh Hòa",
        "Bưu Cục 229 Phước Long-Nam Nha Trang-Khánh Hòa",
        "Bưu Cục 40A Yết Kiêu-Nha Trang-Khánh Hòa",
        "Bưu Cục 466 Đường 23/10-Nha Trang-Khánh Hòa",
        "Bưu Cục Đường 35 Hà Quang 1-Xã Nam Nha Trang-Khánh Hòa",
        "Bưu Cục Phước Đồng-Nha Trang-Khánh Hoà"
    ],
    'BCCK Di Linh': [
        "Bưu Cục 1322 Hùng Vương-Di Linh-Lâm Đồng",
        "(LDO) Đinh Văn Lâm Hà",
        "(LDO) Tân Hà Lâm Hà",
        "Bưu Cục 231 Thôn 1-Xã Hòa Ninh-Lâm Đồng",
        "(LDO) Nam Ban Lâm Hà"
    ],
    'BCCK Đơn Dương': [
        "(LDO) Đơn Dương",
        "(LDO) Hiệp Thạnh"
    ],
    'BCKK Đức Linh-Bình Thuận': [
        "(BTH) Đức Linh"
    ]
}

months_cols = {'T7': 't7', 'T10': 't10', 'T12': 't12'}

for gname, items in groups.items():
    print(f"\n{gname}:")
    for m, key in months_cols.items():
        monthly_vol = 0.0
        for name in items:
            if name in bcs:
                monthly_vol += bcs[name][key]
            else:
                print(f"  ⚠️ Warning: {name} not found in sheet!")
        daily_vol = round(monthly_vol / 30.0)
        chuyen = daily_vol * 0.8 / 47.0
        chuyen_rounded = math.ceil(chuyen)
        xe = math.ceil(chuyen_rounded / 2.0)
        print(f"  Month {m}: Monthly Vol = {monthly_vol:.1f} | Daily = {daily_vol} | Chuyến = {chuyen_rounded} (raw: {chuyen:.2f}) | Xe = {xe}")
