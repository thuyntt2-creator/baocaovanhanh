import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

xlsx_aop_path = r"C:\Users\lap4all\Downloads\AOP_Hang_NTB_T7-T12_2026.xlsx"
wb = openpyxl.load_workbook(xlsx_aop_path, data_only=True)

for sname in ['Volume Giao', 'Volume Lấy']:
    sheet = wb[sname]
    print(f"\n=== Summing Raw Data in {sname} ===")
    
    # Let's sum the raw data starting from row 22
    sums = {}
    for r_idx in range(22, sheet.max_row + 1):
        prov = sheet.cell(r_idx, 1).value
        band = sheet.cell(r_idx, 3).value
        if prov is None or band is None:
            continue
        # Clean band names to match the top ones
        # Top bands: '03.10-15kg  (Hàng vừa)', '04.15-20kg  (Hàng vừa)', '05.>=20kg  (Hàng nặng)'
        # Raw bands: '03.10-15kg', '04.15-20kg', '05.>=20kg' (or similar)
        clean_band = band.strip()
        key = (prov.strip(), clean_band)
        sums.setdefault(key, [0.0]*6)
        for m_idx in range(6):
            val = sheet.cell(r_idx, 4 + m_idx).value
            if val is not None:
                sums[key][m_idx] += float(val)
                
    # Compare with top table
    # Top values
    top_vals = {}
    for r_idx in range(3, 18):
        prov = sheet.cell(r_idx, 1).value
        band_full = sheet.cell(r_idx, 2).value
        # clean band
        clean_b = band_full.split('(')[0].strip()
        key = (prov.strip(), clean_b)
        vals = [sheet.cell(r_idx, 3 + m_idx).value for m_idx in range(6)]
        top_vals[key] = vals
        
    for key in sorted(top_vals.keys()):
        raw_sum = sums.get(key, [0.0]*6)
        top = top_vals[key]
        # round for display
        raw_sum_disp = [round(x, 1) for x in raw_sum]
        print(f"Key: {key}\n  Top: {top}\n  Raw sum: {raw_sum_disp}")

