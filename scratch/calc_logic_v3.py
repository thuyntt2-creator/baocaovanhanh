import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')
print("Loading config...")
wb_cfg = openpyxl.load_workbook(r'C:\Users\lap4all\Downloads\config_psbba_NTB.xlsx', data_only=True)

print("1. Parsing Config...")
sheet_cfg = wb_cfg['1_Config_Chuan']
ward_data = {}
bc_ward_count = {}
for r in range(2, sheet_cfg.max_row+1):
    ward = sheet_cfg.cell(r, 7).value
    bc = sheet_cfg.cell(r, 11).value
    grp = sheet_cfg.cell(r, 8).value
    if ward and bc:
        ward = str(ward).strip()
        bc = str(bc).strip()
        ward_data[ward] = {'bc': bc, 'group': grp}
        bc_ward_count[bc] = bc_ward_count.get(bc, 0) + 1

print("2. Parsing 3_KQ_BC_Detail...")
sheet_3 = wb_cfg['3_KQ_BC_Detail']
# Find T07/2026 column index
t7_col_3 = 0
for c in range(1, 50):
    val = sheet_3.cell(1, c).value
    if val and 'T07/2026' in str(val):
        t7_col_3 = c
        break

bc_reg_vol = {}
for r in range(2, sheet_3.max_row+1):
    bc = sheet_3.cell(r, 4).value
    weight = str(sheet_3.cell(r, 5).value or '')
    if bc and weight in ['10-15kg', '15-30kg', '≥30kg']:
        vol = sheet_3.cell(r, t7_col_3).value or 0
        bc = str(bc).strip()
        bc_reg_vol[bc] = bc_reg_vol.get(bc, 0) + vol

print("3. Parsing 4_KQ_B2B...")
sheet_4 = wb_cfg['4_KQ_B2B']
t7_col_4 = 0
for c in range(1, 50):
    val = sheet_4.cell(1, c).value
    if val and 'T07/2026' in str(val):
        t7_col_4 = c
        break

ward_b2b_vol = {}
for r in range(2, sheet_4.max_row+1):
    ward = sheet_4.cell(r, 4).value
    if ward:
        vol = sheet_4.cell(r, t7_col_4).value or 0
        ward = str(ward).strip()
        ward_b2b_vol[ward] = ward_b2b_vol.get(ward, 0) + vol

print("\n--- RESULTS FOR DI LINH & LÂM HÀ ---")
wards_to_test = [
    'Thị trấn Di Linh', 'Thị trấn Đinh Văn', 'Xã Tân Hà', 'Thị trấn Nam Ban', 
    'Xã Hòa Ninh', 'Xã Đinh Trang Hòa', 'Xã Tân Văn', 'Xã Liên Đầm', 
    'Xã Gia Lâm', 'Xã Đông Thanh', 'Xã Bình Thạnh', 'Xã Đạ Đờn', 
    'Xã Phi Tô', 'Xã Phú Sơn', 'Xã Mê Linh', 'Xã Nam Hà', 'Xã Hòa Bắc'
]

for w in wards_to_test:
    if w in ward_data:
        bc = ward_data[w]['bc']
        grp = ward_data[w]['group']
        w_count = bc_ward_count.get(bc, 1)
        tot_bc_reg = bc_reg_vol.get(bc, 0)
        avg_reg = tot_bc_reg / w_count
        b2b_vol = ward_b2b_vol.get(w, 0)
        
        if grp in ['A', 'B']:
            final_vol = b2b_vol + avg_reg
        else:
            final_vol = avg_reg
            
        print(f"{w} (Nhóm {grp}): BC={bc} | T.Số Xã={w_count} | Tổng BC GTC={tot_bc_reg:.1f} => GTC TB={avg_reg:.1f} | B2B={b2b_vol:.1f} || Tổng = {final_vol:.1f} (Làm tròn: {round(final_vol)})")
    else:
        print(f"Không tìm thấy {w} trong config!")
