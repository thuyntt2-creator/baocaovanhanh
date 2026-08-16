import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

filepath = r"C:\Users\lap4all\Downloads\NTB_Input_Da_Dien_FLM_CRC_2.xlsx"
wb = openpyxl.load_workbook(filepath, data_only=True)
sheet = wb['NTB – Input']

cols = ['D', 'E', 'F', 'G', 'H', 'I']
print("=== NTB_Input_Da_Dien_FLM_CRC_2.xlsx - Rows 30-45 values ===")
for r in range(30, 46):
    label = sheet.cell(r, 2).value or sheet.cell(r, 1).value
    vals = [sheet.cell(r, openpyxl.utils.column_index_from_string(c)).value for c in cols]
    print(f"Row {r:2d} | {str(label)[:35]:<35} | {vals}")
wb.close()
