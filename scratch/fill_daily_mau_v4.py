import openpyxl, sys, shutil
sys.stdout.reconfigure(encoding='utf-8')

config_path = r'C:\Users\lap4all\Downloads\config_psbba_NTB.xlsx'
mau_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026.xlsx'
out_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v4.xlsx'

print("1. Đọc data từ 7_FC_Giao_Daily...")
wb_cfg = openpyxl.load_workbook(config_path, data_only=True)
sheet_fc = wb_cfg['7_FC_Giao_Daily']
mc = sheet_fc.max_column

col_map = {}
for c in range(6, mc):
    header = sheet_fc.cell(1, c).value
    if header:
        parts = str(header).split()
        if len(parts) >= 2:
            col_map[parts[1]] = c

groups = {
    'BCCK Nha Trang': ['Nha Trang'],
    'BCCK Đơn Dương': ['Đơn Dương', 'Hiệp Thạnh'],
    'BCCK Di Linh': ['Di Linh', 'Hòa Ninh', 'Lâm Hà'],
    'BCCK Đức Linh': ['Đức Linh']
}

daily_vol = {g: {} for g in groups.keys()}

for r in range(2, sheet_fc.max_row+1):
    bc = sheet_fc.cell(r, 4).value
    band = sheet_fc.cell(r, 5).value
    if bc and band:
        bc_str = str(bc).strip()
        band_str = str(band).strip()
        
        if 'Bulky' in band_str:
            if bc_str == 'Kho Giao Hàng Nặng - Nha Trang - Khánh Hoà':
                continue
            
            matched_group = None
            for g_name, keywords in groups.items():
                for kw in keywords:
                    if kw in bc_str:
                        matched_group = g_name
                        break
                if matched_group:
                    break
            
            if matched_group:
                for date_str, c in col_map.items():
                    vol = sheet_fc.cell(r, c).value or 0
                    daily_vol[matched_group][date_str] = daily_vol[matched_group].get(date_str, 0) + vol

print("2. Đắp data vào file MAU V4...")
shutil.copy(mau_path, out_path)
wb_mau = openpyxl.load_workbook(out_path)

sheet_param = wb_mau['1. Thông số']
sheet_param['B2'] = 1.4
sheet_param['B3'] = 20
sheet_param['B4'] = 30
sheet_param['B5'] = 47
sheet_param['B6'] = 64
sheet_param['B10'] = 0.855
sheet_param['B11'] = 0.858
sheet_param['B12'] = 0.859
sheet_param['B13'] = 0.858
sheet_param['B14'] = 0.858
sheet_param['B15'] = 0.846
sheet_param['B16'] = 0.825
sheet_param['B19'] = 55
sheet_param['B20'] = 2.0
sheet_param['B21'] = 3.3
sheet_param['B22'] = 150000
sheet_param['B23'] = 1200000
sheet_param['B24'] = 15000000

months = ['T7', 'T8', 'T9', 'T10', 'T11', 'T12']
group_names = list(groups.keys())

for m_idx, m_name in enumerate(months):
    sheet_f = wb_mau[f'Forecast {m_name}']
    
    # In original template, rows 4 to 11 are Kho/BC 1 to Kho/BC 8
    # We will write the 4 groups to rows 4, 5, 6, 7 and clear 8, 9, 10, 11
    
    for i in range(8):
        r_bc = 4 + i
        if i < len(group_names):
            g_name = group_names[i]
            sheet_f.cell(r_bc, 1).value = g_name
            for day in range(1, 32):
                date_col = day + 1
                month_num = 7 + m_idx
                date_str = f"{day:02d}/{month_num:02d}"
                
                if m_name in ['T7', 'T8'] and date_str in daily_vol[g_name]:
                    sheet_f.cell(r_bc, date_col).value = round(daily_vol[g_name][date_str])
                else:
                    if len(daily_vol[g_name]) > 0:
                        avg = sum(daily_vol[g_name].values()) / len(daily_vol[g_name])
                        sheet_f.cell(r_bc, date_col).value = round(avg)
        else:
            # Clear row
            sheet_f.cell(r_bc, 1).value = None
            for day in range(1, 32):
                sheet_f.cell(r_bc, day + 1).value = None

wb_mau.save(out_path)
print(f"Hoàn thành! Đã lưu: {out_path}")
