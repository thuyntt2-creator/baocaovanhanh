import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\V3 AOP_NTB_T70-T12_2026.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
print("Sheets in new V3 file:", wb.sheetnames)

if 'Mặt bằng' in wb.sheetnames:
    sheet = wb['Mặt bằng']
    print(f"\n=== Sheet: Mặt bằng (max_row={sheet.max_row}, max_col={sheet.max_column}) ===")
    for r in range(1, 40):
        row_vals = [sheet.cell(r, c).value for c in range(1, 15)]
        if any(v is not None for v in row_vals):
            print(f"Row {r:2d}: {row_vals}")
else:
    print("Sheet 'Mặt bằng' not found!")

wb.close()
