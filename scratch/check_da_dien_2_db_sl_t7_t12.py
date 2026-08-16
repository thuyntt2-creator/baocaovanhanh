import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

filepath = r"C:\Users\lap4all\Downloads\NTB_Input_Da_Dien_FLM_CRC_2.xlsx"
wb = openpyxl.load_workbook(filepath, data_only=True)
sheet = wb['Định biên & Sản lượng']

print("=== NTB_Input_Da_Dien_FLM_CRC_2.xlsx - Định biên & Sản lượng rows 72-77 (T7-T12) ===")
cols = ['I', 'J', 'K', 'L', 'M', 'N']  # columns 9 to 14, representing months 7 to 12
for r in range(72, 78):
    label = sheet.cell(r, 1).value or sheet.cell(r, 2).value or f"Row {r}"
    vals = [sheet.cell(r, openpyxl.utils.column_index_from_string(c)).value for c in cols]
    print(f"Row {r:2d} | {str(label)[:35]:<35} | {vals}")
wb.close()
