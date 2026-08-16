import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['Chi phí FLM']

wb_eval = openpyxl.load_workbook(file_path, data_only=True)
sheet_eval = wb_eval['Chi phí FLM']

cols = ['J', 'K', 'L', 'M', 'N', 'O']
print("=== Formulas in 'Chi phí FLM' ===")
for r in [36, 37, 38, 39]:
    label = sheet.cell(r, 2).value or sheet.cell(r, 1).value
    row_formulas = [sheet.cell(r, openpyxl.utils.column_index_from_string(c)).value for c in cols]
    row_vals = [sheet_eval.cell(r, openpyxl.utils.column_index_from_string(c)).value for c in cols]
    print(f"Row {r:2d} | Label: '{label}'")
    print(f"  Formulas: {row_formulas}")
    print(f"  Values  : {row_vals}")

wb_eval.close()
wb.close()
