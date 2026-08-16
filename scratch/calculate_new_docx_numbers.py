import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

calculated_path = r"C:\Users\lap4all\Downloads\AOP_Hang_NTB_T7-T12_2026_calculated.xlsx"
wb = openpyxl.load_workbook(calculated_path, data_only=True)

# 1. Table 2.2 calculation
print("=== Table 2.2 ===")
sheet_det = wb['0.3 Bưu cục Detail']

# Group bưu cục lists
nt_bcs = [
    "Bưu Cục 06 Lê Hồng Phong-TP.Nha Trang-Khánh Hòa",
    "Bưu Cục 195 Đường 2/4-Nha Trang-Khánh Hòa",
    "Bưu Cục 229 Phước Long-Nam Nha Trang-Khánh Hòa",
    "Bưu Cục 40A Yết Kiêu-Nha Trang-Khánh Hòa",
    "Bưu Cục 466 Đường 23/10-Nha Trang-Khánh Hòa",
    "Bưu Cục Đường 35 Hà Quang 1-Xã Nam Nha Trang-Khánh Hòa",
    "Bưu Cục Phước Đồng-Nha Trang-Khánh Hoà"
]
dl_bcs = [
    "Bưu Cục 1322 Hùng Vương-Di Linh-Lâm Đồng",
    "(LDO) Đinh Văn Lâm Hà",
    "(LDO) Tân Hà Lâm Hà",
    "Bưu Cục 231 Thôn 1-Xã Hòa Ninh-Lâm Đồng",
    "(LDO) Nam Ban Lâm Hà"
]
dd_bcs = [
    "(LDO) Đơn Dương",
    "(LDO) Hiệp Thạnh"
]
dk_bcs = [
    "(BTH) Đức Linh"
]

# Find row indexes
bc_rows = {}
for r in range(3, sheet_det.max_row + 1):
    name = sheet_det.cell(r, 2).value
    if name:
        bc_rows[name.strip().lower()] = r

def get_group_vol(bcs_list, col_idx):
    vol = 0
    for name in bcs_list:
        r = bc_rows.get(name.strip().lower())
        if r:
            vol += sheet_det.cell(r, col_idx).value
    return vol

months_cols = {'T7': 4, 'T10': 7, 'T12': 9}
groups = {
    'BCCK Nha Trang': nt_bcs,
    'BCCK Di Linh': dl_bcs,
    'BCCK Đơn Dương': dd_bcs,
    'BCCK Đức Linh': dk_bcs
}

for name, bcs_list in groups.items():
    print(f"\n{name}:")
    for m, col in months_cols.items():
        monthly_vol = get_group_vol(bcs_list, col)
        daily_vol = round(monthly_vol / 30.0)
        chuyen = math_chuyen = daily_vol * 0.8 / 47.0
        import math
        chuyen_rounded = math.ceil(chuyen)
        xe = math.ceil(chuyen_rounded / 2.0)
        print(f"  Month {m}: Monthly Vol = {monthly_vol:.1f} | Daily = {daily_vol} | Chuyến = {chuyen_rounded} (raw: {chuyen:.2f}) | Xe = {xe}")

# 2. Table 5.2 calculation
print("\n=== Table 5.2 ===")
sheet_vg = wb['Volume Giao']
provinces = ['Bình Thuận', 'Khánh Hòa', 'Lâm Đồng', 'Ninh Thuận', 'Đắk Nông']

# Read rows
vg_rows = []
for r in range(3, 18):
    prov = sheet_vg.cell(r, 1).value
    band = sheet_vg.cell(r, 2).value
    t7 = sheet_vg.cell(r, 3).value
    t8 = sheet_vg.cell(r, 4).value
    t9 = sheet_vg.cell(r, 5).value
    t10 = sheet_vg.cell(r, 6).value
    t11 = sheet_vg.cell(r, 7).value
    t12 = sheet_vg.cell(r, 8).value
    if prov and band:
        vg_rows.append({'prov': prov.strip(), 'band': band.strip(), 'v': [t7, t8, t9, t10, t11, t12]})

table_5_2_data = {}
for p in provinces:
    table_5_2_data[p] = {
        'Hàng vừa (10–20kg)': [0]*6,
        'Hàng nặng (≥20kg)': [0]*6
    }
    
for r in vg_rows:
    p = r['prov']
    b = r['band']
    if '03.' in b or '04.' in b:
        key = 'Hàng vừa (10–20kg)'
    else:
        key = 'Hàng nặng (≥20kg)'
    if p in table_5_2_data:
        for i in range(6):
            table_5_2_data[p][key][i] += r['v'][i]

total_giao = [0]*6
for p in provinces:
    print(f"Province: {p}")
    for key in ['Hàng vừa (10–20kg)', 'Hàng nặng (≥20kg)']:
        vals = table_5_2_data[p][key]
        print(f"  {key}: " + " | ".join(f"{v:,}" for v in vals))
        for i in range(6):
            total_giao[i] += vals[i]
print(f"TỔNG GIAO BULKY/THÁNG: " + " | ".join(f"{v:,}" for v in total_giao))

# 3. Table 6.3 calculation
print("\n=== Table 6.3 ===")
sheet_n = wb['Nguồn lực & chi phí']
sheet_k = wb['Kênh & nhu cầu']

col_mapping = {'T7': 2, 'T8': 3, 'T9': 4, 'T10': 5, 'T11': 6, 'T12': 7}
sl_hang_nang_ngay = [round(sheet_k.cell(9, c).value) for c in range(2, 8)]
xe_bq_ngay = [sheet_n.cell(4, c).value for c in range(2, 8)]
# Day cao điểm is Xe BQ * 1.4
xe_peak = [math.ceil(x * 1.4) for x in xe_bq_ngay]
nguoi_peak = [x * 2 for x in xe_peak]
mat_bang = [sheet_n.cell(6, c).value for c in range(2, 8)]

print("SL hàng nặng/ngày (BQ): " + " | ".join(str(x) for x in sl_hang_nang_ngay))
print("Tổng đầu xe 1.9T BQ/ngày (4 BCCK): " + " | ".join(str(x) for x in xe_bq_ngay))
print("Tổng đầu xe 1.9T ngày cao điểm: " + " | ".join(f"~{x}" for x in xe_peak))
print("Số người giao (đỉnh × 2 người/xe): " + " | ".join(f"~{x}" for x in nguoi_peak))
print("Mặt bằng tổng 4 BCCK cần (m²): " + " | ".join(f"~{x}" for x in mat_bang))

# 4. Table 7.2 calculation
print("\n=== Table 7.2 ===")
chi_phi_rows = [
    'Chi phí xe tải 1.9T',
    'Chi phí NV giao hàng (38 shipper)',
    'Chi phí NV kho & quản lý (8 người)',
    'Chi phí thuê mặt bằng',
    'TỔNG CHI PHÍ'
]

for r in [11, 15, 14, 16, 13, 17]:
    name = sheet_n.cell(r, 1).value
    vals = [sheet_n.cell(r, c).value for c in range(2, 8)]
    # Convert to Millions (rounded)
    vals_m = [round(v / 1000000.0) for v in vals]
    print(f"Row {r:2d} ({name}): " + " | ".join(f"{v:,}M" for v in vals_m))

