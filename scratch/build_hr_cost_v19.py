import openpyxl, sys, shutil
from openpyxl.styles import Font, PatternFill, Alignment
sys.stdout.reconfigure(encoding='utf-8')

in_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v6.xlsx'
out_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v19.xlsx'

print("Đang tạo Báo cáo V19 (Xóa bảng phụ, gộp công thức thẳng vào Topline)...")
shutil.copy(in_path, out_path)
wb = openpyxl.load_workbook(out_path)

# Fix mặt bằng
sheet_ts = wb['1. Thông số']
sheet_ts['B22'] = 150000 

sheet = wb['Nguồn lực & chi phí']

# Chèn 2 dòng để xé lẻ nhân sự
sheet.insert_rows(18, 2) 

months = ['T7', 'T8', 'T9', 'T10', 'T11', 'T12']
cols = ['B', 'C', 'D', 'E', 'F', 'G']
money_format = '#,##0'

# Gộp công thức tính quân số 4 hub thẳng vào Dòng 11 (Người xử lý kho)
for c_idx, m in enumerate(months):
    end_col = 'AF' if m in ['T7', 'T8', 'T10', 'T12'] else 'AE'
    # Công thức cộng dồn 4 hub
    f_hub1 = f"(3 + ROUNDUP(AVERAGE('Forecast {m}'!B4:{end_col}4)/'1. Thông số'!$B$19,0))"
    f_hub2 = f"(3 + ROUNDUP(AVERAGE('Forecast {m}'!B14:{end_col}14)/'1. Thông số'!$B$19,0))"
    f_hub3 = f"(3 + ROUNDUP(AVERAGE('Forecast {m}'!B24:{end_col}24)/'1. Thông số'!$B$19,0))"
    f_hub4 = f"(3 + ROUNDUP(AVERAGE('Forecast {m}'!B34:{end_col}34)/'1. Thông số'!$B$19,0))"
    
    formula = f"={f_hub1}+{f_hub2}+{f_hub3}+{f_hub4}"
    sheet.cell(11, c_idx+2).value = formula

# Dòng 17: Xử lý kho
sheet.cell(17, 1).value = "Chi phí NV xử lý kho (4 người / 4 BC)"
for c_idx in range(6):
    sheet.cell(17, c_idx+2).value = f"=4*'1. Thông số'!$B$24"
    sheet.cell(17, c_idx+2).number_format = money_format

# Dòng 18: Giao hàng
for c_idx in range(6):
    sheet.cell(18, 1).value = f'="Chi phí NV giao hàng (" & ({cols[c_idx]}11-8) & " người / 4 BC)"'
    sheet.cell(18, c_idx+2).value = f"=({cols[c_idx]}11-8)*'1. Thông số'!$B$24"
    sheet.cell(18, c_idx+2).number_format = money_format

# Dòng 19: Quản lý
sheet.cell(19, 1).value = "Chi phí NV quản lý & backup (4 người / 4 BC)"
for c_idx in range(6):
    sheet.cell(19, c_idx+2).value = f"=4*'1. Thông số'!$B$24"
    sheet.cell(19, c_idx+2).number_format = money_format

# Dòng 20: Tổng
for c_idx in range(6):
    sheet.cell(20, c_idx+2).value = f"={cols[c_idx]}15+{cols[c_idx]}16+{cols[c_idx]}17+{cols[c_idx]}18+{cols[c_idx]}19"
    sheet.cell(20, c_idx+2).number_format = money_format

wb.save(out_path)
print(f"Xong! Đã lưu: {out_path}")
