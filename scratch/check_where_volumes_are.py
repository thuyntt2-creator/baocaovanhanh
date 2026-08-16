import openpyxl
import os
import glob
import sys

sys.stdout.reconfigure(encoding='utf-8')

downloads_dir = r"C:\Users\lap4all\Downloads"

# Find config files
config_files = glob.glob(os.path.join(downloads_dir, "*config_psbba_NTB*"))
print("Config files:", config_files)

targets = [68578, 74620, 77680, 86427, 89586, 99054, 198219, 196204, 204379, 225298, 234987, 261430]

for cf in config_files:
    print(f"\nSearching in {cf}:")
    wb = openpyxl.load_workbook(cf, data_only=True)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # check dimensions
        for r in range(1, min(100, sheet.max_row + 1)):
            for c in range(1, min(20, sheet.max_column + 1)):
                val = sheet.cell(r, c).value
                if val is not None:
                    try:
                        num = float(val)
                        for t in targets:
                            if abs(num - t) < 5:
                                print(f"  Sheet '{sheet_name}' cell {openpyxl.utils.get_column_letter(c)}{r}: {val} (target: {t})")
                    except (ValueError, TypeError):
                        pass
    wb.close()
