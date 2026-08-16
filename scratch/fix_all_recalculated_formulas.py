import openpyxl
import win32com.client
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

downloads_dir = r"C:\Users\lap4all\Downloads"
files_to_fix = [
    os.path.join(downloads_dir, "NTB_Input_Da_Dien_FLM_CRC_2.xlsx"),
    os.path.join(downloads_dir, "NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx")
]

cols = ['D', 'E', 'F', 'G', 'H', 'I']
# chi phi flm cols correspond to: J, K, L, M, N, O
flm_cols = ['J', 'K', 'L', 'M', 'N', 'O']

for filepath in files_to_fix:
    if not os.path.exists(filepath):
        print(f"File not found: {filepath}")
        continue
    
    print(f"\nFixing formulas in: {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath, data_only=False)
    sheet = wb['NTB – Input']
    
    # Row 7 (TỔNG đơn) = SUM(D8:D10)
    for col in cols:
        sheet[f"{col}7"] = f"=SUM({col}8:{col}10)"
        
    # Row 41 (GIAO) = 'Chi phí FLM'!J37*4/1000000
    # Row 42 (LẤY) = 'Chi phí FLM'!J38*4/1000000
    # Row 43 (TOTAL) = D41+D42
    # Row 44 (Cost/unit) = IF(D7=0,0,D43*1000000/D7)
    # Row 45 (Cost/kg) = 'Chi phí FLM'!J46
    for idx, col in enumerate(cols):
        f_col = flm_cols[idx]
        sheet[f"{col}41"] = f"='Chi phí FLM'!{f_col}37*4/1000000"
        sheet[f"{col}42"] = f"='Chi phí FLM'!{f_col}38*4/1000000"
        sheet[f"{col}43"] = f"={col}41+{col}42"
        sheet[f"{col}44"] = f"=IF({col}7=0,0,{col}43*1000000/{col}7)"
        sheet[f"{col}45"] = f"='Chi phí FLM'!{f_col}46"
        
    wb.save(filepath)
    wb.close()
    print("Saved! Recalculating via Excel COM...")
    
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        workbook = excel.Workbooks.Open(filepath)
        workbook.Save()
        workbook.Close()
        print("Recalculation and save completed successfully!")
    except Exception as e:
        print(f"Error during recalculation: {e}")
    finally:
        excel.Quit()

# Verify the values in NTB_Input_Da_Dien_FLM_CRC_2.xlsx
da_dien_2_path = os.path.join(downloads_dir, "NTB_Input_Da_Dien_FLM_CRC_2.xlsx")
if os.path.exists(da_dien_2_path):
    wb = openpyxl.load_workbook(da_dien_2_path, data_only=True)
    sheet = wb['NTB – Input']
    print("\n=== VERIFIED VALUES in NTB_Input_Da_Dien_FLM_CRC_2.xlsx ===")
    for r in [7, 41, 42, 43, 44, 45]:
        label = sheet.cell(r, 2).value or sheet.cell(r, 1).value
        vals = [sheet.cell(r, openpyxl.utils.column_index_from_string(c)).value for c in cols]
        print(f"Row {r:2d} | {str(label)[:35]:<35} | {vals}")
    wb.close()
