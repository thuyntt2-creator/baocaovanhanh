import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

v2_path = r"C:\Users\lap4all\Downloads\[V2] AOP_Hang_NTB_T7-T12_2026.xlsx"
wb = openpyxl.load_workbook(v2_path, data_only=True)
sheet = wb['0.3 Bưu cục Detail']

print("Search in [V2] 0.3 Bưu cục Detail:")
keywords = ["Ba Đình", "Hòa Ninh", "Đinh Văn", "Tân Hà", "Nam Ban", "Phúc Hưng", "Di Linh", "Đơn Dương", "Đức Trọng", "Hiệp Thạnh", "Nha Trang", "Yết Kiêu", "Phước Đồng", "Hà Quang", "Lê Hồng Phong", "Đức Linh", "Tánh Linh", "Phan Thiết", "Phan Rang", "Đà Lạt", "Bảo Lộc"]

for r in range(3, sheet.max_row + 1):
    name = sheet.cell(r, 2).value
    if name is not None:
        matched = [k for k in keywords if k.lower() in name.lower()]
        if matched:
            print(f"Row {r:2d}: {name} | Match: {matched}")

print("\n" + "="*50)
print("Checking number of rows in [V2] sheet:")
print(f"max_row = {sheet.max_row}")

