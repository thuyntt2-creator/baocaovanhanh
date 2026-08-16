import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

filepath = r"C:\Users\lap4all\Downloads\NTB_Input_Da_Dien_FLM_CRC_2.xlsx"
wb = openpyxl.load_workbook(filepath, data_only=False)
sheet = wb['Chi phí FLM']

print("=== NTB_Input_Da_Dien_FLM_CRC_2.xlsx - Chi phí FLM formulas ===")
for r in range(30, 47):
    label = sheet.cell(r, 2).value or sheet.cell(r, 1).value or f"Row {r}"
    c_val = sheet[f"D{r}"].value
    print(f"Row {r:2d} | {str(label)[:35]:<35} | D: {c_val}")
wb.close()
