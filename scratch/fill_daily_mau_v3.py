import openpyxl, sys, shutil
sys.stdout.reconfigure(encoding='utf-8')

# Paths
config_path = r'C:\Users\lap4all\Downloads\config_psbba_NTB.xlsx'
mau_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v2.xlsx'
out_path = r'C:\Users\lap4all\Downloads\AOP_Hang_Nang_MAU_Vung_T7-T12_2026_final_v3.xlsx'

print("1. Đọc data từ 7_FC_Giao_Daily...")
wb_cfg = openpyxl.load_workbook(config_path, data_only=True)
sheet_fc = wb_cfg['7_FC_Giao_Daily']
mc = sheet_fc.max_column

# Map date column
col_map = {}
for c in range(6, mc):
    header = sheet_fc.cell(1, c).value
    if header:
        parts = str(header).split()
        if len(parts) >= 2:
            col_map[parts[1]] = c # 'dd/mm' -> col_idx

# Define Groups
groups = {
    'BCCK Nha Trang': ['Nha Trang'],
    'BCCK Đơn Dương': ['Đơn Dương', 'Hiệp Thạnh'],
    'BCCK Di Linh': ['Di Linh', 'Hòa Ninh', 'Lâm Hà'],
    'BCCK Đức Linh': ['Đức Linh']
}

daily_vol = {g: {} for g in groups.keys()}

# Accumulate Volume
for r in range(2, sheet_fc.max_row+1):
    bc = sheet_fc.cell(r, 4).value
    band = sheet_fc.cell(r, 5).value
    if bc and band:
        bc_str = str(bc).strip()
        band_str = str(band).strip()
        
        # Only take heavy items (Bulky)
        if 'Bulky' in band_str:
            if bc_str == 'Kho Giao Hàng Nặng - Nha Trang - Khánh Hoà':
                continue
            
            # Find which group this BC belongs to
            matched_group = None
            for g_name, keywords in groups.items():
                for kw in keywords:
                    if kw in bc_str:
                        matched_group = g_name
                        break
                if matched_group:
                    break
            
            if matched_group:
                for date_str, c in col_map.items():
                    vol = sheet_fc.cell(r, c).value or 0
                    daily_vol[matched_group][date_str] = daily_vol[matched_group].get(date_str, 0) + vol

print("2. Đắp data vào file MAU V3...")
shutil.copy(mau_path, out_path)
wb_mau = openpyxl.load_workbook(out_path)

months = ['T7', 'T8', 'T9', 'T10', 'T11', 'T12']
group_names = list(groups.keys())

for m_idx, m_name in enumerate(months):
    sheet_f = wb_mau[f'Forecast {m_name}']
    
    # Identify block start rows
    bc_rows = []
    for r in range(4, 300):
        val = sheet_f.cell(r, 1).value
        # In MAU template, default blocks are named "Kho/BC 1", "Kho/BC 2"... or we already overwrote them
        if val and (str(val).startswith('Kho/BC') or str(val).startswith('BCCK') or str(val).startswith('(BTH)')):
            bc_rows.append(r)
            
    # Clear existing names and data if it's from previous run
    for r in bc_rows:
        sheet_f.cell(r, 1).value = f"Kho/BC {bc_rows.index(r) + 1}"
        for day in range(1, 32):
            sheet_f.cell(r + 1, day + 1).value = None

    # Write new groups
    for i, g_name in enumerate(group_names):
        if i < len(bc_rows):
            r_bc = bc_rows[i]
            sheet_f.cell(r_bc, 1).value = g_name
            
            for day in range(1, 32):
                date_col = day + 1
                month_num = 7 + m_idx
                date_str = f"{day:02d}/{month_num:02d}"
                
                if m_name in ['T7', 'T8'] and date_str in daily_vol[g_name]:
                    sheet_f.cell(r_bc + 1, date_col).value = round(daily_vol[g_name][date_str])
                else:
                    if len(daily_vol[g_name]) > 0:
                        avg = sum(daily_vol[g_name].values()) / len(daily_vol[g_name])
                        sheet_f.cell(r_bc + 1, date_col).value = round(avg)

wb_mau.save(out_path)
print(f"Hoàn thành! Đã lưu: {out_path}")
