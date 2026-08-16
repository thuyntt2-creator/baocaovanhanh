import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

config_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB.xlsx"
wb = openpyxl.load_workbook(config_path, data_only=True)

# 1. Sum Topline_Giao_NTB for >=10kg
sheet_tl = wb['Topline_Giao_NTB']
tl_sums = {m: 0.0 for m in ['2026-07-01', '2026-08-01', '2026-09-01', '2026-10-01', '2026-11-01', '2026-12-01']}
for r in range(2, sheet_tl.max_row + 1):
    m = sheet_tl.cell(r, 1).value
    band = sheet_tl.cell(r, 4).value
    vol = sheet_tl.cell(r, 6).value
    # check if month and band are valid
    if m in tl_sums and band in ['03.10-15kg', '04.15-20kg', '05.>=20kg']:
        if vol is not None:
            tl_sums[m] += float(vol)

print("Topline Giao (>=10kg) sums:")
for k, v in sorted(tl_sums.items()):
    print(f"  {k}: {v:.2f}")

# 2. Sum 3_KQ_BC_Detail for weight groups >= 10kg
sheet_kq = wb['3_KQ_BC_Detail']
kq_sums = {col: 0.0 for col in ['T07/2026', 'T08/2026', 'T09/2026', 'T10/2026', 'T11/2026', 'T12/2026']}
for r in range(2, sheet_kq.max_row + 1):
    grp = sheet_kq.cell(r, 5).value
    if grp in ['10-15kg', '15-30kg', '≥30kg']:
        for col_idx, col in enumerate(['T07/2026', 'T08/2026', 'T09/2026', 'T10/2026', 'T11/2026', 'T12/2026'], start=6):
            val = sheet_kq.cell(r, col_idx).value
            if val is not None:
                kq_sums[col] += float(val)

print("\n3_KQ_BC_Detail (>=10kg) sums:")
for k, v in sorted(kq_sums.items()):
    print(f"  {k}: {v:.2f}")

# 3. Sum 3_KQ_BC_Detail grouped by routed_bc_name to see how many bưu cục have total volumes
bc_volumes = {}
for r in range(2, sheet_kq.max_row + 1):
    bc_name = sheet_kq.cell(r, 3).value
    grp = sheet_kq.cell(r, 5).value
    if bc_name and grp in ['10-15kg', '15-30kg', '≥30kg']:
        bc_volumes.setdefault(bc_name, {col: 0.0 for col in ['T07/2026', 'T08/2026', 'T09/2026', 'T10/2026', 'T11/2026', 'T12/2026']})
        for col_idx, col in enumerate(['T07/2026', 'T08/2026', 'T09/2026', 'T10/2026', 'T11/2026', 'T12/2026'], start=6):
            val = sheet_kq.cell(r, col_idx).value
            if val is not None:
                bc_volumes[bc_name][col] += float(val)

print(f"\nNumber of distinct bưu cục in 3_KQ_BC_Detail with >=10kg: {len(bc_volumes)}")
# show first 5
for bc, vols in list(bc_volumes.items())[:5]:
    print(f"  {bc}: {[round(vols[m], 1) for m in ['T07/2026', 'T08/2026', 'T09/2026', 'T10/2026', 'T11/2026', 'T12/2026']]}")

