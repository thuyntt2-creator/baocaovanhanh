import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['Định biên & Sản lượng']

print("=== Search in Định biên & Sản lượng sheet ===")
for r in range(1, sheet.max_row + 1):
    val = sheet.cell(r, 1).value
    if val is not None:
        val_str = str(val).lower()
        if "setup" in val_str or "di dời" in val_str or "dời" in val_str or "tiện ích" in val_str or "utilities" in val_str:
            print(f"Row {r}: {val}")

wb.close()
