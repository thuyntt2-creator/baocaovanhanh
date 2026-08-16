import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['NTB – Input']

cols = ['D', 'E', 'F', 'G', 'H', 'I']
months = ['T7', 'T8', 'T9', 'T10', 'T11', 'T12']

rows_checked = [8, 9, 11, 12, 13, 17, 20, 22, 27, 32, 37, 38, 39, 41, 42, 44, 45]

print("| Dòng | Tên chỉ số | T7 | T8 | T9 | T10 | T11 | T12 |")
print("|---|---|---|---|---|---|---|---|")
for r in rows_checked:
    row_label = sheet.cell(r, 2).value or sheet.cell(r, 1).value or f"Row {r}"
    vals = []
    for c in cols:
        val = sheet[f"{c}{r}"].value
        if isinstance(val, float):
            vals.append(f"{val:,.2f}" if val % 1 != 0 else f"{int(val):,}")
        elif isinstance(val, int):
            vals.append(f"{val:,}")
        else:
            vals.append(str(val))
    print(f"| {r} | {row_label} | " + " | ".join(vals) + " |")

wb.close()
