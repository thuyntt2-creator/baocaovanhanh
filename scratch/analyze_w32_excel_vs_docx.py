import docx
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

docx_path = r'C:\Users\lap4all\Downloads\NTB - Báo Cáo Tuần (HRBP - ARD) - Tuần 32.docx'
excel_path = r'C:\Users\lap4all\Downloads\BaoCao_Tuan_NTB_W32_2026.xlsx'

doc = docx.Document(docx_path)
wb = openpyxl.load_workbook(excel_path, data_only=True)

print("=== CHECKING ALL SHEETS IN EXCEL ===")
for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    print(f"Sheet '{sheetname}': max_row={ws.max_row}, max_col={ws.max_column}")

print("\n=== MATCHING EXCEL SHEETS TO DOCX TABLES ===")
# Let's dump table headers and first 2 rows of each docx table
for idx, table in enumerate(doc.tables):
    rows = len(table.rows)
    cols = len(table.columns)
    r0 = [cell.text.strip().replace('\n', ' ') for cell in table.rows[0].cells]
    r1 = [cell.text.strip().replace('\n', ' ') for cell in table.rows[1].cells] if rows > 1 else []
    print(f"\n--- TABLE {idx} ({rows}x{cols}) ---")
    print(f"R0: {r0}")
    if r1:
        print(f"R1: {r1}")
