import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

xlsx_aop_path = r"C:\Users\lap4all\Downloads\AOP_Hang_NTB_T7-T12_2026.xlsx"
wb = openpyxl.load_workbook(xlsx_aop_path, data_only=True)

for sheet_name in ['Volume Giao', 'Volume Lấy', '0.3 Bưu cục Detail', 'Mật độ & phương án', 'Kênh & nhu cầu', 'Nguồn lực & chi phí']:
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} (max_row: {sheet.max_row}, max_column: {sheet.max_column}) ---")
        # Check how many non-empty rows there are
        non_empty_count = 0
        sample_rows = []
        for r_idx in range(1, sheet.max_row + 1):
            row_vals = [sheet.cell(r_idx, c_idx).value for c_idx in range(1, sheet.max_column + 1)]
            if any(v is not None for v in row_vals):
                non_empty_count += 1
                if len(sample_rows) < 15:
                    sample_rows.append((r_idx, row_vals[:12]))
        print(f"Non-empty rows count: {non_empty_count}")
        print("First few non-empty rows:")
        for r_idx, vals in sample_rows:
            print(f"  Row {r_idx:2d}: {vals}")
    else:
        print(f"\nSheet {sheet_name} not found!")

