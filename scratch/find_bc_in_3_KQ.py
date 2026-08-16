import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

config_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB (1).xlsx"
if not os.path.exists(config_path):
    config_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB.xlsx"

wb = openpyxl.load_workbook(config_path, data_only=True)
sheet = wb['3_KQ_BC_Detail']

print("=== 3_KQ_BC_Detail (Config) ===")
# Print first 20 rows matching target BCCKs
for r in range(1, sheet.max_row + 1):
    vals = [sheet.cell(r, c).value for c in range(1, 12)]
    if any(v is not None for v in vals):
        bc_name = sheet.cell(r, 3).value # routed_bc_name
        if bc_name and any(name in str(bc_name) for name in ['Nha Trang', 'Đơn Dương', 'Di Linh', 'Đức Linh']):
            print(f"Row {r:3d}: {vals[:5]} | T7={vals[5]} | T8={vals[6]} | T9={vals[7]} | T10={vals[8]} | T11={vals[9]} | T12={vals[10]}")
