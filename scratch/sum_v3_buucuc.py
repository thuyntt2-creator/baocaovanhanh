import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\V3_AOP_Hang_NTB_T7-T12_2026 mới.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['Bưu cục detai;']

print("=== Summing Bưu cục detai; by Month ===")
sums = {f'T{m}': 0.0 for m in range(7, 13)}
for r in range(3, sheet.max_row + 1):
    bc_name = sheet.cell(r, 2).value
    if bc_name is None:
        continue
    for col_idx, m_str in enumerate(['T7', 'T8', 'T9', 'T10', 'T11', 'T12'], start=4):
        val = sheet.cell(r, col_idx).value or 0.0
        sums[m_str] += float(val)

print("Sums from Bưu cục detai;:", sums)
wb.close()
