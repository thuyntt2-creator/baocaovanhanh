import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\lap4all\Downloads\NTB_Input_Con_Thieu_Theo_Template_FLM_CRC.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['NTB – Input']

print("=== Non-empty rows in NTB – Input ===")
for r in range(1, sheet.max_row + 1):
    vals = [sheet.cell(r, c).value for c in range(1, 10)]
    if any(v is not None for v in vals):
        print(f"Row {r:2d} | Col A: {vals[0]} | Col B: {vals[1]} | Col D (T7): {vals[3]}")
wb.close()
