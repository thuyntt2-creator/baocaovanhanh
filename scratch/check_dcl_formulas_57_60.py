import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

filepath = r"C:\Users\lap4all\Downloads\Telegram Desktop\CRC_DCL_Input_FLM_2026.xlsx"
wb = openpyxl.load_workbook(filepath, data_only=False)
sheet = wb['ĐCL – Input']

print("=== CRC_DCL_Input_FLM_2026.xlsx rows 57-60 formulas ===")
cols = ['D', 'E', 'F', 'G', 'H', 'I']
for r in [57, 58, 59]:
    label = sheet.cell(r, 2).value or sheet.cell(r, 1).value or f"Row {r}"
    print(f"Row {r:2d} | {str(label)[:35]:<35}")
    for c in cols:
        print(f"  {c}: {sheet[f'{c}{r}'].value}")
wb.close()
