import os
import sys
import io
import openpyxl

# Fix encoding for Windows
os.environ['PYTHONIOENCODING'] = 'utf-8'
try:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
except AttributeError:
    pass

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def main():
    excel_path = os.path.join(BASE_DIR, 'manual_raw.xlsx')
    if not os.path.exists(excel_path):
        print("manual_raw.xlsx not found.")
        return
        
    target_str = '21624000'
    target_num = 21624000
    
    print(f"Scanning manual_raw.xlsx for '{target_str}'...")
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        for sname in wb.sheetnames:
            # Skip large raw sheets if we want to be super fast
            if sname in ['data thô']:
                continue
            ws = wb[sname]
            print(f"  Scanning sheet: {sname}...")
            for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                for c_idx, val in enumerate(row, 1):
                    if val == target_str or val == target_num:
                        print(f"    🎯 Found in sheet '{sname}' at Row {r_idx}, Col {c_idx}! Row content: {list(row)[:10]}")
    except Exception as e:
        print("Error:", e)

if __name__ == "__main__":
    main()
