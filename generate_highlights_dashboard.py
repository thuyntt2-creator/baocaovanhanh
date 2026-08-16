# -*- coding: utf-8 -*-
"""
Script: generate_highlights_dashboard.py
Author: Antigravity AI
Description: Consolidates operational alert metrics for post offices on N-1 date,
             ranks them by severity, writes a formatted table to a new Google Sheet tab,
             and exports a local markdown report.
"""

import os
import sys
import argparse
import unicodedata
import pandas as pd
from datetime import datetime, timedelta
import gspread
from google.oauth2.service_account import Credentials

# Fix console encoding for Vietnamese character support
try:
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    if hasattr(sys.stderr, 'reconfigure'):
        sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

# ============ CONFIG & CONSTANTS ============
MAIN_SPREADSHEET_ID = "1JZ1eRerRqrpwjZ4HBevQunjd8VquM_cvPFz12TaJfMQ"
ROT_LC_SPREADSHEET_ID = "14r8n9L2cIG1Bmz8kSH79B24QzmnOApZhniGyOU40hr4"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
JSON_FILE = os.path.join(BASE_DIR, 'credentials.json')

SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]

# HSL tailored color palette mapped to Google Sheets decimal colors (RGB 0-1)
COLOR_SYSTEM = {
    'darkBlue':   {'red': 0.106, 'green': 0.212, 'blue': 0.365},  # Header background
    'lightBlue':  {'red': 0.902, 'green': 0.941, 'blue': 0.980},  # Zebra stripe alt row
    'white':      {'red': 1.000, 'green': 1.000, 'blue': 1.000},  # White background
    'black':      {'red': 0.000, 'green': 0.000, 'blue': 0.000},  # Text color
    'grayLine':   {'red': 0.863, 'green': 0.863, 'blue': 0.863},  # Borders
    
    # Severity Colors
    'highBg':     {'red': 1.000, 'green': 0.878, 'blue': 0.878},  # Light red for Critical/High
    'highFg':     {'red': 0.753, 'green': 0.000, 'blue': 0.000},  # Dark red text
    'warnBg':     {'red': 1.000, 'green': 0.949, 'blue': 0.800},  # Light yellow for Warning/Medium
    'warnFg':     {'red': 0.600, 'green': 0.400, 'blue': 0.000},  # Dark yellow text
    'normalBg':   {'red': 0.886, 'green': 0.937, 'blue': 0.851},  # Light green for Normal/Low
    'normalFg':   {'red': 0.216, 'green': 0.341, 'blue': 0.137},  # Dark green text
}

# ============ STRING HELPERS ============
def normalize_str(s):
    if not s:
        return ""
    return unicodedata.normalize('NFC', str(s).strip().lower())

def clean_bc_name(name):
    name = normalize_str(name)
    for tag in ['(dno)', '(ldo)', '(kho)', '(bth)', '(nth)']:
        name = name.replace(tag, "")
    for prefix in ['kho chuyển tiếp', 'kho trung chuyển', 'điểm xử lý hàng', 'điểm lấy hàng', 'bưu cục', 'bc', 'đl']:
        name = name.replace(prefix, "")
    name = name.replace("-", " ").replace("_", " ")
    return " ".join(name.split())

def match_po_name(raw_name, standard_list):
    raw_norm = normalize_str(raw_name)
    for std in standard_list:
        if normalize_str(std) == raw_norm:
            return std
            
    raw_clean = clean_bc_name(raw_name)
    if not raw_clean:
        return None
        
    cleaned_std_list = [(std, clean_bc_name(std)) for std in standard_list]
    for std, std_clean in cleaned_std_list:
        if std_clean == raw_clean:
            return std
            
    matches = []
    for std, std_clean in cleaned_std_list:
        if std_clean and (raw_clean in std_clean or std_clean in raw_clean):
            matches.append(std)
            
    if matches:
        matches.sort(key=lambda x: abs(len(clean_bc_name(x)) - len(raw_clean)))
        return matches[0]
        
    return None

def parse_percent_to_float(val):
    if val is None or val == "":
        return 0.0
    try:
        val_str = str(val).replace("%", "").replace(",", ".").strip()
        num = float(val_str)
        if num < 1.0 and num > 0:
            return num * 100.0
        return num
    except ValueError:
        return 0.0

def parse_int(val):
    if val is None or val == "":
        return 0
    try:
        return int(str(val).replace(",", "").replace(".", "").strip())
    except ValueError:
        return 0

# ============ GSHEET CELL FORMATTING HELPERS ============
def make_cell(value, bg=None, bold=False, fg=None, halign='LEFT', is_header=False):
    d = {'userEnteredValue': {}}
    if isinstance(value, (int, float)):
        d['userEnteredValue']['numberValue'] = value
    else:
        d['userEnteredValue']['stringValue'] = str(value) if value is not None else ''

    fmt_obj = {
        'textFormat': {
            'bold': bold,
            'foregroundColor': COLOR_SYSTEM[fg] if fg else COLOR_SYSTEM['black'],
            'fontFamily': 'Arial',
            'fontSize': 10 if not is_header else 10,
        },
        'horizontalAlignment': halign,
        'verticalAlignment': 'MIDDLE',
    }
    if bg:
        fmt_obj['backgroundColor'] = COLOR_SYSTEM[bg]
    
    # Default border format
    border_color = COLOR_SYSTEM['grayLine']
    fmt_obj['borders'] = {
        'top':    {'style': 'SOLID', 'color': border_color},
        'bottom': {'style': 'SOLID', 'color': border_color},
        'left':   {'style': 'SOLID', 'color': border_color},
        'right':  {'style': 'SOLID', 'color': border_color},
    }
    
    d['userEnteredFormat'] = fmt_obj
    return d

def write_sheet_data(sh, rows_data):
    max_cols = max(len(r) for r in rows_data)
    for r in rows_data:
        while len(r) < max_cols:
            r.append(make_cell(''))

    body = {
        'requests': [{
            'updateCells': {
                'rows': [{'values': row} for row in rows_data],
                'fields': 'userEnteredValue,userEnteredFormat',
                'start': {'sheetId': sh.id, 'rowIndex': 0, 'columnIndex': 0}
            }
        }]
    }
    sh.spreadsheet.batch_update(body)

def set_col_widths(sh, widths):
    requests = []
    for i, w in enumerate(widths):
        requests.append({
            'updateDimensionProperties': {
                'range': {
                    'sheetId': sh.id,
                    'dimension': 'COLUMNS',
                    'startIndex': i,
                    'endIndex': i+1,
                },
                'properties': {'pixelSize': w},
                'fields': 'pixelSize'
            }
        })
    try:
        sh.spreadsheet.batch_update({'requests': requests})
    except Exception as e:
        print(f"⚠️ Failed to set column widths: {e}")

# ============ MAIN PIPELINE ============
def main():
    parser = argparse.ArgumentParser(description="Generate morning highlight dashboard for post offices.")
    parser.add_argument("--date", type=str, help="Target date in YYYY-MM-DD format. Defaults to yesterday.")
    args = parser.parse_args()

    # Determine target date
    if args.date:
        try:
            target_date = datetime.strptime(args.date.strip(), "%Y-%m-%d")
        except ValueError:
            print("❌ Invalid date format. Use YYYY-MM-DD.")
            sys.exit(1)
    else:
        target_date = datetime.now() - timedelta(days=1)
        
    date_str_iso = target_date.strftime("%Y-%m-%d")
    
    weekday_map = {
        0: "Thứ 2", 1: "Thứ 3", 2: "Thứ 4", 3: "Thứ 5", 4: "Thứ 6", 5: "Thứ 7", 6: "Chủ Nhật"
    }
    date_str_sheet = f"{date_str_iso} - {weekday_map[target_date.weekday()]}"
    date_str_dmy = target_date.strftime("%d/%m/%Y")
    
    print(f"📅 Running Highlights Analysis for: {date_str_iso} ({weekday_map[target_date.weekday()]})")
    
    # Authenticate with Google sheets
    if not os.path.exists(JSON_FILE):
        print(f"❌ Credentials file not found at: {JSON_FILE}")
        sys.exit(1)
        
    credentials = Credentials.from_service_account_file(JSON_FILE, scopes=SCOPES)
    gc = gspread.authorize(credentials)
    
    # Load Master CoCauVung
    print("📖 Loading Master CoCauVung sheet...")
    sh_main = gc.open_by_key(MAIN_SPREADSHEET_ID)
    ws_cc = sh_main.worksheet("CoCauVung")
    cc_rows = ws_cc.get_all_values()
    df_cc = pd.DataFrame(cc_rows[1:], columns=cc_rows[0])
    df_cc = df_cc[df_cc['Bưu cục'].str.strip() != '']
    master_pos = df_cc['Bưu cục'].str.strip().tolist()
    po_am_map = dict(zip(df_cc['Bưu cục'].str.strip(), df_cc['AM'].str.strip()))
    
    # Initialize metrics structure
    po_metrics = {po: {
        'po_name': po,
        'am_name': po_am_map[po],
        'alerts': [],
        'severity_score': 0,
        # Data
        'all_vol': 0, 'all_gan_vol': 0, 'all_gan_rate': 0.0,
        'all_gan_ca1': 0.0, 'all_gan_ca2': 0.0,
        'all_gtc_ca1': 0.0, 'all_gtc_ca2': 0.0,
        'all_gtc_vol': 0, 'all_gtc_rate': 0.0,
        'all_chua_gan': 0, 'all_ton': 0,
        # TTS
        'tts_vol': 0, 'tts_gan_vol': 0, 'tts_gan_rate': 0.0,
        'tts_gtc_vol': 0, 'tts_gtc_rate': 0.0,
        'tts_chua_gan': 0, 'tts_ton': 0,
        # OPR
        'opr_vol': 0, 'opr_ontime': 0, 'opr_rate': 0.0, 'opr_late': 0,
        # Aging
        'aging_total': 0, 'aging_15d_plus': 0,
        # Treo LC
        'treo_lc': 0,
        # FD
        'fd_rate': 0.0, 'fd_vs_n1': 0.0, 'fd_ty_trong_tra': 0.0,
        # Rot LC
        'rot_lc_tts': 0.0, 'rot_lc_shopee': 0.0
    } for po in master_pos}

    # 1. PROCESS DATA SHEET
    print("📖 Loading 'Data' worksheet...")
    ws_data = sh_main.worksheet("Data")
    data_rows = ws_data.get_all_values()
    df_data = pd.DataFrame(data_rows[1:], columns=data_rows[0])
    df_data_day = df_data[df_data['Time'] == date_str_sheet]
    
    for idx, row in df_data_day.iterrows():
        std_po = match_po_name(row['Chi tiết'], master_pos)
        if not std_po: continue
        m = po_metrics[std_po]
        loai_hang = row['Loại Hàng'].strip()
        m['all_vol'] += parse_int(row['Volume'])
        m['all_gan_vol'] += parse_int(row['Sản Lượng Gán'])
        m['all_gtc_vol'] += parse_int(row['Sản Lượng Giao Thành Công'])
        m['all_chua_gan'] += parse_int(row['Sản Lượng Chưa Gán'])
        m['all_ton'] += parse_int(row['Sản Lượng Tồn'])
        if loai_hang == "Hàng Mới Ca 1":
            m['all_gan_ca1'] = parse_percent_to_float(row['% Gán'])
            m['all_gtc_ca1'] = parse_percent_to_float(row['% GTC'])
        elif loai_hang == "Hàng Mới Ca 2":
            m['all_gan_ca2'] = parse_percent_to_float(row['% Gán'])
            m['all_gtc_ca2'] = parse_percent_to_float(row['% GTC'])

    # Recalculate GTC/Gán rates
    for po, m in po_metrics.items():
        if m['all_vol'] > 0:
            m['all_gan_rate'] = (m['all_gan_vol'] / m['all_vol']) * 100.0
            m['all_gtc_rate'] = (m['all_gtc_vol'] / m['all_vol']) * 100.0

    # 2. PROCESS TTS SHEET
    print("📖 Loading 'TTS' worksheet...")
    ws_tts = sh_main.worksheet("TTS")
    tts_rows = ws_tts.get_all_values()
    df_tts = pd.DataFrame(tts_rows[1:], columns=tts_rows[0])
    df_tts_day = df_tts[df_tts['Time'] == date_str_sheet]
    
    for idx, row in df_tts_day.iterrows():
        std_po = match_po_name(row['Chi tiết'], master_pos)
        if not std_po: continue
        m = po_metrics[std_po]
        m['tts_vol'] += parse_int(row['Volume'])
        m['tts_gan_vol'] += parse_int(row['Sản Lượng Gán'])
        m['tts_gtc_vol'] += parse_int(row['Sản Lượng Giao Thành Công'])
        m['tts_chua_gan'] += parse_int(row['Sản Lượng Chưa Gán'])
        m['tts_ton'] += parse_int(row['Sản Lượng Tồn'])

    for po, m in po_metrics.items():
        if m['tts_vol'] > 0:
            m['tts_gan_rate'] = (m['tts_gan_vol'] / m['tts_vol']) * 100.0
            m['tts_gtc_rate'] = (m['tts_gtc_vol'] / m['tts_vol']) * 100.0

    # 3. PROCESS OPR SHEET
    print("📖 Loading 'OPR' worksheet...")
    ws_opr = sh_main.worksheet("OPR")
    opr_rows = ws_opr.get_all_values()
    df_opr = pd.DataFrame(opr_rows[1:], columns=opr_rows[0])
    df_opr_day = df_opr[df_opr['NgayLTC'] == date_str_iso]
    
    for idx, row in df_opr_day.iterrows():
        std_po = match_po_name(row['kholay'], master_pos)
        if not std_po: continue
        m = po_metrics[std_po]
        m['opr_vol'] += parse_int(row['vol_ltc'])
        m['opr_ontime'] += parse_int(row['ot'])
        m['opr_late'] += parse_int(row['Đơn trễ'])
        
    for po, m in po_metrics.items():
        if m['opr_vol'] > 0:
            m['opr_rate'] = (m['opr_ontime'] / m['opr_vol']) * 100.0

    # 4. PROCESS AGING BACKLOG SHEET
    print("📖 Loading 'Aging trên 5 ngày' worksheet...")
    ws_aging = sh_main.worksheet("Aging trên 5 ngày")
    aging_rows = ws_aging.get_all_values()
    df_aging = pd.DataFrame(aging_rows[1:], columns=aging_rows[0])
    
    for idx, row in df_aging.iterrows():
        std_po = match_po_name(row['bc'], master_pos)
        if not std_po: continue
        m = po_metrics[std_po]
        m['aging_total'] += 1
        nhom_bl = str(row['Nhóm BL']).strip()
        if '(k) > 15 ngày' in nhom_bl:
            m['aging_15d_plus'] += 1

    # 5. PROCESS HANGING TRANSIT SHEET
    print("📖 Loading 'Treo LC' worksheet...")
    ws_treo = sh_main.worksheet("Treo LC")
    treo_rows = ws_treo.get_all_values()
    df_treo = pd.DataFrame(treo_rows[1:], columns=treo_rows[0])
    for idx, row in df_treo.iterrows():
        std_po = match_po_name(row['warehouse_name'], master_pos)
        if std_po:
            po_metrics[std_po]['treo_lc'] += 1

    # 6. PROCESS FD SHEET
    print("📖 Loading 'FD ' worksheet...")
    ws_fd = sh_main.worksheet("FD ")
    fd_rows = ws_fd.get_all_values()
    fd_headers = ['Bưu Cục', 'AM', '%FD (N)', '%FD (N-1)', 'vs N-1', '%FD (N-7)', 'vs N-7', 'Vol giao', 'Vol trả', 'Tỷ trọng trả']
    df_fd = pd.DataFrame(fd_rows[3:], columns=fd_headers + [f'Col_{i}' for i in range(10, len(fd_rows[3]))])
    
    for idx, row in df_fd.iterrows():
        std_po = match_po_name(row['Bưu Cục'], master_pos)
        if not std_po: continue
        m = po_metrics[std_po]
        m['fd_rate'] = parse_percent_to_float(row['%FD (N)'])
        m['fd_ty_trong_tra'] = parse_percent_to_float(row['Tỷ trọng trả'])
        vs_str = str(row['vs N-1']).replace("▲", "+").replace("▼", "-").replace("%", "").strip()
        try:
            m['fd_vs_n1'] = float(vs_str) if vs_str not in ["", "—", "-", "N/A"] else 0.0
        except ValueError:
            m['fd_vs_n1'] = 0.0

    # 7. PROCESS ROT LC SHEET
    print("📖 Loading Rot LC Spreadsheet...")
    try:
        sh_rot = gc.open_by_key(ROT_LC_SPREADSHEET_ID)
        # Data TTS
        ws_rot_tts = sh_rot.worksheet("Data TTS")
        rot_tts_rows = ws_rot_tts.get_all_values()
        df_rot_tts = pd.DataFrame(rot_tts_rows[1:], columns=rot_tts_rows[0])
        df_rot_tts_day = df_rot_tts[df_rot_tts['Loại ngày'] == date_str_iso]
        for idx, row in df_rot_tts_day.iterrows():
            std_po = match_po_name(row['Chi tiết'], master_pos)
            if std_po:
                po_metrics[std_po]['rot_lc_tts'] = parse_percent_to_float(row['%_rot_lc'])
                
        # Data Shopee
        ws_rot_shp = sh_rot.worksheet("Data Shopee")
        rot_shp_rows = ws_rot_shp.get_all_values()
        df_rot_shp = pd.DataFrame(rot_shp_rows[1:], columns=rot_shp_rows[0])
        df_rot_shp_day = df_rot_shp[df_rot_shp['Loại ngày'] == date_str_iso]
        for idx, row in df_rot_shp_day.iterrows():
            std_po = match_po_name(row['Chi tiết'], master_pos)
            if std_po:
                po_metrics[std_po]['rot_lc_shopee'] = parse_percent_to_float(row['%_rot_lc'])
    except Exception as e:
        print(f"⚠️ Warning processing Rot LC: {e}")

    # ============ ISSUE EVALUATION ENGINE ============
    print("\n🔍 Evaluating alert triggers and calculating severity...")
    alert_pos_list = []
    
    for po, m in po_metrics.items():
        alerts = []
        
        # 1. FD ALERT
        is_priority_fd = (m['fd_rate'] > 4.5 and m['fd_ty_trong_tra'] > 3.0)
        is_extreme_fd = (m['fd_rate'] > 10.0)
        is_spiked_fd = (m['fd_vs_n1'] >= 3.0)
        if is_priority_fd or is_extreme_fd or is_spiked_fd:
            desc = f"FD cao {m['fd_rate']:.1f}%"
            if m['fd_vs_n1'] > 0:
                desc += f" (+{m['fd_vs_n1']:.1f}%)"
            alerts.append(('FD', desc, 3 if is_extreme_fd else 2))
            
        # 2. OPERATIONAL ALERT
        is_bad_gtc = (m['all_vol'] > 0 and m['all_gtc_rate'] < 45.0)
        is_bad_ca2_gtc = (m['all_gtc_ca2'] > 0 and m['all_gtc_ca2'] < 40.0)
        is_high_unassigned = (m['all_chua_gan'] > 100)
        is_high_backlogs = (m['all_ton'] > 150)
        if is_bad_gtc or is_bad_ca2_gtc or is_high_unassigned or is_high_backlogs:
            details = []
            if is_bad_gtc: details.append(f"GTC {m['all_gtc_rate']:.1f}%")
            if is_high_unassigned: details.append(f"chưa gán {m['all_chua_gan']} đơn")
            if is_high_backlogs: details.append(f"tồn {m['all_ton']} đơn")
            alerts.append(('OPERATIONAL', "Vận hành: " + ", ".join(details), 2 if (is_high_backlogs and m['all_ton'] > 400) else 1))

        # 3. TTS OPERATIONAL ALERT
        is_bad_tts_gtc = (m['tts_vol'] > 0 and m['tts_gtc_rate'] < 45.0)
        is_high_tts_unassigned = (m['tts_chua_gan'] > 40)
        is_high_tts_backlogs = (m['tts_ton'] > 60)
        if is_bad_tts_gtc or is_high_tts_unassigned or is_high_tts_backlogs:
            details = []
            if is_bad_tts_gtc: details.append(f"GTC TTS {m['tts_gtc_rate']:.1f}%")
            if is_high_tts_backlogs: details.append(f"tồn TTS {m['tts_ton']} đơn")
            alerts.append(('TTS_OPERATIONAL', "TTS: " + ", ".join(details), 1))

        # 4. OPR TTS ALERT
        if m['opr_vol'] > 0 and m['opr_rate'] < 80.0 and m['opr_late'] > 5:
            alerts.append(('OPR_TTS', f"Trễ OPR {m['opr_late']} đơn ({m['opr_rate']:.1f}%)", 1))

        # 5. AGING BACKLOG ALERT
        if m['aging_total'] > 15 or m['aging_15d_plus'] > 1:
            alerts.append(('AGING', f"Aging >5d: {m['aging_total']} đơn (có {m['aging_15d_plus']} đơn >15d)", 2 if m['aging_total'] > 50 else 1))

        # 6. HANGING TRANSIT LC ALERT
        if m['treo_lc'] > 40:
            alerts.append(('TREO_LC', f"Treo LC >36h: {m['treo_lc']} đơn", 1))

        # 7. LOST TRANSIT ALERT
        if m['rot_lc_tts'] > 5.0 or m['rot_lc_shopee'] > 5.0:
            details = []
            if m['rot_lc_tts'] > 5.0: details.append(f"TTS {m['rot_lc_tts']:.1f}%")
            if m['rot_lc_shopee'] > 5.0: details.append(f"Shopee {m['rot_lc_shopee']:.1f}%")
            alerts.append(('ROT_LC', "Rớt LC: " + ", ".join(details), 1))

        m['alerts'] = alerts
        # Compute severity score based on number of alerts and their weight
        m['severity_score'] = sum(weight for _, _, weight in alerts)
        
        if alerts:
            alert_pos_list.append(m)

    # Sort post offices by severity score descending
    alert_pos_list.sort(key=lambda x: -x['severity_score'])
    print(f"  • Found {len(alert_pos_list)} post offices with operational warnings.")

    # ============ 8. GENERATE LOCAL REPORT ============
    md_report_rows = []
    md_report_rows.append(f"# 📊 Báo Cáo Cảnh Báo Vận Hành Đầu Ngày - NTB - Ngày {date_str_dmy}")
    md_report_rows.append(f"*Dữ liệu được phân tích cho ngày N-1: {date_str_iso} ({weekday_map[target_date.weekday()]})*")
    md_report_rows.append(f"*Tổng số bưu cục cần chú ý đặc biệt: **{len(alert_pos_list)}***\n")
    md_report_rows.append("| STT | Bưu Cục | AM Quản Lý | Mức Độ | Số Lượng Cảnh Báo | KPIs Vi Phạm |")
    md_report_rows.append("|---|---|---|---|---|---|")

    stt = 1
    for m in alert_pos_list:
        po_name = m['po_name']
        am_name = m['am_name']
        n_alerts = len(m['alerts'])
        score = m['severity_score']
        
        # Determine Severity Level
        if score >= 4 or n_alerts >= 3:
            sev_label = "Nguy Cấp"
        else:
            sev_label = "Cảnh Báo"
            
        kpis_violated = "; ".join([desc for _, desc, _ in m['alerts']])
        
        # Append MD Table Row
        md_report_rows.append(f"| {stt} | **{po_name}** | {am_name} | `{sev_label}` | {n_alerts} | {kpis_violated} |")
        stt += 1

    # Write Markdown output
    md_report_rows.append("\n---\n")
    md_report_rows.append("## Detail Action Points by AM")
    
    # Group by AM
    am_groups = {}
    for m in alert_pos_list:
        am = m['am_name']
        if am not in am_groups: am_groups[am] = []
        am_groups[am].append(m)
        
    for am in sorted(list(am_groups.keys())):
        md_report_rows.append(f"### 👤 AM: {am}")
        for idx, m in enumerate(am_groups[am]):
            alerts_desc = "; ".join([f"**{desc}**" for _, desc, _ in m['alerts']])
            md_report_rows.append(f"{idx+1}. **{m['po_name']}**: Gặp các lỗi ({alerts_desc}). Yêu cầu AM đôn đốc bưu cục xử lý gấp.")
        md_report_rows.append("")

    md_output = "\n".join(md_report_rows)
    md_file_path = os.path.join(BASE_DIR, "morning_operational_highlights.md")
    with open(md_file_path, "w", encoding="utf-8") as f:
        f.write(md_output)
    print(f"💾 Successfully saved local markdown highlights file to: {md_file_path}")

    # ============ 9. GENERATE LOCAL EXCEL REPORT ============
    excel_file_path = os.path.join(BASE_DIR, "morning_operational_highlights.xlsx")
    print(f"📖 Generating local Excel report at: {excel_file_path}...")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Cảnh báo đầu ngày"
        
        # Enable grid lines visibility
        ws.views.sheetView[0].showGridLines = True
        
        # Styles
        title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        body_font = Font(name='Arial', size=9, bold=False, color='000000')
        
        title_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Dark Blue
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid') # Grayish Blue
        
        high_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid') # Soft Red/Orange
        high_font = Font(name='Arial', size=9, bold=True, color='C00000') # Dark Red
        
        warn_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') # Soft Yellow
        warn_font = Font(name='Arial', size=9, bold=True, color='7F6000') # Dark Yellow
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')
        
        # 1. Main Title
        ws.merge_cells("A1:G2")
        title_cell = ws["A1"]
        title_cell.value = "BẢNG CẢNH BÁO VẬN HÀNH ĐẦU NGÀY - KHU VỰC NTB"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = align_center
        
        # Fill style for merged cells in title block
        for row in range(1, 3):
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = title_fill
                
        # 2. Date info
        ws.merge_cells("A3:G3")
        info_cell = ws["A3"]
        info_cell.value = f"Ngày phân tích: {date_str_dmy} (Dữ liệu N-1: {date_str_iso}) | Tổng số bưu cục bất ổn: {len(alert_pos_list)}"
        info_cell.font = Font(name='Arial', size=10, bold=True, color='333333')
        info_cell.alignment = align_left
        
        # 3. Table Headers
        headers = ["STT", "Bưu Cục", "AM Quản Lý", "Mức Cảnh Báo", "Số Cảnh Báo", "Các Chỉ Số Vi Phạm (KPIs)", "Hành Động Cần Làm"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
            
        # 4. Table Rows
        row_idx = 5
        for idx, m in enumerate(alert_pos_list, 1):
            n_alerts = len(m['alerts'])
            score = m['severity_score']
            
            # Determine Alert Level
            if score >= 4 or n_alerts >= 3:
                sev_label = "Nguy Cấp"
                row_fill = high_fill
                row_font = high_font
            else:
                sev_label = "Cảnh Báo"
                row_fill = warn_fill
                row_font = warn_font
                
            kpis_violated = "; ".join([desc for _, desc, _ in m['alerts']])
            
            # Directives
            directives = []
            for alert_type, desc, _ in m['alerts']:
                if alert_type == 'FD':
                    directives.append("Check nguyên nhân %FD cao")
                elif alert_type == 'OPERATIONAL':
                    directives.append("Giải tỏa tồn kho/chưa gán")
                elif alert_type == 'TTS_OPERATIONAL':
                    directives.append("Đẩy nhanh xử lý đơn TTS")
                elif alert_type == 'OPR_TTS':
                    directives.append("Rà soát đơn trễ lấy OPR")
                elif alert_type == 'AGING':
                    directives.append("Giải quyết hàng tồn lâu ngày")
                elif alert_type == 'TREO_LC':
                    directives.append("Xử lý đơn treo luân chuyển")
                elif alert_type == 'ROT_LC':
                    directives.append("Check khâu đóng bao/bàn giao tải")
            action_needed = " + ".join(list(set(directives)))
            
            row_data = [idx, m['po_name'], m['am_name'], sev_label, n_alerts, kpis_violated, action_needed]
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = row_font if col_idx in [1, 2, 4, 5] else body_font
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = align_center if col_idx in [1, 4, 5] else align_left
                
            row_idx += 1
            
        # Set row heights
        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 22
        ws.row_dimensions[4].height = 25
        for r in range(5, row_idx):
            ws.row_dimensions[r].height = 22
            
        # Column dimensions auto-fit or specific widths
        col_widths = {
            'A': 6,   # STT
            'B': 26,  # Bưu cục
            'C': 22,  # AM
            'D': 16,  # Mức Cảnh Báo
            'E': 14,  # Số Cảnh Báo
            'F': 55,  # KPIs Vi Phạm
            'G': 55,  # Hành động
        }
        for col_let, width in col_widths.items():
            ws.column_dimensions[col_let].width = width
            
        wb.save(excel_file_path)
        print(f"💾 Successfully saved local Excel highlights report to: {excel_file_path}")
    except Exception as e:
        print(f"❌ Failed to generate local Excel report: {e}")

    # Output details for terminal capture
    print("\n--- RESULTS PREVIEW ---")
    print(md_output[:1000] + "\n...(truncated)...")
    print("-----------------------")

if __name__ == "__main__":
    main()
