import sys
import os
import shutil
import openpyxl
from datetime import datetime, date

# Set stdout encoding to UTF-8
sys.stdout.reconfigure(encoding='utf-8')

config_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB (1).xlsx"
backup_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB (1)_backup.xlsx"

def parse_header_date(col_name):
    if not isinstance(col_name, str):
        return None
    parts = col_name.strip().split()
    if len(parts) == 2:
        try:
            day_str, month_str = parts[1].split('/')
            day = int(day_str)
            month = int(month_str)
            return date(2026, month, day)
        except Exception:
            return None
    return None

def main():
    print("🚀 Bắt đầu tiến trình ghi thẳng kết quả đã tách vào file config...")

    if not os.path.exists(config_path):
        print(f"❌ Không tìm thấy file cấu hình tại: {config_path}")
        sys.exit(1)

    # 1. Tạo file backup trước khi ghi đè
    print(f"📦 Tạo file backup tại: {backup_path}...")
    shutil.copyfile(config_path, backup_path)
    print("✅ Đã sao lưu file gốc thành công.")

    # 2. Đọc dữ liệu (dùng data_only=True để đọc đúng giá trị số)
    print("📖 Đọc dữ liệu tính toán từ file config...")
    wb_data = openpyxl.load_workbook(config_path, data_only=True)
    
    # 2.1 Đọc 1_Config_Chuan
    sheet_chuan = wb_data['1_Config_Chuan']
    chuan_rows = list(sheet_chuan.iter_rows(values_only=True))
    chuan_headers = chuan_rows[0]
    h_idx = {h: i for i, h in enumerate(chuan_headers)}
    
    nt_wards = []
    for row in chuan_rows[1:]:
        if row[h_idx['Quận_Tên']] == 'Thành phố Nha Trang':
            nt_wards.append({
                'Phường_ID': row[h_idx['Phường_ID']],
                'Phường_Tên': row[h_idx['Phường_Tên']],
                'BC_Lấy_ID': row[h_idx['BC_Lấy_ID']],
                'BC_Lấy_Tên': row[h_idx['BC_Lấy_Tên']],
                'BC_Giao_ID': row[h_idx['BC_Giao_ID']],
                'BC_Giao_Tên': row[h_idx['BC_Giao_Tên']]
            })
            
    # 2.2 Đọc 4_KQ_B2B
    sheet_b2b = wb_data['4_KQ_B2B']
    b2b_rows = list(sheet_b2b.iter_rows(values_only=True))
    b2b_headers = b2b_rows[0]
    month_cols = ['T07/2026', 'T08/2026', 'T09/2026', 'T10/2026', 'T11/2026', 'T12/2026']
    
    ward_monthly_vols = {} # ward_monthly_vols[p_id][month] = vol
    for row in b2b_rows[1:]:
        if row[1] == 'Khánh Hòa' and row[2] == 1548.0:
            p_id = row[3]
            ward_monthly_vols.setdefault(p_id, {col: 0.0 for col in month_cols})
            for col in month_cols:
                col_idx = b2b_headers.index(col)
                val = row[col_idx]
                ward_monthly_vols[p_id][col] += float(val) if val is not None else 0.0

    # 2.3 Tính tổng monthly B2B của các phường trực thuộc mỗi BC Lấy và BC Giao
    bc_wards_monthly_lay = {}
    bc_wards_monthly_giao = {}
    
    for w in nt_wards:
        p_id = w['Phường_ID']
        lay_id = w['BC_Lấy_ID']
        giao_id = w['BC_Giao_ID']
        w_vols = ward_monthly_vols.get(p_id, {col: 0.0 for col in month_cols})
        
        # BC Lấy
        if lay_id:
            bc_wards_monthly_lay.setdefault(lay_id, {col: 0.0 for col in month_cols})
            for col in month_cols:
                bc_wards_monthly_lay[lay_id][col] += w_vols[col]
                
        # BC Giao
        if giao_id:
            bc_wards_monthly_giao.setdefault(giao_id, {col: 0.0 for col in month_cols})
            for col in month_cols:
                bc_wards_monthly_giao[giao_id][col] += w_vols[col]

    # Tính ratio cho từng phường trong mỗi BC
    ratio_lay_map = {}
    ratio_giao_map = {}
    
    for w in nt_wards:
        p_id = w['Phường_ID']
        lay_id = w['BC_Lấy_ID']
        giao_id = w['BC_Giao_ID']
        w_vols = ward_monthly_vols.get(p_id, {col: 0.0 for col in month_cols})
        
        # BC Lấy
        if lay_id:
            ratio_lay_map.setdefault(lay_id, {})
            ratio_lay_map[lay_id].setdefault(p_id, {})
            same_bc_wards = [x['Phường_ID'] for x in nt_wards if x['BC_Lấy_ID'] == lay_id]
            for col in month_cols:
                tot_bc = bc_wards_monthly_lay[lay_id][col]
                if tot_bc > 0:
                    ratio_lay_map[lay_id][p_id][col] = w_vols[col] / tot_bc
                else:
                    ratio_lay_map[lay_id][p_id][col] = 1.0 / len(same_bc_wards)
                    
        # BC Giao
        if giao_id:
            ratio_giao_map.setdefault(giao_id, {})
            ratio_giao_map[giao_id].setdefault(p_id, {})
            same_bc_wards = [x['Phường_ID'] for x in nt_wards if x['BC_Giao_ID'] == giao_id]
            for col in month_cols:
                tot_bc = bc_wards_monthly_giao[giao_id][col]
                if tot_bc > 0:
                    ratio_giao_map[giao_id][p_id][col] = w_vols[col] / tot_bc
                else:
                    ratio_giao_map[giao_id][p_id][col] = 1.0 / len(same_bc_wards)

    # 3. Mở workbook gốc dạng data_only=False để giữ công thức của các sheet khác
    print("📖 Mở file config để ghi các sheet mới (giữ nguyên công thức)...")
    wb_write = openpyxl.load_workbook(config_path, data_only=False)

    # Xóa các sheet cũ nếu đã tồn tại để tránh ghi trùng
    for sheet_name in ['6_FC_Lay_NhaTrang_Phuong', '7_FC_Giao_NhaTrang_Phuong']:
        if sheet_name in wb_write.sheetnames:
            print(f"  - Xóa sheet cũ '{sheet_name}'...")
            del wb_write[sheet_name]

    # 3.1 Xử lý ghi sheet 6_FC_Lay_NhaTrang_Phuong
    print("✏️ Tạo sheet '6_FC_Lay_NhaTrang_Phuong'...")
    ws_lay_phuong = wb_write.create_sheet('6_FC_Lay_NhaTrang_Phuong')
    
    sheet_lay = wb_data['6_FC_Lay_Daily']
    lay_rows = list(sheet_lay.iter_rows(values_only=True))
    lay_headers = lay_rows[0]
    lay_row2 = lay_rows[1]
    
    # Ghi 2 dòng header đầu tiên
    new_headers = ['Vùng', 'Tỉnh/Quận', 'Phường_ID', 'Phường_Tên', 'BC_ID', 'BC_Tên', 'Sàn'] + list(lay_headers[5:])
    new_row2 = [None, None, None, None, None, None, 'Source'] + list(lay_row2[5:])
    ws_lay_phuong.append(new_headers)
    ws_lay_phuong.append(new_row2)
    
    # Tìm các cột ngày và bưu cục Nha Trang
    date_cols_lay = []
    for col_idx, col_name in enumerate(lay_headers):
        d = parse_header_date(col_name)
        if d:
            date_cols_lay.append((col_idx, d))
            
    nt_bc_ids = set(w['BC_Lấy_ID'] for w in nt_wards if w['BC_Lấy_ID'])
    
    for row in lay_rows[2:]:
        bc_id = row[2]
        bc_name = row[3]
        san = row[4]
        
        if bc_id in nt_bc_ids and san and 'Bulky' in str(san):
            serving_wards = [w for w in nt_wards if w['BC_Lấy_ID'] == bc_id]
            for w in serving_wards:
                p_id = w['Phường_ID']
                p_name = w['Phường_Tên']
                
                # Tạo hàng mới
                new_row = ['NTB', 'Khánh Hòa', p_id, p_name, bc_id, bc_name, san]
                
                # Tính giá trị cho từng ngày
                for col_idx, d_date in date_cols_lay:
                    val = row[col_idx]
                    bc_vol = float(val) if val is not None else 0.0
                    
                    m_str = f"T{d_date.month:02d}/2026"
                    if m_str not in month_cols:
                        m_str = 'T07/2026' if d_date.month == 6 else 'T12/2026'
                        
                    ratio = ratio_lay_map[bc_id][p_id][m_str]
                    new_row.append(bc_vol * ratio)
                    
                ws_lay_phuong.append(new_row)

    # 3.2 Xử lý ghi sheet 7_FC_Giao_NhaTrang_Phuong
    print("✏️ Tạo sheet '7_FC_Giao_NhaTrang_Phuong'...")
    ws_giao_phuong = wb_write.create_sheet('7_FC_Giao_NhaTrang_Phuong')
    
    sheet_giao = wb_data['7_FC_Giao_Daily']
    giao_rows = list(sheet_giao.iter_rows(values_only=True))
    giao_headers = giao_rows[0]
    giao_row2 = giao_rows[1]
    
    new_headers_g = ['Vùng', 'Tỉnh/Quận', 'Phường_ID', 'Phường_Tên', 'BC_ID', 'BC_Tên', 'Sàn'] + list(giao_headers[5:])
    new_row2_g = [None, None, None, None, None, None, 'Source'] + list(giao_row2[5:])
    ws_giao_phuong.append(new_headers_g)
    ws_giao_phuong.append(new_row2_g)
    
    date_cols_giao = []
    for col_idx, col_name in enumerate(giao_headers):
        d = parse_header_date(col_name)
        if d:
            date_cols_giao.append((col_idx, d))
            
    nt_giao_bc_ids = set(w['BC_Giao_ID'] for w in nt_wards if w['BC_Giao_ID'])
    
    for row in giao_rows[2:]:
        bc_id = row[2]
        bc_name = row[3]
        san = row[4]
        
        if bc_id in nt_giao_bc_ids and san and 'Bulky' in str(san):
            serving_wards = [w for w in nt_wards if w['BC_Giao_ID'] == bc_id]
            for w in serving_wards:
                p_id = w['Phường_ID']
                p_name = w['Phường_Tên']
                
                new_row = ['NTB', 'Khánh Hòa', p_id, p_name, bc_id, bc_name, san]
                
                for col_idx, d_date in date_cols_giao:
                    val = row[col_idx]
                    bc_vol = float(val) if val is not None else 0.0
                    
                    m_str = f"T{d_date.month:02d}/2026"
                    if m_str not in month_cols:
                        m_str = 'T07/2026' if d_date.month == 6 else 'T12/2026'
                        
                    ratio = ratio_giao_map[bc_id][p_id][m_str]
                    new_row.append(bc_vol * ratio)
                    
                ws_giao_phuong.append(new_row)

    # 4. Lưu lại đè lên file config cũ
    print(f"💾 Đang lưu trực tiếp vào file config: {config_path}...")
    wb_write.save(config_path)
    print("🎉 Hoàn thành ghi đè dữ liệu thành công!")

if __name__ == "__main__":
    main()
