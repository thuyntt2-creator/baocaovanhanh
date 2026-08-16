import sys
import os
import shutil
import openpyxl

# Set stdout encoding to UTF-8
sys.stdout.reconfigure(encoding='utf-8')

config_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB (1).xlsx"
aop_path = r"C:\Users\lap4all\Downloads\AOP_Hang_NTB_T7-T12_2026 (1).xlsx"
backup_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB (1)_backup.xlsx"

def main():
    print("🚀 Bắt đầu tiến trình tách dự báo tháng (sheet 0.3 Bưu cục Detail) theo Phường...")

    if not os.path.exists(config_path):
        print(f"❌ Không tìm thấy file cấu hình tại: {config_path}")
        sys.exit(1)
    if not os.path.exists(aop_path):
        print(f"❌ Không tìm thấy file AOP tại: {aop_path}")
        sys.exit(1)

    # 1. Sao lưu file config cũ trước khi sửa
    print(f"📦 Tạo file backup tại: {backup_path}...")
    shutil.copyfile(config_path, backup_path)
    print("✅ Đã sao lưu file gốc thành công.")

    # 2. Đọc dữ liệu tính toán
    print("📖 Đọc dữ liệu từ file config và file AOP...")
    wb_cfg = openpyxl.load_workbook(config_path, data_only=True)
    wb_aop = openpyxl.load_workbook(aop_path, data_only=True)

    # 2.1 Đọc 1_Config_Chuan từ file config
    sheet_chuan = wb_cfg['1_Config_Chuan']
    chuan_rows = list(sheet_chuan.iter_rows(values_only=True))
    chuan_headers = chuan_rows[0]
    h_idx = {h: i for i, h in enumerate(chuan_headers)}

    nt_wards = []
    for row in chuan_rows[1:]:
        if row[h_idx['Quận_Tên']] == 'Thành phố Nha Trang':
            nt_wards.append({
                'Phường_ID': row[h_idx['Phường_ID']],
                'Phường_Tên': row[h_idx['Phường_Tên']],
                'BC_Giao_Tên': row[h_idx['BC_Giao_Tên']]
            })

    # 2.2 Đọc 4_KQ_B2B từ file config để làm trọng số ratio
    sheet_b2b = wb_cfg['4_KQ_B2B']
    b2b_rows = list(sheet_b2b.iter_rows(values_only=True))
    b2b_headers = b2b_rows[0]
    month_cols = ['T07/2026', 'T08/2026', 'T09/2026', 'T10/2026', 'T11/2026', 'T12/2026']
    # Mapping tháng trong B2B (vd T07/2026) sang AOP (vd T7)
    month_map = {
        'T7': 'T07/2026',
        'T8': 'T08/2026',
        'T9': 'T09/2026',
        'T10': 'T10/2026',
        'T11': 'T11/2026',
        'T12': 'T12/2026'
    }
    aop_months = ['T7', 'T8', 'T9', 'T10', 'T11', 'T12']

    ward_monthly_vols = {} # ward_monthly_vols[p_id][month] = vol
    for row in b2b_rows[1:]:
        if row[1] == 'Khánh Hòa' and row[2] == 1548.0:
            p_id = row[3]
            ward_monthly_vols.setdefault(p_id, {col: 0.0 for col in month_cols})
            for col in month_cols:
                col_idx = b2b_headers.index(col)
                val = row[col_idx]
                ward_monthly_vols[p_id][col] += float(val) if val is not None else 0.0

    # 2.3 Tính tổng monthly B2B của các phường trực thuộc mỗi BC Giao (Khớp theo BC_Giao_Tên)
    bc_wards_monthly_giao = {}
    for w in nt_wards:
        p_id = w['Phường_ID']
        g_name = w['BC_Giao_Tên']
        w_vols = ward_monthly_vols.get(p_id, {col: 0.0 for col in month_cols})
        
        if g_name:
            bc_wards_monthly_giao.setdefault(g_name, {col: 0.0 for col in month_cols})
            for col in month_cols:
                bc_wards_monthly_giao[g_name][col] += w_vols[col]

    # Tính ratio cho từng phường trong mỗi BC Giao
    ratio_giao_map = {}
    for w in nt_wards:
        p_id = w['Phường_ID']
        g_name = w['BC_Giao_Tên']
        w_vols = ward_monthly_vols.get(p_id, {col: 0.0 for col in month_cols})
        
        if g_name:
            ratio_giao_map.setdefault(g_name, {})
            ratio_giao_map[g_name].setdefault(p_id, {})
            same_bc_wards = [x['Phường_ID'] for x in nt_wards if x['BC_Giao_Tên'] == g_name]
            for col in month_cols:
                tot_bc = bc_wards_monthly_giao[g_name][col]
                if tot_bc > 0:
                    ratio_giao_map[g_name][p_id][col] = w_vols[col] / tot_bc
                else:
                    # Fallback chia đều
                    ratio_giao_map[g_name][p_id][col] = 1.0 / len(same_bc_wards)

    # 3. Mở workbook gốc dạng data_only=False để ghi sheet mới và giữ công thức
    print("📖 Mở file config để ghi sheet mới...")
    wb_write = openpyxl.load_workbook(config_path, data_only=False)

    target_sheet_name = '0.3_BuuCuc_Detail_Phuong'
    if target_sheet_name in wb_write.sheetnames:
        print(f"  - Xóa sheet cũ '{target_sheet_name}'...")
        del wb_write[target_sheet_name]

    ws_new = wb_write.create_sheet(target_sheet_name)
    
    # Ghi header
    new_headers = ['Bưu cục', 'Phường_ID', 'Phường_Tên', 'T7', 'T8', 'T9', 'T10', 'T11', 'T12']
    ws_new.append(new_headers)

    # Đọc sheet 0.3 Bưu cục Detail từ file AOP
    sheet_detail = wb_aop['0.3 Bưu cục Detail']
    detail_rows = list(sheet_detail.iter_rows(values_only=True))
    detail_headers = detail_rows[1] # Row 2 là header chính
    
    # Danh sách tên các bưu cục giao ở Nha Trang
    nt_bc_names = set(w['BC_Giao_Tên'] for w in nt_wards if w['BC_Giao_Tên'])

    count_rows = 0
    for row in detail_rows[2:]: # Dòng 3 trở đi
        bc_name = row[1]
        
        # Chỉ xử lý các bưu cục Nha Trang
        if bc_name in nt_bc_names:
            serving_wards = [w for w in nt_wards if w['BC_Giao_Tên'] == bc_name]
            
            for w in serving_wards:
                p_id = w['Phường_ID']
                p_name = w['Phường_Tên']
                
                # Tạo dòng mới
                new_row = [bc_name, p_id, p_name]
                
                # Tính giá trị các tháng
                for aop_m in aop_months:
                    col_idx = detail_headers.index(aop_m)
                    bc_val = row[col_idx]
                    bc_vol = float(bc_val) if bc_val is not None else 0.0
                    
                    b2b_month = month_map[aop_m]
                    ratio = ratio_giao_map[bc_name][p_id][b2b_month]
                    new_row.append(bc_vol * ratio)
                    
                ws_new.append(new_row)
                count_rows += 1

    # 4. Lưu lại đè lên file config gốc
    print(f"💾 Đang lưu trực tiếp vào file config: {config_path}...")
    wb_write.save(config_path)
    print(f"🎉 Hoàn thành! Đã tạo sheet '{target_sheet_name}' với {count_rows} dòng dữ liệu.")

if __name__ == "__main__":
    main()
