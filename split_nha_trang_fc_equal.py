import sys
import os
import shutil
import openpyxl
from datetime import datetime, date

sys.stdout.reconfigure(encoding='utf-8')

config_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB (1).xlsx"
backup_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB (1)_backup.xlsx"

def parse_header_date(col_name):
    if not isinstance(col_name, str):
        return None
    parts = col_name.strip().split()
    if len(parts) == 2:
        try:
            day_str, month_str = parts[1].split('/')
            day = int(day_str)
            month = int(month_str)
            return date(2026, month, day)
        except Exception:
            return None
    return None

def main():
    print("🚀 Bắt đầu tiến trình tách dự báo Nha Trang theo phương pháp CHIA ĐỀU (Equal Split)...")

    if not os.path.exists(config_path):
        print(f"❌ Không tìm thấy file cấu hình tại: {config_path}")
        sys.exit(1)

    # 1. Tạo backup file
    print(f"📦 Tạo file backup tại: {backup_path}...")
    shutil.copyfile(config_path, backup_path)
    print("✅ Đã sao lưu file gốc thành công.")

    # 2. Đọc dữ liệu từ file config
    print("📖 Đọc dữ liệu 1_Config_Chuan...")
    wb_data = openpyxl.load_workbook(config_path, data_only=True)
    sheet_chuan = wb_data['1_Config_Chuan']
    chuan_rows = list(sheet_chuan.iter_rows(values_only=True))
    chuan_headers = chuan_rows[0]
    h_idx = {h: i for i, h in enumerate(chuan_headers)}
    
    # Danh sách phường Nha Trang
    nt_wards = []
    for row in chuan_rows[1:]:
        if row[h_idx['Quận_Tên']] == 'Thành phố Nha Trang':
            nt_wards.append({
                'Phường_ID': row[h_idx['Phường_ID']],
                'Phường_Tên': row[h_idx['Phường_Tên']],
                'BC_Lấy_ID': row[h_idx['BC_Lấy_ID']],
                'BC_Lấy_Tên': row[h_idx['BC_Lấy_Tên']],
                'BC_Giao_ID': row[h_idx['BC_Giao_ID']],
                'BC_Giao_Tên': row[h_idx['BC_Giao_Tên']]
            })

    # Mở workbook ghi đè dạng giữ công thức
    print("📖 Mở file config để ghi các sheet mới...")
    wb_write = openpyxl.load_workbook(config_path, data_only=False)

    # Xóa các sheet cũ nếu đã tồn tại để cập nhật
    for s_name in ['3_KQ_BC_Detail_Phuong', '6_FC_Lay_NhaTrang_Phuong', '7_FC_Giao_NhaTrang_Phuong']:
        if s_name in wb_write.sheetnames:
            print(f"  - Xóa sheet cũ '{s_name}'...")
            del wb_write[s_name]

    # ==================== 1. TÁCH SHEET 3_KQ_BC_Detail (Hàng tháng) ====================
    print("✏️ Tạo sheet '3_KQ_BC_Detail_Phuong' (Chia đều)...")
    ws_new_monthly = wb_write.create_sheet('3_KQ_BC_Detail_Phuong')
    
    sheet_detail = wb_data['3_KQ_BC_Detail']
    detail_rows = list(sheet_detail.iter_rows(values_only=True))
    detail_headers = detail_rows[0]
    month_cols = ['T07/2026', 'T08/2026', 'T09/2026', 'T10/2026', 'T11/2026', 'T12/2026']
    
    # Header
    ws_new_monthly.append(['vung', 'routed_bc_id', 'routed_bc_name', 'bc_type', 'weight_group', 'Phường_ID', 'Phường_Tên'] + month_cols)
    
    nt_giao_bc_ids = set(w['BC_Giao_ID'] for w in nt_wards if w['BC_Giao_ID'])
    bulky_groups = {'10-15kg', '15-30kg', '≥30kg'}
    
    count_m = 0
    for row in detail_rows[1:]:
        vung = row[0]
        bc_id = row[1]
        bc_name = row[2]
        bc_type = row[3]
        weight_group = row[4]
        
        if bc_id in nt_giao_bc_ids and weight_group in bulky_groups:
            # Tìm danh sách phường Nha Trang được bưu cục giao này phục vụ
            serving_wards = [w for w in nt_wards if w['BC_Giao_ID'] == bc_id]
            num_wards = len(serving_wards)
            
            for w in serving_wards:
                p_id = w['Phường_ID']
                p_name = w['Phường_Tên']
                new_row = [vung, bc_id, bc_name, bc_type, weight_group, p_id, p_name]
                
                # Chia đều
                for col in month_cols:
                    col_idx = detail_headers.index(col)
                    val = row[col_idx]
                    bc_vol = float(val) if val is not None else 0.0
                    new_row.append(bc_vol / num_wards if num_wards > 0 else 0.0)
                    
                ws_new_monthly.append(new_row)
                count_m += 1

    # ==================== 2. TÁCH SHEET 6_FC_Lay_Daily (Hàng ngày) ====================
    print("✏️ Tạo sheet '6_FC_Lay_NhaTrang_Phuong' (Chia đều)...")
    ws_new_lay = wb_write.create_sheet('6_FC_Lay_NhaTrang_Phuong')
    
    sheet_lay = wb_data['6_FC_Lay_Daily']
    lay_rows = list(sheet_lay.iter_rows(values_only=True))
    lay_headers = lay_rows[0]
    lay_row2 = lay_rows[1]
    
    ws_new_lay.append(['Vùng', 'Tỉnh/Quận', 'Phường_ID', 'Phường_Tên', 'BC_ID', 'BC_Tên', 'Sàn'] + list(lay_headers[5:]))
    ws_new_lay.append([None, None, None, None, None, None, 'Source'] + list(lay_row2[5:]))
    
    nt_lay_bc_ids = set(w['BC_Lấy_ID'] for w in nt_wards if w['BC_Lấy_ID'])
    
    count_lay = 0
    for row in lay_rows[2:]:
        bc_id = row[2]
        bc_name = row[3]
        san = row[4]
        
        if bc_id in nt_lay_bc_ids and san and 'Bulky' in str(san):
            serving_wards = [w for w in nt_wards if w['BC_Lấy_ID'] == bc_id]
            num_wards = len(serving_wards)
            
            for w in serving_wards:
                p_id = w['Phường_ID']
                p_name = w['Phường_Tên']
                new_row = ['NTB', 'Khánh Hòa', p_id, p_name, bc_id, bc_name, san]
                
                for val in row[5:]:
                    bc_vol = float(val) if val is not None else 0.0
                    new_row.append(bc_vol / num_wards if num_wards > 0 else 0.0)
                    
                ws_new_lay.append(new_row)
                count_lay += 1

    # ==================== 3. TÁCH SHEET 7_FC_Giao_Daily (Hàng ngày) ====================
    print("✏️ Tạo sheet '7_FC_Giao_NhaTrang_Phuong' (Chia đều)...")
    ws_new_giao = wb_write.create_sheet('7_FC_Giao_NhaTrang_Phuong')
    
    sheet_giao = wb_data['7_FC_Giao_Daily']
    giao_rows = list(sheet_giao.iter_rows(values_only=True))
    giao_headers = giao_rows[0]
    giao_row2 = giao_rows[1]
    
    ws_new_giao.append(['Vùng', 'Tỉnh/Quận', 'Phường_ID', 'Phường_Tên', 'BC_ID', 'BC_Tên', 'Sàn'] + list(giao_headers[5:]))
    ws_new_giao.append([None, None, None, None, None, None, 'Source'] + list(giao_row2[5:]))
    
    count_giao = 0
    for row in giao_rows[2:]:
        bc_id = row[2]
        bc_name = row[3]
        san = row[4]
        
        if bc_id in nt_giao_bc_ids and san and 'Bulky' in str(san):
            serving_wards = [w for w in nt_wards if w['BC_Giao_ID'] == bc_id]
            num_wards = len(serving_wards)
            
            for w in serving_wards:
                p_id = w['Phường_ID']
                p_name = w['Phường_Tên']
                new_row = ['NTB', 'Khánh Hòa', p_id, p_name, bc_id, bc_name, san]
                
                for val in row[5:]:
                    bc_vol = float(val) if val is not None else 0.0
                    new_row.append(bc_vol / num_wards if num_wards > 0 else 0.0)
                    
                ws_new_giao.append(new_row)
                count_giao += 1

    # 4. Lưu lại
    print(f"💾 Đang lưu trực tiếp vào file config: {config_path}...")
    wb_write.save(config_path)
    print(f"🎉 Hoàn thành chia đều! Đã tạo:")
    print(f"  - Sheet '3_KQ_BC_Detail_Phuong': {count_m} dòng")
    print(f"  - Sheet '6_FC_Lay_NhaTrang_Phuong': {count_lay} dòng")
    print(f"  - Sheet '7_FC_Giao_NhaTrang_Phuong': {count_giao} dòng")

if __name__ == "__main__":
    main()
