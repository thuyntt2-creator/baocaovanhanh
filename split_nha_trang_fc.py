import sys
import os
import openpyxl
import pandas as pd
from datetime import datetime, date

# Set output encoding to UTF-8
sys.stdout.reconfigure(encoding='utf-8')

# Input paths
config_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB (1).xlsx"
aop_path = r"C:\Users\lap4all\Downloads\AOP_MAU_NTB_Cach1_Dung.xlsx"
output_path = r"C:\Users\lap4all\Downloads\FC_HangCongKenh_TheoPhuong_NhaTrang.xlsx"

def parse_header_date(col_name):
    """
    Parses column headers like 'T2 22/06' to a datetime.date object.
    Returns None if the header is not a date.
    """
    if not isinstance(col_name, str):
        return None
    parts = col_name.strip().split()
    if len(parts) == 2:
        try:
            day_str, month_str = parts[1].split('/')
            day = int(day_str)
            month = int(month_str)
            # Default to 2026 as per workspace context
            return date(2026, month, day)
        except Exception:
            return None
    return None

def main():
    print("🚀 Bắt đầu tiến trình tách dự báo (FC) hàng cồng kềnh Nha Trang theo Phường...")

    if not os.path.exists(config_path):
        print(f"❌ Không tìm thấy file cấu hình tại: {config_path}")
        sys.exit(1)
    if not os.path.exists(aop_path):
        print(f"❌ Không tìm thấy file AOP mẫu tại: {aop_path}")
        sys.exit(1)

    # 1. Đọc dữ liệu từ config_psbba_NTB (1).xlsx
    print("📖 Đọc dữ liệu từ file cấu hình...")
    wb_cfg = openpyxl.load_workbook(config_path, data_only=True)
    
    # 1.1 Đọc 1_Config_Chuan để lấy thông tin các phường ở Nha Trang và bưu cục phục vụ
    print("  - Đọc sheet '1_Config_Chuan'...")
    sheet_chuan = wb_cfg['1_Config_Chuan']
    chuan_rows = list(sheet_chuan.iter_rows(values_only=True))
    chuan_headers = chuan_rows[0]
    
    # Map headers to index
    h_idx = {h: i for i, h in enumerate(chuan_headers)}
    
    # Lọc các dòng thuộc Nha Trang (Quận_Tên = 'Thành phố Nha Trang')
    nt_wards = []
    # Lưu danh sách ward_id duy nhất để đối chiếu
    nt_ward_ids = set()
    for row in chuan_rows[1:]:
        if row[h_idx['Quận_Tên']] == 'Thành phố Nha Trang':
            p_id = row[h_idx['Phường_ID']]
            p_name = row[h_idx['Phường_Tên']]
            bc_lay_id = row[h_idx['BC_Lấy_ID']]
            bc_lay_name = row[h_idx['BC_Lấy_Tên']]
            bc_giao_id = row[h_idx['BC_Giao_ID']]
            bc_giao_name = row[h_idx['BC_Giao_Tên']]
            
            nt_wards.append({
                'Phường_ID': p_id,
                'Phường_Tên': p_name,
                'BC_Lấy_ID': bc_lay_id,
                'BC_Lấy_Tên': bc_lay_name,
                'BC_Giao_ID': bc_giao_id,
                'BC_Giao_Tên': bc_giao_name
            })
            if p_id:
                nt_ward_ids.add(p_id)
                
    df_wards = pd.DataFrame(nt_wards)
    print(f"    Tìm thấy {len(df_wards)} phường/xã thuộc Nha Trang.")

    # 1.2 Đọc 4_KQ_B2B để lấy sản lượng cồng kềnh dự kiến theo tháng của từng phường
    print("  - Đọc sheet '4_KQ_B2B'...")
    sheet_b2b = wb_cfg['4_KQ_B2B']
    b2b_rows = list(sheet_b2b.iter_rows(values_only=True))
    b2b_headers = b2b_rows[0]
    
    # Lọc các phường Nha Trang và lấy sản lượng theo các tháng
    b2b_data = []
    month_cols = ['T07/2026', 'T08/2026', 'T09/2026', 'T10/2026', 'T11/2026', 'T12/2026']
    
    for row in b2b_rows[1:]:
        # row[2] là Quận_ID (1548.0 = Nha Trang) hoặc row[1] là Khánh Hòa
        if row[1] == 'Khánh Hòa' and row[2] == 1548.0:
            p_id = row[3]
            band = row[10] # Weight_Band
            
            # Đọc giá trị các tháng
            m_vols = {}
            for col in month_cols:
                col_idx = b2b_headers.index(col)
                val = row[col_idx]
                m_vols[col] = float(val) if val is not None else 0.0
                
            b2b_data.append({
                'Phường_ID': p_id,
                'Weight_Band': band,
                **m_vols
            })
            
    df_b2b = pd.DataFrame(b2b_data)
    print(f"    Tìm thấy {len(df_b2b)} dòng sản lượng B2B của Nha Trang.")

    # Gom nhóm theo Phường_ID để tính tổng sản lượng bulky hàng tháng của từng phường
    # (Vì mỗi phường có thể có 2 dòng tương ứng với 2 band cân nặng khác nhau)
    df_ward_monthly = df_b2b.groupby('Phường_ID')[month_cols].sum().reset_index()
    
    # Map Phường_ID sang Phường_Tên
    ward_name_map = df_wards.set_index('Phường_ID')['Phường_Tên'].to_dict()
    df_ward_monthly['Phường_Tên'] = df_ward_monthly['Phường_ID'].map(ward_name_map)
    
    print("📊 Bảng sản lượng cồng kềnh (B2B) hàng tháng theo Phường:")
    print(df_ward_monthly[['Phường_ID', 'Phường_Tên'] + month_cols].head(10))

    # Tính tỷ lệ phân bổ hàng tháng của các phường Nha Trang
    # Để Method A hoạt động: lấy tỷ lệ của từng phường trên tổng Nha Trang
    ratios_city = {}
    for col in month_cols:
        total_city = df_ward_monthly[col].sum()
        if total_city > 0:
            df_ward_monthly[f'Ratio_City_{col}'] = df_ward_monthly[col] / total_city
        else:
            df_ward_monthly[f'Ratio_City_{col}'] = 0.0
            
    # Tính tỷ lệ phân bổ hàng tháng cho bưu cục phục vụ (Method B)
    # Lấy (Pick-up) và Giao (Delivery) riêng biệt
    # Cho mỗi bưu cục, tính tổng sản lượng của các phường Nha Trang mà nó phục vụ.
    # Sau đó tính tỷ lệ đóng góp của mỗi phường trong bưu cục đó.
    
    # 2. Đọc AOP mẫu để lấy dự báo hàng ngày của BCCK Nha Trang (Method A)
    print("📖 Đọc dữ liệu từ file AOP mẫu...")
    wb_aop = openpyxl.load_workbook(aop_path, data_only=True)
    
    daily_forecasts = []
    # Quét các sheet Forecast T7 -> Forecast T12
    for m_idx, col_month in enumerate(month_cols, start=7):
        sheet_name = f"Forecast T{m_idx}"
        if sheet_name in wb_aop.sheetnames:
            print(f"  - Đọc sheet '{sheet_name}'...")
            sheet_fc = wb_aop[sheet_name]
            fc_rows = list(sheet_fc.iter_rows(values_only=True))
            
            # Tìm dòng 'BCCK Nha Trang'
            nt_row = None
            for row in fc_rows:
                if row[0] == 'BCCK Nha Trang':
                    nt_row = row
                    break
            
            if nt_row is None:
                print(f"    ⚠️ Cảnh báo: Không tìm thấy dòng 'BCCK Nha Trang' trong sheet '{sheet_name}'.")
                continue
                
            # Đọc hàng Ngày (dòng 2) và Thứ (dòng 3) để lấy thông tin các ngày
            dates_row = fc_rows[1] # Dòng 2
            days_row = fc_rows[2]  # Dòng 3
            
            # Bắt đầu đọc các cột từ cột 1 (index 1) trở đi
            for col_idx in range(1, len(nt_row)):
                d_val = dates_row[col_idx]
                day_name = days_row[col_idx]
                fc_val = nt_row[col_idx]
                
                # Kiểm tra xem có phải là ngày hợp lệ không
                if isinstance(d_val, datetime):
                    d_date = d_val.date()
                elif isinstance(d_val, date):
                    d_date = d_val
                else:
                    # Hết các cột ngày
                    break
                    
                if fc_val is None:
                    fc_val = 0.0
                    
                daily_forecasts.append({
                    'Month_Col': col_month,
                    'Date': d_date,
                    'DayOfWeek': day_name,
                    'BCCK_NhaTrang_Total': float(fc_val)
                })
                
    df_daily_aop = pd.DataFrame(daily_forecasts)
    print(f"    Tổng cộng đọc được {len(df_daily_aop)} ngày dự báo của BCCK Nha Trang.")

    # 3. Tiến hành tách Method A (City-Level)
    print("🔮 Đang thực hiện tách sản lượng theo Method A (City-Level)...")
    method_a_records = []
    
    # Tạo map tỷ lệ city để tra cứu nhanh: ratios_city_map[Phường_ID][Month_Col] = Ratio
    ratios_city_map = {}
    for idx, row in df_ward_monthly.iterrows():
        p_id = row['Phường_ID']
        p_name = row['Phường_Tên']
        ratios_city_map[p_id] = {}
        for col in month_cols:
            ratios_city_map[p_id][col] = row[f'Ratio_City_{col}']

    # Tách cho tất cả 28 phường (đối với các phường không có sản lượng trong B2B, tỷ lệ sẽ là 0)
    for idx, d_row in df_daily_aop.iterrows():
        m_col = d_row['Month_Col']
        d_date = d_row['Date']
        day_name = d_row['DayOfWeek']
        total_vol = d_row['BCCK_NhaTrang_Total']
        
        for w_row in nt_wards:
            p_id = w_row['Phường_ID']
            p_name = w_row['Phường_Tên']
            
            # Lấy tỷ lệ
            ratio = ratios_city_map.get(p_id, {}).get(m_col, 0.0)
            ward_vol = total_vol * ratio
            
            method_a_records.append({
                'Date': d_date,
                'DayOfWeek': day_name,
                'Month': m_col,
                'Phường_ID': p_id,
                'Phường_Tên': p_name,
                'BCCK_NhaTrang_Total': total_vol,
                'Ward_Ratio': ratio,
                'Ward_Forecast': ward_vol
            })
            
    df_method_a = pd.DataFrame(method_a_records)
    print(f"    Method A hoàn thành. Tổng số dòng dữ liệu: {len(df_method_a)}.")

    # 4. Tiến hành tách Method B (Post-Office Level)
    print("🔮 Đang thực hiện tách sản lượng theo Method B (Post-Office Level)...")
    
    # 4.1 Đọc dữ liệu 6_FC_Lay_Daily và 7_FC_Giao_Daily từ file cấu hình
    method_b_records = []
    
    # Chuẩn bị mapping tỷ lệ cho từng BC và các phường trực thuộc
    # Tính tổng monthly B2B của các phường trực thuộc mỗi BC Lấy và BC Giao
    # bc_wards_monthly_lay[bc_id][month] = sum of b2b volumes of served wards
    bc_wards_monthly_lay = {}
    bc_wards_monthly_giao = {}
    
    # Lấy thông tin b2b cho từng phường
    ward_monthly_map = {} # ward_monthly_map[p_id][month] = vol
    for idx, row in df_ward_monthly.iterrows():
        p_id = row['Phường_ID']
        ward_monthly_map[p_id] = {col: row[col] for col in month_cols}
        
    for w in nt_wards:
        p_id = w['Phường_ID']
        lay_id = w['BC_Lấy_ID']
        giao_id = w['BC_Giao_ID']
        
        # BC Lấy
        if lay_id:
            bc_wards_monthly_lay.setdefault(lay_id, {})
            w_vols = ward_monthly_map.get(p_id, {col: 0.0 for col in month_cols})
            for col in month_cols:
                bc_wards_monthly_lay[lay_id].setdefault(col, 0.0)
                bc_wards_monthly_lay[lay_id][col] += w_vols[col]
                
        # BC Giao
        if giao_id:
            bc_wards_monthly_giao.setdefault(giao_id, {})
            w_vols = ward_monthly_map.get(p_id, {col: 0.0 for col in month_cols})
            for col in month_cols:
                bc_wards_monthly_giao[giao_id].setdefault(col, 0.0)
                bc_wards_monthly_giao[giao_id][col] += w_vols[col]

    # Tính ratio cho từng phường trong mỗi BC
    # ratio_lay_map[bc_id][p_id][month] = ratio
    ratio_lay_map = {}
    ratio_giao_map = {}
    
    for w in nt_wards:
        p_id = w['Phường_ID']
        lay_id = w['BC_Lấy_ID']
        giao_id = w['BC_Giao_ID']
        w_vols = ward_monthly_map.get(p_id, {col: 0.0 for col in month_cols})
        
        # BC Lấy
        if lay_id:
            ratio_lay_map.setdefault(lay_id, {})
            ratio_lay_map[lay_id].setdefault(p_id, {})
            # Tìm tổng số lượng các phường thuộc BC Lấy này
            # Tìm danh sách các phường cùng thuộc BC Lấy này
            same_bc_wards = [x['Phường_ID'] for x in nt_wards if x['BC_Lấy_ID'] == lay_id]
            for col in month_cols:
                tot_bc = bc_wards_monthly_lay[lay_id][col]
                if tot_bc > 0:
                    ratio_lay_map[lay_id][p_id][col] = w_vols[col] / tot_bc
                else:
                    # Fallback chia đều cho các phường cùng bưu cục
                    ratio_lay_map[lay_id][p_id][col] = 1.0 / len(same_bc_wards)
                    
        # BC Giao
        if giao_id:
            ratio_giao_map.setdefault(giao_id, {})
            ratio_giao_map[giao_id].setdefault(p_id, {})
            same_bc_wards = [x['Phường_ID'] for x in nt_wards if x['BC_Giao_ID'] == giao_id]
            for col in month_cols:
                tot_bc = bc_wards_monthly_giao[giao_id][col]
                if tot_bc > 0:
                    ratio_giao_map[giao_id][p_id][col] = w_vols[col] / tot_bc
                else:
                    ratio_giao_map[giao_id][p_id][col] = 1.0 / len(same_bc_wards)

    # 4.2 Xử lý sheet 6_FC_Lay_Daily
    print("  - Đọc và tách dự báo sheet '6_FC_Lay_Daily'...")
    sheet_lay = wb_cfg['6_FC_Lay_Daily']
    lay_rows = list(sheet_lay.iter_rows(values_only=True))
    lay_headers = lay_rows[0]
    
    # Tìm các cột ngày
    date_cols_lay = []
    for col_idx, col_name in enumerate(lay_headers):
        d = parse_header_date(col_name)
        if d:
            date_cols_lay.append((col_idx, col_name, d))
            
    print(f"    Tìm thấy {len(date_cols_lay)} cột ngày trong sheet Lấy.")
    
    # Lọc các dòng bulky của bưu cục Nha Trang
    # Nha Trang post office IDs:
    nt_bc_ids = set(w['BC_Lấy_ID'] for w in nt_wards if w['BC_Lấy_ID'])
    
    for row in lay_rows[1:]:
        bc_id = row[2]
        bc_name = row[3]
        san = row[4] # Sàn (loại hàng)
        
        # Chỉ xử lý Nha Trang và các dòng bulky (chứa từ 'Bulky')
        if bc_id in nt_bc_ids and san and 'Bulky' in str(san):
            # Tìm các phường thuộc bưu cục này
            serving_wards = [w for w in nt_wards if w['BC_Lấy_ID'] == bc_id]
            
            for col_idx, col_name, d_date in date_cols_lay:
                val = row[col_idx]
                bc_vol = float(val) if val is not None else 0.0
                
                # Xác định tháng để lấy tỷ lệ (vd: 22/06 -> T07/2026 vì chưa có T06, hoặc map theo tháng)
                m_str = f"T{d_date.month:02d}/2026"
                if m_str not in month_cols:
                    # Nếu ngày tháng 6/2026 thì lấy tỷ lệ tháng 7/2026 để làm xấp xỉ
                    if d_date.month == 6:
                        m_str = 'T07/2026'
                    else:
                        m_str = 'T12/2026'
                        
                for w in serving_wards:
                    p_id = w['Phường_ID']
                    p_name = w['Phường_Tên']
                    
                    ratio = ratio_lay_map[bc_id][p_id][m_str]
                    ward_vol = bc_vol * ratio
                    
                    method_b_records.append({
                        'Type': 'Lấy',
                        'BC_ID': bc_id,
                        'BC_Tên': bc_name,
                        'Sàn_Bulky': san,
                        'Phường_ID': p_id,
                        'Phường_Tên': p_name,
                        'Date': d_date,
                        'DayOfWeek': col_name.split()[0],
                        'BC_Total_Volume': bc_vol,
                        'Ward_Ratio': ratio,
                        'Ward_Volume': ward_vol
                    })

    # 4.3 Xử lý sheet 7_FC_Giao_Daily
    print("  - Đọc và tách dự báo sheet '7_FC_Giao_Daily'...")
    sheet_giao = wb_cfg['7_FC_Giao_Daily']
    giao_rows = list(sheet_giao.iter_rows(values_only=True))
    giao_headers = giao_rows[0]
    
    date_cols_giao = []
    for col_idx, col_name in enumerate(giao_headers):
        d = parse_header_date(col_name)
        if d:
            date_cols_giao.append((col_idx, col_name, d))
            
    print(f"    Tìm thấy {len(date_cols_giao)} cột ngày trong sheet Giao.")
    
    nt_giao_bc_ids = set(w['BC_Giao_ID'] for w in nt_wards if w['BC_Giao_ID'])
    
    for row in giao_rows[1:]:
        bc_id = row[2]
        bc_name = row[3]
        san = row[4]
        
        if bc_id in nt_giao_bc_ids and san and 'Bulky' in str(san):
            serving_wards = [w for w in nt_wards if w['BC_Giao_ID'] == bc_id]
            
            for col_idx, col_name, d_date in date_cols_giao:
                val = row[col_idx]
                bc_vol = float(val) if val is not None else 0.0
                
                m_str = f"T{d_date.month:02d}/2026"
                if m_str not in month_cols:
                    if d_date.month == 6:
                        m_str = 'T07/2026'
                    else:
                        m_str = 'T12/2026'
                        
                for w in serving_wards:
                    p_id = w['Phường_ID']
                    p_name = w['Phường_Tên']
                    
                    ratio = ratio_giao_map[bc_id][p_id][m_str]
                    ward_vol = bc_vol * ratio
                    
                    method_b_records.append({
                        'Type': 'Giao',
                        'BC_ID': bc_id,
                        'BC_Tên': bc_name,
                        'Sàn_Bulky': san,
                        'Phường_ID': p_id,
                        'Phường_Tên': p_name,
                        'Date': d_date,
                        'DayOfWeek': col_name.split()[0],
                        'BC_Total_Volume': bc_vol,
                        'Ward_Ratio': ratio,
                        'Ward_Volume': ward_vol
                    })

    df_method_b = pd.DataFrame(method_b_records)
    print(f"    Method B hoàn thành. Tổng số dòng dữ liệu: {len(df_method_b)}.")

    # 5. Xuất ra file Excel kết quả
    print(f"💾 Đang xuất kết quả ra file: {output_path} ...")
    
    # Để thuận tiện, chúng ta sẽ tạo các sheet:
    # 1. README: giải thích
    # 2. Method_A_Tidy: Dạng hàng dọc Method A
    # 3. Method_A_Wide: Dạng hàng ngang Method A (dễ nhìn)
    # 4. Method_B_Tidy: Dạng hàng dọc Method B
    
    # Tạo bản Wide cho Method A
    print("  - Xây dựng bảng hiển thị ngang (Wide format) cho Method A...")
    df_method_a_wide = df_method_a.pivot_table(
        index=['Phường_ID', 'Phường_Tên'],
        columns='Date',
        values='Ward_Forecast',
        aggfunc='sum'
    ).reset_index()
    
    # Tạo bản Wide cho Method B (Tách riêng Lấy và Giao)
    print("  - Xây dựng bảng hiển thị ngang (Wide format) cho Method B...")
    df_method_b_lay_wide = df_method_b[df_method_b['Type'] == 'Lấy'].pivot_table(
        index=['BC_ID', 'BC_Tên', 'Sàn_Bulky', 'Phường_ID', 'Phường_Tên'],
        columns='Date',
        values='Ward_Volume',
        aggfunc='sum'
    ).reset_index()
    
    df_method_b_giao_wide = df_method_b[df_method_b['Type'] == 'Giao'].pivot_table(
        index=['BC_ID', 'BC_Tên', 'Sàn_Bulky', 'Phường_ID', 'Phường_Tên'],
        columns='Date',
        values='Ward_Volume',
        aggfunc='sum'
    ).reset_index()

    # Tạo bảng giải thích README
    readme_data = [
        ["Tên Sheet", "Mô tả", "Cách dùng"],
        ["Method_A_Wide", "Dự báo hàng ngày chiều Giao (Bulky) tại Nha Trang theo từng Phường (dạng ngang). Tách từ BCCK Nha Trang.", "Dành cho sếp và team theo dõi nhanh theo từng ngày."],
        ["Method_A_Tidy", "Dữ liệu dạng hàng dọc (tidy data) của Method A.", "Phù hợp để làm Pivot Table hoặc nạp vào PowerBI/Looker Studio."],
        ["Method_B_Lay_Wide", "Dự báo LẤY hàng ngày chiều Lấy (Bulky) Nha Trang theo từng Phường (dạng ngang). Tách từ bưu cục thường.", "Dành cho điều phối chặng lấy chặng chót."],
        ["Method_B_Giao_Wide", "Dự báo GIAO hàng ngày chiều Giao (Bulky) Nha Trang theo từng Phường (dạng ngang). Tách từ bưu cục thường.", "Dành cho điều phối chặng giao chặng chót."],
        ["Method_B_Tidy", "Dữ liệu dạng hàng dọc (tidy data) của Method B (gồm cả Lấy và Giao).", "Phù hợp để làm phân tích chuyên sâu."],
        ["Config_NhaTrang_Wards", "Cơ cấu cấu hình bưu cục lấy/giao và tỷ lệ B2B của các phường Nha Trang.", "Tham chiếu cấu hình."]
    ]
    df_readme = pd.DataFrame(readme_data[1:], columns=readme_data[0])

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_readme.to_excel(writer, sheet_name='README', index=False)
        df_method_a_wide.to_excel(writer, sheet_name='Method_A_Wide', index=False)
        df_method_a.to_excel(writer, sheet_name='Method_A_Tidy', index=False)
        df_method_b_lay_wide.to_excel(writer, sheet_name='Method_B_Lay_Wide', index=False)
        df_method_b_giao_wide.to_excel(writer, sheet_name='Method_B_Giao_Wide', index=False)
        df_method_b.to_excel(writer, sheet_name='Method_B_Tidy', index=False)
        df_wards.to_excel(writer, sheet_name='Config_NhaTrang_Wards', index=False)

    print(f"🎉 Xuất báo cáo thành công tại: {output_path}")
    print("📢 Đã phân tách xong dự báo hàng cồng kềnh Nha Trang theo Phường!")

if __name__ == "__main__":
    main()
