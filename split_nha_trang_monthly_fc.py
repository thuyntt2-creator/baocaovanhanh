import sys
import os
import shutil
import openpyxl

# Set stdout encoding to UTF-8
sys.stdout.reconfigure(encoding='utf-8')

config_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB (1).xlsx"
backup_path = r"C:\Users\lap4all\Downloads\config_psbba_NTB (1)_backup.xlsx"

def main():
    print("🚀 Bắt đầu tiến trình tách dự báo hàng tháng (sheet 3_KQ_BC_Detail) theo Phường...")

    if not os.path.exists(config_path):
        print(f"❌ Không tìm thấy file cấu hình tại: {config_path}")
        sys.exit(1)

    # 1. Tạo file backup trước khi ghi đè
    print(f"📦 Tạo file backup tại: {backup_path}...")
    shutil.copyfile(config_path, backup_path)
    print("✅ Đã sao lưu file gốc thành công.")

    # 2. Đọc dữ liệu (dùng data_only=True để đọc đúng giá trị số)
    print("📖 Đọc dữ liệu tính toán từ file config...")
    wb_data = openpyxl.load_workbook(config_path, data_only=True)
    
    # 2.1 Đọc 1_Config_Chuan
    sheet_chuan = wb_data['1_Config_Chuan']
    chuan_rows = list(sheet_chuan.iter_rows(values_only=True))
    chuan_headers = chuan_rows[0]
    h_idx = {h: i for i, h in enumerate(chuan_headers)}
    
    nt_wards = []
    for row in chuan_rows[1:]:
        if row[h_idx['Quận_Tên']] == 'Thành phố Nha Trang':
            nt_wards.append({
                'Phường_ID': row[h_idx['Phường_ID']],
                'Phường_Tên': row[h_idx['Phường_Tên']],
                'BC_Giao_ID': row[h_idx['BC_Giao_ID']],
                'BC_Giao_Tên': row[h_idx['BC_Giao_Tên']]
            })
            
    # 2.2 Đọc 4_KQ_B2B
    sheet_b2b = wb_data['4_KQ_B2B']
    b2b_rows = list(sheet_b2b.iter_rows(values_only=True))
    b2b_headers = b2b_rows[0]
    month_cols = ['T07/2026', 'T08/2026', 'T09/2026', 'T10/2026', 'T11/2026', 'T12/2026']
    
    ward_monthly_vols = {} # ward_monthly_vols[p_id][month] = vol
    for row in b2b_rows[1:]:
        if row[1] == 'Khánh Hòa' and row[2] == 1548.0:
            p_id = row[3]
            ward_monthly_vols.setdefault(p_id, {col: 0.0 for col in month_cols})
            for col in month_cols:
                col_idx = b2b_headers.index(col)
                val = row[col_idx]
                ward_monthly_vols[p_id][col] += float(val) if val is not None else 0.0

    # 2.3 Tính tổng monthly B2B của các phường trực thuộc mỗi BC Giao
    bc_wards_monthly_giao = {}
    for w in nt_wards:
        p_id = w['Phường_ID']
        giao_id = w['BC_Giao_ID']
        w_vols = ward_monthly_vols.get(p_id, {col: 0.0 for col in month_cols})
        
        if giao_id:
            bc_wards_monthly_giao.setdefault(giao_id, {col: 0.0 for col in month_cols})
            for col in month_cols:
                bc_wards_monthly_giao[giao_id][col] += w_vols[col]

    # Tính ratio cho từng phường trong mỗi BC Giao
    ratio_giao_map = {}
    for w in nt_wards:
        p_id = w['Phường_ID']
        giao_id = w['BC_Giao_ID']
        w_vols = ward_monthly_vols.get(p_id, {col: 0.0 for col in month_cols})
        
        if giao_id:
            ratio_giao_map.setdefault(giao_id, {})
            ratio_giao_map[giao_id].setdefault(p_id, {})
            same_bc_wards = [x['Phường_ID'] for x in nt_wards if x['BC_Giao_ID'] == giao_id]
            for col in month_cols:
                tot_bc = bc_wards_monthly_giao[giao_id][col]
                if tot_bc > 0:
                    ratio_giao_map[giao_id][p_id][col] = w_vols[col] / tot_bc
                else:
                    # Fallback chia đều nếu không có sản lượng B2B
                    ratio_giao_map[giao_id][p_id][col] = 1.0 / len(same_bc_wards)

    # 3. Mở workbook gốc dạng data_only=False để ghi sheet mới và giữ công thức
    print("📖 Mở file config để ghi sheet mới...")
    wb_write = openpyxl.load_workbook(config_path, data_only=False)

    # Xóa sheet cũ nếu đã tồn tại
    target_sheet_name = '3_KQ_BC_Detail_Phuong'
    if target_sheet_name in wb_write.sheetnames:
        print(f"  - Xóa sheet cũ '{target_sheet_name}'...")
        del wb_write[target_sheet_name]

    ws_new = wb_write.create_sheet(target_sheet_name)
    
    # Đọc sheet 3_KQ_BC_Detail để lấy dữ liệu phân bổ
    sheet_detail = wb_data['3_KQ_BC_Detail']
    detail_rows = list(sheet_detail.iter_rows(values_only=True))
    detail_headers = detail_rows[0]
    
    # Tạo header cho sheet mới
    new_headers = ['vung', 'routed_bc_id', 'routed_bc_name', 'bc_type', 'weight_group', 'Phường_ID', 'Phường_Tên'] + month_cols
    ws_new.append(new_headers)

    # Các bưu cục Nha Trang cần tách
    nt_giao_bc_ids = set(w['BC_Giao_ID'] for w in nt_wards if w['BC_Giao_ID'])
    # Nhóm cân nặng bulky
    bulky_groups = {'10-15kg', '15-30kg', '≥30kg'}

    count_rows = 0
    for row in detail_rows[1:]:
        vung = row[0]
        bc_id = row[1]
        bc_name = row[2]
        bc_type = row[3]
        weight_group = row[4]
        
        # Chỉ xử lý Nha Trang và các nhóm cồng kềnh (>= 10kg)
        if bc_id in nt_giao_bc_ids and weight_group in bulky_groups:
            serving_wards = [w for w in nt_wards if w['BC_Giao_ID'] == bc_id]
            for w in serving_wards:
                p_id = w['Phường_ID']
                p_name = w['Phường_Tên']
                
                # Tạo dòng mới
                new_row = [vung, bc_id, bc_name, bc_type, weight_group, p_id, p_name]
                
                # Tách sản lượng của từng tháng
                for col in month_cols:
                    col_idx = detail_headers.index(col)
                    val = row[col_idx]
                    bc_monthly_vol = float(val) if val is not None else 0.0
                    
                    ratio = ratio_giao_map[bc_id][p_id][col]
                    new_row.append(bc_monthly_vol * ratio)
                    
                ws_new.append(new_row)
                count_rows += 1

    # 4. Lưu lại đè lên file config cũ
    print(f"💾 Đang lưu trực tiếp vào file config: {config_path}...")
    wb_write.save(config_path)
    print(f"🎉 Hoàn thành! Đã tạo sheet '{target_sheet_name}' với {count_rows} dòng dữ liệu.")

if __name__ == "__main__":
    main()
