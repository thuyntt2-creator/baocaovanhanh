import os
import sys
import io
import openpyxl

# Fix encoding for Windows
os.environ['PYTHONIOENCODING'] = 'utf-8'
try:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
except AttributeError:
    pass

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def main():
    excel_files = ['manual_raw.xlsx', 'temp_ghn.xlsx']
    for f in excel_files:
        path = os.path.join(BASE_DIR, f)
        if os.path.exists(path):
            print(f"\nExcel File: {f}")
            try:
                wb = openpyxl.load_workbook(path, read_only=True)
                print(f"  Sheets: {wb.sheetnames}")
                for name in wb.sheetnames:
                    ws = wb[name]
                    # Read first row
                    for row in ws.iter_rows(max_row=1, values_only=True):
                        print(f"    Sheet '{name}' Headers: {list(row)[:15]}")
            except Exception as e:
                print(f"  Error reading {f}: {e}")
        else:
            print(f"\nExcel File: {f} NOT found!")

if __name__ == "__main__":
    main()
